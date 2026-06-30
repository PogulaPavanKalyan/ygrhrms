from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from datetime import date
from .models import *


def hr_login(request):
    return redirect('viewlogins')


@login_required(login_url="hr_login")
def logout_view(request):
    logout(request)
    return redirect("viewlogins")


from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required

from django.views.decorators.cache import never_cache

@login_required(login_url="hr_login")
@never_cache
def home(request):

    if request.user.role != "HR":
        logout(request)
        messages.error(request, "Only HR can access Home Page")
        return redirect("hr_login")


    hr_user = request.user

    manager_count = User.objects.filter(role="Manager").count()
    teamlead_count = User.objects.filter(role="TeamLead").count()
    employee_count = User.objects.filter(role="Employee").count()
    total_users = manager_count + teamlead_count + employee_count

    users = User.objects.filter(
        role__in=["Manager", "TeamLead", "Employee"],
        is_active=True
    )

    today = date.today()
    latest_reports = []

    for user in users:
        report = (
            DailyReport.objects
            .filter(user=user, report_date=today)
            .order_by("-id")
            .first()
        )
        if report:
            latest_reports.append(report)

    holiday_stats = {
        "all": Holiday.objects.count(),
        "pending": Holiday.objects.filter(status='Pending').count(),
        "approved": Holiday.objects.filter(status='Approved').count(),
        "draft": Holiday.objects.filter(status='Draft').count(),
    }

    attendance = Attendance.objects.filter(user=hr_user, date=today).first()

    context = {
        "hr_user": hr_user,
        "total_users": total_users,
        "manager_count": manager_count,
        "teamlead_count": teamlead_count,
        "employee_count": employee_count,
        "latest_reports": latest_reports,
        "today": today,
        "holiday_stats": holiday_stats,
        "attendance": attendance,
    }

    return render(request, "home.html", context)

# ---------------- REGISTRATION HANDLER ----------------
def register_user(request, role, template_name, redirect_url):
    if request.method == "POST":
        username = request.POST.get("fullname")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        gender = request.POST.get("gender")
        date_of_birth = request.POST.get("date_of_birth") or None
        date_of_joining = request.POST.get("date_of_joining") or None
        status = request.POST.get("status")
        address = request.POST.get("address")
        salary = request.POST.get("salary")
        department = request.POST.get("department")
        team_name = request.POST.get("team_name")
        profile_pic = request.FILES.get("profile_pic")
        document = request.FILES.get("document")
        experience_years = request.POST.get("experience_years") or None
        previous_company = request.POST.get("previous_company") or None
        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, template_name)

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return render(request, template_name)

        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
            first_name=username,
            phone=phone,
            gender=gender,
            date_of_birth=date_of_birth,
            date_of_joining=date_of_joining,
            status=status or "Fresher",
            experience_years=experience_years,
            previous_company=previous_company,
            address=address,
            salary=salary,
            department=department,
            team_name=team_name,
            profile_pic=profile_pic,
            document=document,
            role=role,
        )

        messages.success(request, f"{role} registered successfully!")
        return redirect(redirect_url)

    return render(request, template_name)


# ---------------- ROLE REGISTRATIONS ----------------
@login_required
def manager_register(request):
    return register_user(request, "Manager", "manager_register.html", "manager_list")


@login_required
def teamlead_register(request):
    return register_user(request, "TeamLead", "teamlead_register.html", "teamlead_list")


@login_required
def employee_register(request):
    return register_user(request, "Employee", "employee_register.html", "employee_list")





# ---------------- LISTS ----------------
@login_required
def manager_list(request):
    managers = User.objects.filter(role="Manager")
    return render(request, "manager_list.html", {"managers": managers})


@login_required
def teamlead_list(request):
    leads = User.objects.filter(role="TeamLead")
    return render(request, "teamlead_list.html", {"leads": leads})


@login_required
def employee_list(request):
    
    employees = User.objects.filter(role="Employee")
    return render(request, "employee_list.html", {"employees": employees})


# ---------------- PROFILES ----------------
@login_required
def manager_profile(request, pk):
    manager = get_object_or_404(User, pk=pk, role="Manager")

    attendance = calculate_attendance(manager)

    salary = f"₹{manager.salary or 0:,.2f}"

    # Projects assigned to this manager
    projects = Project.objects.filter(
        assigned_manager=manager
    ).select_related("assigned_team")

    # Teams working under this manager (from projects)
    teams = Team.objects.filter(
        projects__assigned_manager=manager
    ).distinct().prefetch_related("members", "lead")

    return render(request, "manager_profile.html", {
        "manager": manager,
        "attendance": attendance,
        "salary": salary,
        "projects": projects,
        "teams": teams,
        "reporting_manager": manager.reporting_manager,
    })


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from hr.models import ProjectWorkUpdate, Project, Team
from django.contrib.auth import get_user_model

User = get_user_model()

@login_required
def teamlead_profile(request, pk):
    teamlead = get_object_or_404(User, pk=pk, role="TeamLead")

    salary = f"₹{teamlead.salary or 0:,.2f}"

    team = Team.objects.filter(lead=teamlead).first()

    attendance = calculate_attendance(teamlead)

    members = team.members.all() if team else User.objects.none()

    projects = Project.objects.filter(assigned_team=team) if team else Project.objects.none()

    project_data = []

    for project in projects:
        member_updates = []
        for member in members:
            # Use ProjectWorkUpdate instead of WorkUpdate
            latest_update = ProjectWorkUpdate.objects.filter(
                project=project,
                team_lead=teamlead  # assuming team lead submitted update
            ).order_by('-created_at').first()
            
            member_updates.append({
                "member": member,
                "has_update": bool(latest_update),
                "latest_update": latest_update,
            })

        project_data.append({
            "project": project,
            "members": member_updates,
        })

    return render(request, "teamlead_profile.html", {
        "teamlead": teamlead,
        "salary": salary,
        "attendance": attendance,
        "project_data": project_data,
        "members": members,
        "reporting_manager": teamlead.reporting_manager,
    })



@login_required
def delete_employee(request, pk):
    employee = get_object_or_404(User, pk=pk)
    employee.delete()
    messages.success(request, "Employee deleted successfully.")
    return redirect("employee_list")



@login_required
def delete_teamlead(request, pk):
    teamlead = get_object_or_404(User, pk=pk)
    teamlead.delete()
    messages.success(request, "Team Lead deleted successfully.")
    return redirect("teamlead_list")

@login_required
def delete_manager(request, pk):
    manager = get_object_or_404(User, pk=pk)
    manager.delete()
    messages.success(request, "Manager deleted successfully.")
    return redirect("manager_list")



@login_required
def employee_profile(request, pk):
    employee = get_object_or_404(User, pk=pk, role="Employee")

   
    salary = f"₹{employee.salary or 0:,.2f}"

   
    teams = Team.objects.filter(members=employee)

  
    employee_projects = Project.objects.filter(assigned_team__in=teams)
    attendance = calculate_attendance(employee)


    return render(request, "employee_profile.html", {
        "employee": employee,
        "salary": salary,
        "employee_projects": employee_projects,
         "attendance": attendance,
         "reporting_manager": employee.reporting_manager,
    })





@login_required
def edit_profile(request, pk):
    """Allow HR to edit any user's profile."""
    user_obj = get_object_or_404(User, pk=pk)

    # Restrict access to HR only
    if request.user.role != "HR":
        messages.error(request, "Only HR can edit profiles.")
        return redirect("home")

    if request.method == "POST":
        user_obj.first_name = request.POST.get("first_name")
        user_obj.email = request.POST.get("email")
        user_obj.phone = request.POST.get("phone")
        user_obj.address = request.POST.get("address")
        user_obj.department = request.POST.get("department")
        user_obj.team_name = request.POST.get("team_name")
        user_obj.gender = request.POST.get("gender")
        user_obj.salary = request.POST.get("salary") or user_obj.salary
        user_obj.date_of_birth = request.POST.get("date_of_birth") or user_obj.date_of_birth

        if request.FILES.get("profile_pic"):
            user_obj.profile_pic = request.FILES.get("profile_pic")

        user_obj.save()
        messages.success(request, f"{user_obj.first_name}'s profile updated successfully.")

        # Redirect to the right list based on role
        if user_obj.role == "Manager":
            return redirect("manager_list")
        elif user_obj.role == "TeamLead":
            return redirect("teamlead_list")
        else:
            return redirect("employee_list")

    return render(request, "edit_profile.html", {"user": user_obj})





@login_required
def assign_project(request):
    managers = User.objects.filter(role="Manager")

    if request.method == "POST":
        name = request.POST.get("title")
        description = request.POST.get("description")
        assigned_to_id = request.POST.get("assigned_to")
        startdate=request.POST.get("startdate")
        deadline = request.POST.get("deadline")
        document = request.FILES.get("document")  # ✅ FILE HERE

        assigned_to = get_object_or_404(User, id=assigned_to_id)

        last_project = Project.objects.order_by("-id").first()
        next_number = last_project.id + 1 if last_project else 1
        project_id = f"PRJ{next_number:04d}"

        Project.objects.create(
            project_id=project_id,
            name=name,
            description=description,
            assigned_manager=assigned_to,
            startdate=startdate,
            deadline=deadline,
            document=document,   # ✅ SAVE FILE
            status="Pending"
        )

        messages.success(request, f"Project {project_id} assigned successfully.")
        return redirect("assign_project")

    projects = Project.objects.filter(
        assigned_manager__role="Manager"
    ).order_by("-created_at")

    return render(request, "assign_project.html", {
        "managers": managers,
        "projects": projects
    })



@login_required
def delete_project(request, pk):
    project = get_object_or_404(Project, pk=pk)
    project.delete()
    messages.success(request, "Project deleted successfully.")
    return redirect("assign_project")

# views.py
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.auth import get_user_model
import json
from .models import ChatMessage
from django.utils import timezone
from django.db.models import Count
from django.utils.timezone import localtime
from datetime import timedelta
from django.db.models import Max
from django.db.models.functions import Coalesce


User = get_user_model()
@login_required
def communication_page(request, user_id=None):
    from .models import ChatRoom, GroupMessage, ChatMessage, UserPresence, CompanyAnnouncement, Attendance, Team
    
    # Self-seed department channels
    standard_channels = [
        {"name": "General", "desc": "Company-wide general discussions", "ann": False},
        {"name": "HR", "desc": "HR announcements and queries", "ann": False},
        {"name": "Development", "desc": "Software development discussions", "ann": False},
        {"name": "QA", "desc": "Quality assurance and testing discussions", "ann": False},
        {"name": "Announcements", "desc": "Official company announcements (Read-only for employees)", "ann": True},
        {"name": "Payroll", "desc": "Payroll and salary discussions", "ann": False},
        {"name": "Events", "desc": "Upcoming company events and gatherings", "ann": False},
        {"name": "Support", "desc": "IT support and troubleshooting help", "ann": False}
    ]

    for ch in standard_channels:
        room, created = ChatRoom.objects.get_or_create(
            name=ch["name"],
            room_type='channel',
            defaults={
                'description': ch["desc"],
                'is_announcement_only': ch["ann"]
            }
        )
        if created:
            room.users.set(User.objects.all())

    # Get active/online status counts
    now = timezone.now()
    all_presences = {p.user_id: p for p in UserPresence.objects.all()}

    def get_status(u):
        p = all_presences.get(u.id)
        if not p:
            return 'Offline'
        if now - p.last_activity > timedelta(minutes=5):
            return 'Offline'
        return p.status

    # Get all company users (excluding self)
    all_users = User.objects.exclude(id=request.user.id).order_by('username')
    users_data = []
    for u in all_users:
        unread = ChatMessage.objects.filter(
            sender=u,
            receiver=request.user,
            is_read=False
        ).exclude(deleted_for=request.user).count()

        last_msg = ChatMessage.objects.filter(
            (Q(sender=request.user, receiver=u) | Q(sender=u, receiver=request.user))
        ).order_by('-id').first()

        last_time = last_msg.created_at if last_msg else timezone.make_aware(timezone.datetime(2000, 1, 1))

        users_data.append({
            "user": u,
            "status": get_status(u),
            "unread": unread,
            "last_time": last_time,
            "last_msg_text": last_msg.text if last_msg else ""
        })

    # Sort users by last message time (most recent first)
    users_data.sort(key=lambda x: x["last_time"], reverse=True)

    # Get channels (room_type = 'channel')
    channels = ChatRoom.objects.filter(room_type='channel')

    # Get teams (room_type = 'team') that the user is a member of
    teams = ChatRoom.objects.filter(room_type='team', users=request.user)

    # Announcements for MD/HR dashboard or communication page
    announcements = CompanyAnnouncement.objects.all().order_by('-created_at')[:10]

    # Shared files directory
    shared_files = []
    
    # DM files
    dm_files = ChatMessage.objects.filter(
        (Q(sender=request.user) | Q(receiver=request.user)),
        file__isnull=False
    ).exclude(deleted_for=request.user).exclude(deleted_for_everyone=True).order_by('-created_at')[:10]
    
    for f in dm_files:
        if f.file and f.file.name:
            shared_files.append({
                "name": f.file.name.split('/')[-1],
                "url": f.file.url,
                "sender": f.sender.get_full_name() or f.sender.username,
                "timestamp": localtime(f.created_at).strftime("%d %b %H:%M")
            })

    # Group files
    group_rooms = ChatRoom.objects.filter(users=request.user)
    gm_files = GroupMessage.objects.filter(
        room__in=group_rooms,
        file__isnull=False,
        is_deleted=False
    ).order_by('-created_at')[:10]

    for f in gm_files:
        if f.file and f.file.name:
            shared_files.append({
                "name": f.file.name.split('/')[-1],
                "url": f.file.url,
                "sender": f.sender.get_full_name() or f.sender.username,
                "timestamp": localtime(f.created_at).strftime("%d %b %H:%M")
            })

    role_to_base = {
        "MD": ("md/md_dashboard.html", "md_dashboard"),
        "HR": ("base.html", "home"),
        "Manager": ("main/mddash.html", "index"),
        "TeamLead": ("tl/side_bar.html", "tl_dashboard"),
        "Employee": ("employee/employee_base.html", "employee_dashboard"),
    }
    base_template, back_url = role_to_base.get(
        request.user.role,
        ("employee/employee_base.html", "employee_dashboard")
    )

    active_room = None
    active_private_user = None

    room_id = request.GET.get('room_id')
    if room_id:
        try:
            active_room = ChatRoom.objects.get(id=room_id)
        except ChatRoom.DoesNotExist:
            pass

    if user_id:
        try:
            active_private_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            pass

    # Fetch call history for display in the sidebar / UI
    from .models import CallSession
    raw_calls = CallSession.objects.filter(
        Q(caller=request.user) | Q(receiver=request.user)
    ).order_by('-created_at')[:20]

    call_logs = []
    for c in raw_calls:
        is_caller = (c.caller == request.user)
        peer = c.receiver if is_caller else c.caller
        duration_str = ""
        if c.started_at and c.ended_at:
            duration = (c.ended_at - c.started_at).total_seconds()
            m = int(duration // 60)
            s = int(duration % 60)
            duration_str = f"{m}m {s}s"
        else:
            duration_str = "Missed" if c.status == 'missed' else ("Rejected" if c.status == 'rejected' else "No answer")

        call_logs.append({
            "id": c.id,
            "call_type": c.call_type,
            "direction": "Outgoing" if is_caller else "Incoming",
            "peer_name": (peer.get_full_name() or peer.username) if peer else "Group Meeting",
            "status": c.status,
            "timestamp": localtime(c.created_at).strftime("%d %b %H:%M"),
            "duration": duration_str
        })

    return render(request, "unified_chat.html", {
        "users_data": users_data,
        "channels": channels,
        "teams": teams,
        "announcements": announcements,
        "shared_files": shared_files,
        "base_template": base_template,
        "back_url": back_url,
        "active_room": active_room,
        "active_private_user": active_private_user,
        "all_users_plain": User.objects.exclude(id=request.user.id),
        "call_logs": call_logs,
    })


@login_required
def send_message_ajax(request):
    if request.method == "POST" and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        text = request.POST.get("text")
        receiver_id = request.POST.get("receiver_id")
        if text and receiver_id:
            msg = ChatMessage.objects.create(
                sender=request.user,
                receiver_id=receiver_id,
                text=text
            )
            return JsonResponse({
                "status": "success",
                "message": msg.text,
                "created_at": msg.created_at.strftime("%H:%M")
            })
    return JsonResponse({"status": "error"})


@csrf_exempt
@login_required
def send_file_ajax(request):
    if request.method == "POST":
        receiver_id = request.POST.get("receiver_id")
        file = request.FILES.get("file")
        if not receiver_id or not file:
            return JsonResponse({"status": "error", "msg": "Missing data"}, status=400)
        try:
            receiver = User.objects.get(id=receiver_id)
        except User.DoesNotExist:
            return JsonResponse({"status": "error", "msg": "User not found"}, status=404)

        ChatMessage.objects.create(
            sender=request.user,
            receiver=receiver,
            text="",
            file=file
        )
        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "error"}, status=400)


@csrf_exempt
@login_required
def typing_status(request):
    receiver_id = request.GET.get("receiver")
    if receiver_id:
        request.session[f"typing_{receiver_id}"] = True
    return JsonResponse({"status": "ok"})


@require_POST
@login_required
def edit_message(request, msg_id):
    msg = ChatMessage.objects.get(id=msg_id, sender=request.user)
    data = json.loads(request.body)
    msg.text = data['text']
    msg.save()
    return JsonResponse({'status': 'ok'})

@require_POST
@login_required
def delete_message(request, msg_id):
    try:
        msg = ChatMessage.objects.get(id=msg_id)

        if request.user != msg.sender and request.user != msg.receiver:
            return JsonResponse({"status": "error"}, status=403)

        msg.deleted_for.add(request.user)

        return JsonResponse({"status": "success"})
    except ChatMessage.DoesNotExist:
        return JsonResponse({"status": "error"}, status=404)


# views.py
@require_POST
@login_required
def delete_message_everyone(request, msg_id):
    try:
        msg = ChatMessage.objects.get(id=msg_id)

        if msg.sender != request.user:
            return JsonResponse({"status": "error"}, status=403)

        msg.deleted_for_everyone = True
        msg.deleted_by_user = request.user
        msg.text = ""
        msg.file = None
        msg.save()

        return JsonResponse({"status": "success"})
    except ChatMessage.DoesNotExist:
        return JsonResponse({"status": "error"}, status=404)






@csrf_exempt
@login_required
def forward_message_ajax(request):
    if request.method == "POST":
        data = json.loads(request.body)
        msg_ids = data.get("msg_ids", [])
        usernames = data.get("users", [])
        for username in usernames:
            try:
                receiver = User.objects.get(username=username.strip())
            except User.DoesNotExist:
                continue
            for msg_id in msg_ids:
                try:
                    msg = ChatMessage.objects.get(id=msg_id)
                    ChatMessage.objects.create(
                        sender=request.user,
                        receiver=receiver,
                        text=msg.text,
                        file=msg.file if msg.file else None
                    )
                except ChatMessage.DoesNotExist:
                    continue
        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "error"}, status=400)

from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
import json
from .models import ChatMessage

@require_POST
@login_required
def delete_multiple_messages(request):
    data = json.loads(request.body)
    msg_ids = data.get("msg_ids", [])
    if not msg_ids:
        return JsonResponse({"status": "error", "msg": "No messages selected"}, status=400)

    for msg_id in msg_ids:
        try:
            msg = ChatMessage.objects.get(id=msg_id)
            # Only hide for current user
            msg.deleted_for.add(request.user)
            msg.save()
        except ChatMessage.DoesNotExist:
            continue

    return JsonResponse({"status": "success"})

from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.shortcuts import render
from .models import ChatRoom, GroupMessage
from datetime import timedelta
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.shortcuts import render

@login_required
def group_chat(request):
    request.session["last_chat_seen"] = timezone.now().isoformat()

    user = request.user
    User = get_user_model()

    room, _ = ChatRoom.objects.get_or_create(name="All Users")
    room.users.set(User.objects.all())

    messages = GroupMessage.objects.filter(
        room=room
    ).order_by("created_at")

    # ✅ DATE LOGIC
    today = timezone.localdate()
    yesterday = today - timedelta(days=1)

    # 🔙 BACK BUTTON LOGIC (ROLE WISE)
    role_to_base = {
        "MD": "md_dashboard",
        "HR": "home",
        "Manager": "index",
        "TeamLead": "tl_dashboard",
        "Employee": "employee_dashboard",
    }

    back_url = role_to_base.get(user.role, "home")

    # Enrich context for enterprise chat dashboard
    all_users = User.objects.filter(is_active=True).order_by('first_name', 'username')
    team_groups = Team.objects.filter(is_active=True)
    
    # Identify online users from active check-ins today
    online_user_ids = list(Attendance.objects.filter(
        date=today,
        check_in_time__isnull=False,
        check_out_time__isnull=True
    ).values_list('user_id', flat=True))

    role_counts = {
        "MD": User.objects.filter(role="MD", is_active=True).count(),
        "Manager": User.objects.filter(role="Manager", is_active=True).count(),
        "TeamLead": User.objects.filter(role="TeamLead", is_active=True).count(),
        "Employee": User.objects.filter(role="Employee", is_active=True).count(),
        "HR": User.objects.filter(role="HR", is_active=True).count(),
    }

    return render(request, "group_chat.html", {
        "messages": messages,
        "back_url": back_url,
        "today": today,          
        "yesterday": yesterday,
        "all_users": all_users,
        "team_groups": team_groups,
        "online_user_ids": online_user_ids,
        "role_counts": role_counts,
    })



@login_required
@require_POST
def send_group_message(request):
    room = ChatRoom.objects.get(name="All Users")
    text = request.POST.get("text")

    GroupMessage.objects.create(
        room=room,
        sender=request.user,
        text=text
    )
    return JsonResponse({"status": "ok"})


@login_required
@require_POST
def delete_group_message(request, msg_id):
    msg = get_object_or_404(GroupMessage, id=msg_id, sender=request.user)

    msg.is_deleted = True
    msg.deleted_by = request.user
    msg.text = ""     # remove actual text
    msg.save()

    return JsonResponse({"status": "deleted"})

@login_required
def edit_group_message(request, id):
    if request.method == "POST":
        msg = get_object_or_404(GroupMessage, id=id)

        if msg.sender != request.user:
            return JsonResponse({"error":"Not allowed"}, status=403)

        new_text = request.POST.get("text")
        if new_text:
            msg.text = new_text
            msg.save()

        return JsonResponse({"success":True})

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone

@login_required
def chat_notification_count(request):
    user = request.user

    # LAST SEEN TIME (from session)
    last_seen = request.session.get("last_chat_seen")

    # -------- PERSONAL CHAT --------
    personal_qs = ChatMessage.objects.filter(
        receiver=user,
        deleted_for_everyone=False
    )

    if last_seen:
        personal_qs = personal_qs.filter(created_at__gt=last_seen)

    personal_unread = personal_qs.count()

    # -------- GROUP CHAT --------
    group_qs = GroupMessage.objects.exclude(
        sender=user
    ).filter(
        is_deleted=False
    )

    if last_seen:
        group_qs = group_qs.filter(created_at__gt=last_seen)

    group_unread = group_qs.count()

    return JsonResponse({
        "personal": personal_unread,
        "group": group_unread,
        "total": personal_unread + group_unread
    })


from django.http import JsonResponse
from django.utils.dateparse import parse_datetime
from django.utils import timezone

@login_required
def group_chat_unread_api(request):
    last_seen_str = request.session.get("last_chat_seen")

    if last_seen_str:
        last_seen = parse_datetime(last_seen_str)
    else:
        last_seen = timezone.make_aware(timezone.datetime(2000, 1, 1))

    room = ChatRoom.objects.filter(name="All Users").first()

    count = 0
    if room:
        count = GroupMessage.objects.filter(
            room=room,
            created_at__gt=last_seen
        ).exclude(sender=request.user).count()

    return JsonResponse({"count": count})

@login_required
def private_chat_unread_api(request):
    last_seen = request.session.get("last_private_chat_seen")

    if last_seen:
        last_seen = parse_datetime(last_seen)
    else:
        last_seen = timezone.make_aware(timezone.datetime(2000, 1, 1))

    count = ChatMessage.objects.filter(
        receiver=request.user,
        created_at__gt=last_seen,
        deleted_for_everyone=False
    ).exclude(deleted_for=request.user).count()

    return JsonResponse({"count": count})


# ==================== UNIFIED COLLABORATION AND CHAT VIEW APIS ====================

import os
import json
from django.conf import settings
from django.db.models import Q
from django.utils import timezone
from django.utils.timezone import localtime
from datetime import timedelta
from .models import ChatRoom, GroupMessage, ChatMessage, UserPresence, CompanyAnnouncement

TYPING_FILE_PATH = os.path.join(settings.BASE_DIR, 'tmp', 'typing_status.json')

def update_typing_indicator(user_id, target_type, target_id, is_typing):
    try:
        os.makedirs(os.path.dirname(TYPING_FILE_PATH), exist_ok=True)
        if os.path.exists(TYPING_FILE_PATH):
            with open(TYPING_FILE_PATH, 'r') as f:
                data = json.load(f)
        else:
            data = {}
    except Exception:
        data = {}

    user_key = str(user_id)
    if is_typing:
        data[user_key] = {
            "target_type": target_type,
            "target_id": int(target_id),
            "timestamp": timezone.now().isoformat()
        }
    else:
        if user_key in data:
            del data[user_key]

    try:
        with open(TYPING_FILE_PATH, 'w') as f:
            json.dump(data, f)
    except Exception:
        pass

def get_typing_users(target_type, target_id, current_user_id):
    try:
        if not os.path.exists(TYPING_FILE_PATH):
            return []
        with open(TYPING_FILE_PATH, 'r') as f:
            data = json.load(f)
    except Exception:
        return []

    typing_names = []
    now = timezone.now()
    for uid_str, info in list(data.items()):
        uid = int(uid_str)
        if uid == current_user_id:
            continue
        try:
            ts = timezone.datetime.fromisoformat(info["timestamp"])
            if now - ts > timedelta(seconds=6):
                continue
            if info["target_type"] == target_type and info["target_id"] == int(target_id):
                u = User.objects.get(id=uid)
                typing_names.append(u.get_full_name() or u.username)
        except Exception:
            continue
    return typing_names


@login_required
def get_chat_history(request):
    room_id = request.GET.get('room_id')
    user_id = request.GET.get('user_id')
    last_checked_id = request.GET.get('last_checked_id')
    
    try:
        last_checked_id = int(last_checked_id)
    except (TypeError, ValueError):
        last_checked_id = 0

    messages_data = []
    typing_users = []

    if room_id:
        room = get_object_or_404(ChatRoom, id=room_id)
        if room.room_type == 'team' and not room.users.filter(id=request.user.id).exists() and room.created_by != request.user:
            return JsonResponse({"error": "You do not have access to this team chat."}, status=403)
        
        messages = GroupMessage.objects.filter(room=room)
        if last_checked_id:
            messages = messages.filter(id__gt=last_checked_id)
        messages = messages.order_by('id')

        request.session["last_chat_seen"] = timezone.now().isoformat()
        typing_users = get_typing_users('room', room_id, request.user.id)

        for msg in messages:
            reactions_dict = msg.reactions or {}
            reactions_with_usernames = {}
            for emoji, uids in reactions_dict.items():
                if isinstance(uids, list):
                    names = list(User.objects.filter(id__in=uids).values_list('username', flat=True))
                    reactions_with_usernames[emoji] = {
                        "users": uids,
                        "usernames": names
                    }
                else:
                    reactions_with_usernames[emoji] = {"users": [], "usernames": []}

            reply_info = None
            if msg.reply_to:
                reply_info = {
                    "id": msg.reply_to.id,
                    "sender_name": msg.reply_to.sender.get_full_name() or msg.reply_to.sender.username,
                    "text_preview": msg.reply_to.text[:50] if msg.reply_to.text else ("Attachment" if msg.reply_to.file else "Message")
                }

            messages_data.append({
                "id": msg.id,
                "is_group": True,
                "sender_id": msg.sender.id,
                "sender_name": msg.sender.get_full_name() or msg.sender.username,
                "sender_avatar": msg.sender.profile_pic.url if msg.sender.profile_pic else None,
                "sender_role": msg.sender.role,
                "text": msg.text,
                "file_url": msg.file.url if (msg.file and msg.file.name) else None,
                "file_name": msg.file.name.split('/')[-1] if (msg.file and msg.file.name) else None,
                "created_at": localtime(msg.created_at).strftime("%d %b %H:%M"),
                "edited": msg.edited,
                "reactions": reactions_with_usernames,
                "is_deleted": msg.is_deleted,
                "reply_to": reply_info
            })

    elif user_id:
        target_user = get_object_or_404(User, id=user_id)
        
        # Mark received messages from target_user as read
        ChatMessage.objects.filter(
            sender=target_user,
            receiver=request.user,
            is_read=False
        ).update(is_read=True, is_delivered=True)

        messages = ChatMessage.objects.filter(
            (Q(sender=request.user, receiver=target_user) | Q(sender=target_user, receiver=request.user))
        )
        if last_checked_id:
            messages = messages.filter(id__gt=last_checked_id)
        
        messages = messages.order_by('id')

        request.session["last_private_chat_seen"] = timezone.now().isoformat()
        typing_users = get_typing_users('private', user_id, request.user.id)

        for msg in messages:
            if request.user in msg.deleted_for.all():
                continue
            
            reactions_dict = msg.reactions or {}
            reactions_with_usernames = {}
            for emoji, uids in reactions_dict.items():
                if isinstance(uids, list):
                    names = list(User.objects.filter(id__in=uids).values_list('username', flat=True))
                    reactions_with_usernames[emoji] = {
                        "users": uids,
                        "usernames": names
                    }
                else:
                    reactions_with_usernames[emoji] = {"users": [], "usernames": []}

            reply_info = None
            if msg.reply_to:
                reply_info = {
                    "id": msg.reply_to.id,
                    "sender_name": msg.reply_to.sender.get_full_name() or msg.reply_to.sender.username,
                    "text_preview": msg.reply_to.text[:50] if msg.reply_to.text else ("Attachment" if msg.reply_to.file else "Message")
                }

            messages_data.append({
                "id": msg.id,
                "is_group": False,
                "sender_id": msg.sender.id,
                "sender_name": msg.sender.get_full_name() or msg.sender.username,
                "sender_avatar": msg.sender.profile_pic.url if msg.sender.profile_pic else None,
                "sender_role": msg.sender.role,
                "text": msg.text,
                "file_url": msg.file.url if (msg.file and msg.file.name) else None,
                "file_name": msg.file.name.split('/')[-1] if (msg.file and msg.file.name) else None,
                "created_at": localtime(msg.created_at).strftime("%d %b %H:%M"),
                "edited": msg.edited,
                "reactions": reactions_with_usernames,
                "is_deleted": msg.deleted_for_everyone,
                "reply_to": reply_info,
                "is_read": msg.is_read,
                "is_delivered": msg.is_delivered
            })

    return JsonResponse({"status": "success", "messages": messages_data, "typing_users": typing_users})


@csrf_exempt
@login_required
def send_chat_message(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    room_id = request.POST.get('room_id')
    receiver_id = request.POST.get('receiver_id')
    text = request.POST.get('text', '')
    reply_to_id = request.POST.get('reply_to_id')
    file = request.FILES.get('file')

    voice_duration = request.POST.get('voice_duration')
    voice_data = request.POST.get('voice_data')
    if voice_data and voice_duration:
        text = f"[VOICE_MESSAGE]:{voice_duration}:{voice_data}"

    reply_to = None
    if reply_to_id:
        try:
            if room_id:
                reply_to = GroupMessage.objects.get(id=reply_to_id)
            else:
                reply_to = ChatMessage.objects.get(id=reply_to_id)
        except Exception:
            pass

    if room_id:
        room = get_object_or_404(ChatRoom, id=room_id)
        if room.is_announcement_only and request.user.role not in ['MD', 'HR']:
            return JsonResponse({"error": "Only MD and HR can post in the announcements channel."}, status=403)

        msg = GroupMessage.objects.create(
            room=room,
            sender=request.user,
            text=text,
            file=file,
            reply_to=reply_to
        )
        # Broadcast via WebSocket
        broadcast_ws_message(
            f"room_{room.id}",
            {
                "type": "room_message",
                "message": {
                    "id": msg.id,
                    "sender_id": msg.sender.id,
                    "sender_name": msg.sender.get_full_name() or msg.sender.username,
                    "sender_avatar": msg.sender.profile_pic.url if msg.sender.profile_pic else None,
                    "text": msg.text,
                    "created_at": localtime(msg.created_at).strftime("%I:%M %p"),
                    "room_id": room.id
                }
            }
        )
        return JsonResponse({
            "status": "success",
            "message": {
                "id": msg.id,
                "is_group": True,
                "sender_id": msg.sender.id,
                "sender_name": msg.sender.get_full_name() or msg.sender.username,
                "text": msg.text,
                "file_url": msg.file.url if (msg.file and msg.file.name) else None,
                "file_name": msg.file.name.split('/')[-1] if (msg.file and msg.file.name) else None,
                "created_at": localtime(msg.created_at).strftime("%d %b %H:%M"),
                "edited": msg.edited,
                "reactions": msg.reactions,
                "is_deleted": msg.is_deleted
            }
        })
    elif receiver_id:
        receiver = get_object_or_404(User, id=receiver_id)
        msg = ChatMessage.objects.create(
            sender=request.user,
            receiver=receiver,
            text=text,
            file=file,
            reply_to=reply_to
        )
        
        # Broadcast via WebSocket
        broadcast_ws_message(
            f"user_{receiver.id}",
            {
                "type": "chat_message",
                "message": {
                    "id": msg.id,
                    "sender_id": msg.sender.id,
                    "sender_name": msg.sender.get_full_name() or msg.sender.username,
                    "sender_avatar": msg.sender.profile_pic.url if msg.sender.profile_pic else None,
                    "text": msg.text,
                    "created_at": localtime(msg.created_at).strftime("%I:%M %p"),
                    "room_id": None
                }
            }
        )
        
        return JsonResponse({
            "status": "success",
            "message": {
                "id": msg.id,
                "is_group": False,
                "sender_id": msg.sender.id,
                "sender_name": msg.sender.get_full_name() or msg.sender.username,
                "text": msg.text,
                "file_url": msg.file.url if (msg.file and msg.file.name) else None,
                "file_name": msg.file.name.split('/')[-1] if (msg.file and msg.file.name) else None,
                "created_at": localtime(msg.created_at).strftime("%d %b %H:%M"),
                "edited": msg.edited,
                "reactions": msg.reactions,
                "is_deleted": msg.deleted_for_everyone,
                "is_read": msg.is_read,
                "is_delivered": msg.is_delivered
            }
        })

    return JsonResponse({"error": "Invalid request parameters"}, status=400)


@csrf_exempt
@login_required
def toggle_reaction(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = request.POST

    message_id = data.get('message_id')
    is_group = str(data.get('is_group')).lower() == 'true'
    emoji = data.get('emoji')

    if not message_id or not emoji:
        return JsonResponse({"error": "Missing parameters"}, status=400)

    if is_group:
        msg = get_object_or_404(GroupMessage, id=message_id)
    else:
        msg = get_object_or_404(ChatMessage, id=message_id)

    reactions = msg.reactions or {}
    user_id = request.user.id

    if emoji in reactions:
        uids = reactions[emoji]
        if not isinstance(uids, list):
            uids = []
        if user_id in uids:
            uids.remove(user_id)
            if not uids:
                del reactions[emoji]
            else:
                reactions[emoji] = uids
        else:
            uids.append(user_id)
            reactions[emoji] = uids
    else:
        reactions[emoji] = [user_id]

    msg.reactions = reactions
    msg.save()

    reactions_with_usernames = {}
    for emo, uids in reactions.items():
        if isinstance(uids, list):
            names = list(User.objects.filter(id__in=uids).values_list('username', flat=True))
            reactions_with_usernames[emo] = {
                "users": uids,
                "usernames": names
            }

    return JsonResponse({"status": "success", "reactions": reactions_with_usernames})


@csrf_exempt
@login_required
def edit_chat_message(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = request.POST

    message_id = data.get('message_id')
    is_group = str(data.get('is_group')).lower() == 'true'
    new_text = data.get('text', '').strip()

    if not message_id or not new_text:
        return JsonResponse({"error": "Missing parameters"}, status=400)

    if is_group:
        msg = get_object_or_404(GroupMessage, id=message_id)
    else:
        msg = get_object_or_404(ChatMessage, id=message_id)

    if msg.sender != request.user:
        return JsonResponse({"error": "Forbidden: You are not the sender of this message."}, status=403)

    time_diff = timezone.now() - msg.created_at
    if time_diff.total_seconds() > 600:
        return JsonResponse({"error": "Cannot edit message: 10-minute limit exceeded."}, status=400)

    msg.text = new_text
    msg.edited = True
    msg.save()

    return JsonResponse({"status": "success", "text": msg.text})


@csrf_exempt
@login_required
def delete_chat_message(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = request.POST

    message_id = data.get('message_id')
    is_group = str(data.get('is_group')).lower() == 'true'
    mode = data.get('mode', 'everyone')

    if not message_id:
        return JsonResponse({"error": "Missing parameters"}, status=400)

    if is_group:
        msg = get_object_or_404(GroupMessage, id=message_id)
        if msg.sender != request.user:
            return JsonResponse({"error": "Forbidden: You are not the sender of this message."}, status=403)
        msg.is_deleted = True
        msg.deleted_by = request.user
        msg.text = ""
        msg.file = None
        msg.save()
    else:
        msg = get_object_or_404(ChatMessage, id=message_id)
        if mode == 'everyone':
            if msg.sender != request.user:
                return JsonResponse({"error": "Forbidden: You are not the sender of this message."}, status=403)
            msg.deleted_for_everyone = True
            msg.deleted_by_user = request.user
            msg.text = ""
            msg.file = None
            msg.save()
        else:
            msg.deleted_for.add(request.user)
            msg.save()

    return JsonResponse({"status": "success"})


@csrf_exempt
@login_required
def update_presence(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = request.POST

    status = data.get('status', 'Online')
    valid_statuses = ['Online', 'Offline', 'Away', 'Busy', 'In Meeting', 'Working From Home']
    if status not in valid_statuses:
        status = 'Online'

    presence, created = UserPresence.objects.get_or_create(user=request.user)
    presence.status = status
    presence.save()

    typing_room_id = data.get('typing_room_id')
    typing_receiver_id = data.get('typing_receiver_id')
    is_typing = str(data.get('is_typing')).lower() == 'true'

    if typing_room_id:
        update_typing_indicator(request.user.id, 'room', typing_room_id, is_typing)
    elif typing_receiver_id:
        update_typing_indicator(request.user.id, 'private', typing_receiver_id, is_typing)

    return JsonResponse({"status": "success"})


@csrf_exempt
@login_required
def create_team_group(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    if request.user.role not in ['MD', 'HR', 'Manager', 'TeamLead']:
        return JsonResponse({"error": "Forbidden: You are not authorized to create team groups."}, status=403)

    name = request.POST.get('name', '').strip()
    description = request.POST.get('description', '').strip()
    user_ids = request.POST.getlist('users')

    if not name:
        return JsonResponse({"error": "Room name is required."}, status=400)

    room = ChatRoom.objects.create(
        name=name,
        room_type='team',
        description=description,
        created_by=request.user
    )
    room.users.add(request.user)
    room.admins.add(request.user)

    for uid in user_ids:
        try:
            u = User.objects.get(id=int(uid))
            room.users.add(u)
        except Exception:
            continue

    return JsonResponse({
        "status": "success",
        "room": {
            "id": room.id,
            "name": room.name,
            "description": room.description
        }
    })


@csrf_exempt
@login_required
def post_announcement(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    if request.user.role not in ['MD', 'HR']:
        return JsonResponse({"error": "Forbidden: Only MD and HR can post announcements."}, status=403)

    title = request.POST.get('title', '').strip()
    content = request.POST.get('content', '').strip()
    ann_type = request.POST.get('type', 'General')

    if not title or not content:
        return JsonResponse({"error": "Title and content are required."}, status=400)

    ann = CompanyAnnouncement.objects.create(
        title=title,
        content=content,
        announcement_type=ann_type,
        created_by=request.user
    )

    ann_room = ChatRoom.objects.filter(room_type='channel', name='Announcements').first()
    if not ann_room:
        ann_room = ChatRoom.objects.create(
            name='Announcements',
            room_type='channel',
            description='Official company announcements (Read-only for employees)',
            is_announcement_only=True
        )

    GroupMessage.objects.create(
        room=ann_room,
        sender=request.user,
        text=f"📢 **{title}**\n\n{content}"
    )

    return JsonResponse({
        "status": "success",
        "announcement": {
            "id": ann.id,
            "title": ann.title,
            "created_at": localtime(ann.created_at).strftime("%d %b %Y %H:%M")
        }
    })


def broadcast_ws_message(group_name, payload):
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(group_name, payload)
    except Exception as e:
        print("WebSocket broadcast error:", e)


@csrf_exempt
@login_required
def call_signal(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = request.POST

    action = data.get('action')
    if not action:
        return JsonResponse({"error": "Missing action"}, status=400)

    from .models import CallSession

    if action == 'initiate':
        receiver_id = data.get('receiver_id')
        room_id = data.get('room_id')
        call_type = data.get('call_type', 'video')
        sdp = data.get('sdp')

        if receiver_id:
            receiver = get_object_or_404(User, id=receiver_id)
            if request.user.role == 'Employee' and receiver.role == 'MD':
                return JsonResponse({"error": "Forbidden: Employees are not allowed to call MD directly."}, status=403)
            
            # Check if the receiver is online (via presence)
            from .models import UserPresence
            presence = UserPresence.objects.filter(user=receiver).first()
            is_online = False
            if presence:
                # Active in last 5 minutes and not Offline
                if timezone.now() - presence.last_activity <= timezone.timedelta(minutes=5) and presence.status != 'Offline':
                    is_online = True
            else:
                # Default to online if presence tracking is not active/seeded (safety fallback)
                is_online = True
                
            if not is_online:
                return JsonResponse({"status": "offline", "error": "Employee is offline."})

            # Check if the receiver is busy in another call
            active_receiver_call = CallSession.objects.filter(
                (Q(caller=receiver) | Q(receiver=receiver)),
                status__in=['ringing', 'active']
            ).exists()
            if active_receiver_call:
                return JsonResponse({"status": "busy", "error": "User is currently on another call."})
        else:
            receiver = None

        # Clean up any existing active calls of the caller
        CallSession.objects.filter(caller=request.user, status__in=['ringing', 'active']).update(status='ended', ended_at=timezone.now())

        session = CallSession.objects.create(
            caller=request.user,
            receiver=receiver,
            room_id=room_id,
            call_type=call_type,
            status='ringing',
            caller_sdp=sdp
        )

        # Broadcast incoming call to receiver WebSocket group
        if receiver:
            broadcast_ws_message(
                f"user_{receiver.id}",
                {
                    "type": "call_incoming",
                    "session_id": session.id,
                    "caller_id": request.user.id,
                    "caller_name": request.user.get_full_name() or request.user.username,
                    "caller_avatar": request.user.profile_pic.url if request.user.profile_pic else None,
                    "caller_role": request.user.designation or request.user.role or "Staff",
                    "caller_dept": request.user.get_department_display() if hasattr(request.user, 'get_department_display') else "Development",
                    "call_type": call_type
                }
            )

        return JsonResponse({"status": "success", "session_id": session.id})

    elif action == 'accept':
        session_id = data.get('session_id')
        sdp = data.get('sdp')
        session = get_object_or_404(CallSession, id=session_id)
        session.status = 'active'
        session.receiver_sdp = sdp
        session.started_at = timezone.now()
        session.save()

        # Broadcast call accept to caller group
        broadcast_ws_message(
            f"user_{session.caller.id}",
            {
                "type": "call_state_change",
                "session_id": session.id,
                "status": "active",
                "receiver_sdp": sdp
            }
        )
        return JsonResponse({"status": "success"})

    elif action == 'reject':
        session_id = data.get('session_id')
        session = get_object_or_404(CallSession, id=session_id)
        session.status = 'rejected'
        session.ended_at = timezone.now()
        session.save()

        # Broadcast call reject to caller group
        broadcast_ws_message(
            f"user_{session.caller.id}",
            {
                "type": "call_state_change",
                "session_id": session.id,
                "status": "rejected"
            }
        )
        return JsonResponse({"status": "success"})

    elif action == 'missed':
        session_id = data.get('session_id')
        session = get_object_or_404(CallSession, id=session_id)
        session.status = 'missed'
        session.ended_at = timezone.now()
        session.save()

        # Broadcast call missed to receiver group
        broadcast_ws_message(
            f"user_{session.receiver.id}",
            {
                "type": "call_state_change",
                "session_id": session.id,
                "status": "missed"
            }
        )
        return JsonResponse({"status": "success"})

    elif action == 'end':
        session_id = data.get('session_id')
        session = get_object_or_404(CallSession, id=session_id)
        session.status = 'ended'
        session.ended_at = timezone.now()
        session.save()

        # Broadcast call ended to peer group
        peer_id = session.receiver.id if request.user == session.caller else session.caller.id
        broadcast_ws_message(
            f"user_{peer_id}",
            {
                "type": "call_state_change",
                "session_id": session.id,
                "status": "ended"
            }
        )
        return JsonResponse({"status": "success"})

    elif action == 'add_ice':
        session_id = data.get('session_id')
        candidate = data.get('candidate')
        session = get_object_or_404(CallSession, id=session_id)
        
        if request.user == session.caller:
            candidates = session.caller_ice or []
            candidates.append(candidate)
            session.caller_ice = candidates
        else:
            candidates = session.receiver_ice or []
            candidates.append(candidate)
            session.receiver_ice = candidates
        
        session.save()
        return JsonResponse({"status": "success"})

    elif action == 'poll':
        session_id = data.get('session_id')
        response_data = {"status": "success"}

        if session_id:
            try:
                session = CallSession.objects.get(id=session_id)
                response_data["session"] = {
                    "id": session.id,
                    "status": session.status,
                    "caller_id": session.caller.id,
                    "caller_name": session.caller.get_full_name() or session.caller.username,
                    "receiver_id": session.receiver.id if session.receiver else None,
                    "receiver_name": session.receiver.get_full_name() or session.receiver.username if session.receiver else None,
                    "call_type": session.call_type,
                    "caller_sdp": session.caller_sdp,
                    "receiver_sdp": session.receiver_sdp,
                    "caller_ice": session.caller_ice,
                    "receiver_ice": session.receiver_ice,
                }
            except CallSession.DoesNotExist:
                response_data["status"] = "not_found"
            return JsonResponse(response_data)

        # Ringing check (for incoming calls)
        incoming = CallSession.objects.filter(receiver=request.user, status='ringing').order_by('-id').first()
        if incoming:
            response_data["incoming_call"] = {
                "session_id": incoming.id,
                "caller_id": incoming.caller.id,
                "caller_name": incoming.caller.get_full_name() or incoming.caller.username,
                "caller_avatar": incoming.caller.profile_pic.url if incoming.caller.profile_pic else None,
                "call_type": incoming.call_type
            }

        # Active check
        active = CallSession.objects.filter(
            (Q(caller=request.user) | Q(receiver=request.user)),
            status='active'
        ).order_by('-id').first()
        if active:
            is_caller = (active.caller == request.user)
            response_data["active_call"] = {
                "session_id": active.id,
                "status": active.status,
                "is_caller": is_caller,
                "peer_name": (active.receiver.get_full_name() or active.receiver.username) if is_caller else (active.caller.get_full_name() or active.caller.username),
                "call_type": active.call_type,
                "caller_sdp": active.caller_sdp,
                "receiver_sdp": active.receiver_sdp,
                "caller_ice": active.caller_ice,
                "receiver_ice": active.receiver_ice
            }

        return JsonResponse(response_data)

    return JsonResponse({"error": "Invalid action"}, status=400)


@csrf_exempt
@login_required
def schedule_meeting(request):
    from .models import ScheduledMeeting, ChatRoom, ChatMessage
    
    if request.method == "POST":
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        scheduled_time_str = request.POST.get('scheduled_time')
        duration_minutes = request.POST.get('duration_minutes', 30)
        user_ids = request.POST.getlist('users')
        recurrence = request.POST.get('recurrence', 'none')
        waiting_room = request.POST.get('waiting_room') == 'true'

        if not title or not scheduled_time_str:
            return JsonResponse({"error": "Title and date/time are required."}, status=400)

        try:
            scheduled_time = timezone.datetime.fromisoformat(scheduled_time_str.replace('Z', '+00:00'))
        except ValueError:
            return JsonResponse({"error": "Invalid date/time format."}, status=400)

        # Create room for this meeting
        room = ChatRoom.objects.create(
            name=f"📅 {title}",
            room_type='team',
            description=description or f"Meeting room for {title}",
            created_by=request.user
        )
        room.users.add(request.user)
        room.admins.add(request.user)

        # Create scheduled meeting
        meeting = ScheduledMeeting.objects.create(
            title=title,
            description=description,
            room=room,
            scheduled_time=scheduled_time,
            duration_minutes=int(duration_minutes),
            creator=request.user,
            recurrence=recurrence,
            waiting_room_enabled=waiting_room
        )

        # Add participants
        for uid in user_ids:
            try:
                u = User.objects.get(id=int(uid))
                meeting.participants.add(u)
                room.users.add(u)
                
                # Notify the participant via private DM
                notification_text = (
                    f"🔔 **Meeting Invitation**\n\n"
                    f"You have been invited to a meeting by **{request.user.get_full_name() or request.user.username}**.\n"
                    f"📅 **Title**: {title}\n"
                    f"🕒 **Time**: {localtime(scheduled_time).strftime('%d %b %Y %H:%M')}\n"
                    f"⌛ **Duration**: {duration_minutes} minutes\n"
                    f"🔗 Click the Meetings tab in Collaboration Hub to join."
                )
                ChatMessage.objects.create(
                    sender=request.user,
                    receiver=u,
                    text=notification_text
                )
            except Exception:
                continue

        return JsonResponse({
            "status": "success",
            "meeting": {
                "id": meeting.id,
                "title": meeting.title,
                "scheduled_time": localtime(meeting.scheduled_time).strftime("%d %b %Y %H:%M"),
                "room_id": room.id
            }
        })

    # For GET
    meetings = ScheduledMeeting.objects.filter(
        Q(creator=request.user) | Q(participants=request.user)
    ).distinct().order_by('scheduled_time')

    meetings_data = []
    for m in meetings:
        meetings_data.append({
            "id": m.id,
            "title": m.title,
            "description": m.description,
            "room_id": m.room.id if m.room else None,
            "room_name": m.room.name if m.room else "",
            "scheduled_time": localtime(m.scheduled_time).strftime("%d %b %Y %H:%M"),
            "duration": m.duration_minutes,
            "creator_name": m.creator.get_full_name() or m.creator.username,
            "waiting_room": m.waiting_room_enabled,
            "is_creator": (m.creator == request.user)
        })

    return JsonResponse({"status": "success", "meetings": meetings_data})


# ==============invicode===============

from django.shortcuts import render, redirect
from .forms import InvoiceForm
from .models import Service, Invoice

def create_invoice(request):
    if request.method == "POST":
        form = InvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save(commit=False)

            service = invoice.service
            invoice.base_amount = service.amount

            invoice.gst_amount = service.amount * invoice.gst_percent / 100
            invoice.total_amount = service.amount + invoice.gst_amount

            invoice.save()
            return redirect('invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceForm()

    return render(request, 'invoice_create.html', {"form": form})



from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import Invoice
from .forms import InvoiceForm, InvoiceItemFormSet

from django.template.loader import get_template
from xhtml2pdf import pisa  # for PDF generation
# views.py
from django.shortcuts import render, redirect
from .forms import InvoiceForm, InvoiceItemFormSet
from .models import Invoice

def create_invoice(request):
    if request.method == "POST":
        form = InvoiceForm(request.POST)
        formset = InvoiceItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            invoice = form.save()
            items = formset.save(commit=False)
            for item in items:
                item.invoice = invoice
                item.save()
            return redirect('invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceForm()
        formset = InvoiceItemFormSet()

    return render(request, 'invoce/invoice_create.html', {
        'form': form,
        'formset': formset
    })


def invoice_detail(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    return render(request, 'invoce/invoice_detail.html', {'invoice': invoice})


from django.contrib.staticfiles import finders

def download_invoice_pdf(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)

    # Get absolute path of the logo using finders
    logo_file = finders.find('images/Logo.png')
    if logo_file:
        logo_path = "file:///" + logo_file.replace("\\", "/")
    else:
        logo_path = ""  # fallback

    template_path = 'invoce/invoice_pdf.html'
    context = {
        'invoice': invoice,
        'logo_path': logo_path,
    }

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.invoice_number}.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response, encoding='UTF-8')

    if pisa_status.err:
        return HttpResponse('Error generating PDF <pre>' + html + '</pre>')

    return response


def create_menu(request):
    return render(request, 'invoce/create_menu.html')



from twilio.rest import Client
from twilio.rest import Client

def send_invoice_whatsapp(invoice, phone_number):
    account_sid = 'YOUR_TWILIO_SID'
    auth_token = 'YOUR_TWILIO_AUTH_TOKEN'
    client = Client(account_sid, auth_token)

    pdf_url = f"http://yourdomain.com/invoice/{invoice.id}/pdf/"

    message = client.messages.create(
        from_='whatsapp:+917794053340',  # Twilio sandbox number
        body=f"Hello {invoice.client.name}, your invoice #{invoice.invoice_number} is ready.\nDownload here: {pdf_url}",
        to=f'whatsapp:+{phone_number}'
    )

    return message.sid
def send_invoice_whatsapp_view(request, id):
    invoice = get_object_or_404(Invoice, id=id)
    phone = invoice.client.phone

    send_invoice_whatsapp(invoice, phone)

    messages.success(request, "Invoice sent to WhatsApp successfully!")
    return redirect("invoice_view", id=id)


# views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import Client, Service
from .forms import ClientForm, ServiceForm

# ---- Client CRUD ----
def client_list(request):
    clients = Client.objects.all()
    return render(request, 'invoce/client_list.html', {'clients': clients})

def client_create(request):
    form = ClientForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('client_list')
    return render(request, 'invoce/client_form.html', {'form': form})

def client_update(request, pk):
    client = get_object_or_404(Client, pk=pk)
    form = ClientForm(request.POST or None, instance=client)
    if form.is_valid():
        form.save()
        return redirect('client_list')
    return render(request, 'invoce/client_form.html', {'form': form})

def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == "POST":
        client.delete()
        return redirect('client_list')
    return render(request, 'invoce/client_confirm_delete.html', {'client': client})

# ---- Service CRUD ----
def service_list(request):
    services = Service.objects.all()
    return render(request, 'invoce/service_list.html', {'services': services})

def service_create(request):
    form = ServiceForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('service_list')
    return render(request, 'invoce/service_form.html', {'form': form})

def service_update(request, pk):
    service = get_object_or_404(Service, pk=pk)
    form = ServiceForm(request.POST or None, instance=service)
    if form.is_valid():
        form.save()
        return redirect('service_list')
    return render(request, 'invoce/service_form.html', {'form': form})

def service_delete(request, pk):
    service = get_object_or_404(Service, pk=pk)
    if request.method == "POST":
        service.delete()
        return redirect('service_list')
    return render(request, 'invoce/service_confirm_delete.html', {'service': service})



def create_invoice(request):
    if request.method == "POST":
        form = InvoiceForm(request.POST)
        formset = InvoiceItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            invoice = form.save()   # Saves discount also

            # Save invoice items
            items = formset.save(commit=False)
            for item in items:
                item.invoice = invoice
                item.save()

            return redirect("invoice_detail", invoice.id)

    else:
        form = InvoiceForm()
        formset = InvoiceItemFormSet()

    return render(request, "invoce/invoice_create.html", {"form": form, "formset": formset})


def invoice_view(request, id):
    invoice = get_object_or_404(Invoice, id=id)
    return render(request, "invoce/view_inoice.html", {"invoice": invoice})

def invoice_edit(request, id):
    invoice = get_object_or_404(Invoice, id=id)

    if request.method == "POST":
        invoice.invoice_number = request.POST.get("invoice_number")
        invoice.discount_percent = request.POST.get("discount_percent")
        invoice.note = request.POST.get("note")
        invoice.created_at = request.POST.get("created_at")
        invoice.save()

        # Update items
        for item in invoice.items.all():
            amount = request.POST.get(f"amount_{item.id}")
            if amount:
                item.amount = amount
                item.save()

        return redirect("invoice_view", id=id)

    
    return render(request, "invoce/invoice_edit.html", { "invoice": invoice})
from django.shortcuts import redirect, get_object_or_404
from .models import Invoice

def invoice_delete(request, id):
    invoice = get_object_or_404(Invoice, id=id)
    invoice.delete()
    return redirect("viewall_invoices")  # ← Redirect to your invoice list page


def viewall_invoices(request):
    invoices = Invoice.objects.all()  # Make sure this returns real invoices
    return render(request, "invoce/all_view.html", {"invoices": invoices})



#-------------------------------------#
    #manager_dashboard_views
#-------------------------------------#

from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db.models import Q
from .forms import LoginForm

User = get_user_model()

def login_user(request):
    return redirect('viewlogins')



@login_required(login_url='manager_login')
def logout_user(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('viewlogins')


from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Project, Team, User, Leave

@login_required(login_url='manager_login')
def index(request):
    manager = request.user

    # 🔹 Projects
    manager_projects = Project.objects.filter(assigned_manager=manager)

    projects_received = manager_projects.filter(assigned_team__isnull=True).count()
    projects_assigned = manager_projects.filter(assigned_team__isnull=False).count()
    projects_completed = manager_projects.filter(status="Completed").count()

    # 🔹 Teams
    # teams = Team.objects.filter(lead__managed_projects__assigned_manager=manager).distinct() # This might be too restrictive
    teams = Team.objects.all()

    # 🔹 Employees and Team Leads
    employees = User.objects.filter(role='Employee')
    team_leads_count = User.objects.filter(role='TeamLead').count()

    # 🔹 Leaves waiting for manager approval
    employee_leave_count = Leave.objects.filter(user__role='Employee', status="Pending Manager Approval").count()
    teamlead_leave_count = Leave.objects.filter(user__role='TeamLead', status="Pending Manager Approval").count()

    # 🔸 Recent Daily Reports
    daily_reports = DailyReport.objects.filter(
        project__in=manager_projects
    ).select_related("user", "project").order_by("-report_date")[:6]

    today_date = timezone.localdate()
    attendance = Attendance.objects.filter(user=manager, date=today_date).first()

    context = {
        "manager": manager,
        "projects_received": projects_received,
        "projects_assigned": projects_assigned,
        "projects_completed": projects_completed,
        "employees": employees,
        "teams": teams,
        "team_leads_count": team_leads_count,
        "employee_leave_count": employee_leave_count,
        "teamlead_leave_count": teamlead_leave_count,
        "daily_reports": daily_reports,
        "attendance": attendance,
    }

    return render(request, "main/index.html", context)

@login_required(login_url='manager_login')
def teams(request):
    user = request.user
    teams = Team.objects.filter(is_active=True).prefetch_related('members')
    chart_data = {
        "labels": [t.name for t in teams],
        "values": [t.members.count() for t in teams],
    }
    return render(request, "main/teams.html", {
        "manager": user,
        "teams": teams,
        "chart_data": json.dumps(chart_data),
    })


@login_required(login_url='manager_login')
def delete_team(request, id):
    team = get_object_or_404(Team, id=id)
    if request.method == "POST":
        team.delete()
        messages.success(request, f"✅ Team '{team.name}' has been deleted.")
        return redirect("main:teams")
    return render(request, "main/confirm_delete.html", {"team": team})




@login_required(login_url='manager_login')
def archived_teams(request):
    teams = Team.objects.filter(is_active=False)
    return render(request, "archived_teams.html", {"teams": teams})



@login_required(login_url='manager_login')
def restore_team(request, id):
    team = get_object_or_404(Team, id=id)
    team.is_active = True
    team.save()
    messages.success(request, "✅ Team restored successfully")
    return redirect("archived_teams")


@login_required(login_url='manager_login')
def cteam(request):
    if request.method == "POST":
        team_name = request.POST.get("team_name")
        team_lead_id = request.POST.get("team_lead")
        member_ids = request.POST.getlist("team_members")

        # Validation
        if not team_name:
            messages.error(request, "⚠ Team name is required.")
            return redirect("cteam")

        if not team_lead_id:
            messages.error(request, "⚠ Please select a team lead.")
            return redirect("cteam")

        if len(member_ids) < 2 or len(member_ids) > 10:
            messages.error(request, "⚠ Team must have 2–10 members.")
            return redirect("cteam")

        lead = get_object_or_404(User, id=team_lead_id, role='TeamLead')

        if Team.objects.filter(lead=lead).exists():
            messages.warning(request, f"{lead.username} already leads a team.")
            return redirect("cteam")

        # Create team
        team = Team.objects.create(name=team_name, lead=lead)

        members = User.objects.filter(
            id__in=member_ids,
            role='Employee'
        ).exclude(id=lead.id)

        team.members.set(members)

        messages.success(request, f"Team '{team_name}' created successfully.")
        return redirect("teams")

    # Get TeamLead users not already leading a team
    assigned_lead_ids = Team.objects.values_list('lead__id', flat=True)
    leads = User.objects.filter(role='TeamLead').exclude(id__in=assigned_lead_ids)

    # Get Employee users not already in any team
    assigned_member_ids = Team.objects.values_list('members__id', flat=True)
    employees = User.objects.filter(role='Employee').exclude(id__in=assigned_member_ids)

    return render(request, "main/cteam.html", {
        "leads": leads,
        "employees": employees
    })

from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages

User = get_user_model()



@login_required(login_url='manager_login')
def edit_team(request, team_id):
    team = get_object_or_404(Team, id=team_id)

    # -------- AVAILABLE TEAM LEADS --------
    other_team_lead_ids = Team.objects.exclude(
        id=team.id
    ).values_list('lead_id', flat=True)

    available_leads = User.objects.filter(
        role='TeamLead'
    ).exclude(id__in=other_team_lead_ids)

    # -------- AVAILABLE MEMBERS --------
    assigned_member_ids = Team.objects.exclude(
        id=team.id
    ).values_list('members', flat=True)

    available_members = User.objects.filter(
        role='Employee'
    ).exclude(id__in=assigned_member_ids)

    # -------- AVAILABLE PROJECTS --------
    projects = (
        Project.objects.filter(assigned_team__isnull=True) |
        Project.objects.filter(assigned_team=team)
    ).distinct()

    if request.method == "POST":
        name = request.POST.get("name")
        lead_id = request.POST.get("lead")
        member_ids = request.POST.getlist("members")
        project_ids = request.POST.getlist("projects")

        if not name:
            messages.error(request, "Team name is required.")
            return redirect('edit_team', team_id=team.id)

        if not lead_id:
            messages.error(request, "Please select a team lead.")
            return redirect('edit_team', team_id=team.id)

        if len(member_ids) < 2 or len(member_ids) > 10:
            messages.error(request, "Team must have 2–10 members.")
            return redirect('edit_team', team_id=team.id)

        lead = get_object_or_404(User, id=lead_id, role='TeamLead')

        if Team.objects.filter(lead=lead).exclude(id=team.id).exists():
            messages.warning(request, f"{lead.username} already leads another team.")
            return redirect('edit_team', team_id=team.id)

        # -------- SAVE TEAM --------
        team.name = name
        team.lead = lead
        team.save()

        team.members.set(
            User.objects.filter(id__in=member_ids, role='Employee')
        )

        # -------- PROJECT ASSIGN --------
        Project.objects.filter(id__in=project_ids).update(assigned_team=team)
        Project.objects.filter(assigned_team=team).exclude(
            id__in=project_ids
        ).update(assigned_team=None)

        messages.success(request, f'Team "{team.name}" updated successfully.')
        return redirect('teams')

    return render(request, 'main/edit_team.html', {
        'team': team,
        'available_leads': available_leads,
        'available_members': available_members,
        'projects': projects,
    })


# ======================
# 🧑‍💼 Team Lead Management
# ======================
 
@login_required(login_url='manager_login')
def teamlead(request):
    if request.method == "POST":

        if "create_lead" in request.POST:
            user_id = request.POST.get("emp_id")
            user = get_object_or_404(User, id=user_id, role='Employee')

            user.role = 'TeamLead'
            user.save()

            messages.success(request, f"✅ {user.username} promoted to Team Lead.")
            return redirect("teamlead")

        elif "remove_lead" in request.POST:
            lead_id = request.POST.get("lead_id")
            lead = User.objects.filter(id=lead_id, role='TeamLead').first()

            if lead:
                Team.objects.filter(lead=lead).update(lead=None)
                lead.role = 'Employee'
                lead.save()
                messages.success(request, "✅ Team Lead removed successfully.")
            else:
                messages.warning(request, "⚠ No Team Lead found.")

            return redirect("teamlead")

    leads = User.objects.filter(role='TeamLead')
    employees = User.objects.filter(role='Employee')

    return render(request, "main/teamlead.html", {
        "leads": leads,
        "employees": employees
    })




@login_required(login_url='manager_login')
def assignp(request):
    if request.method == "POST":
        project_id = request.POST.get("project")
        team_id = request.POST.get("team")
        deadline = request.POST.get("deadline")

        project = Project.objects.filter(id=project_id, assigned_team__isnull=True).first()
        team = Team.objects.filter(id=team_id).first()

        if not project:
            messages.error(request, "Invalid or already assigned project.")
            return redirect("main:assignp")

        if not team:
            messages.error(request, "Please select a valid team.")
            return redirect("assignp")

        project.assigned_team = team
        project.deadline = deadline
        project.status = "Assigned"
        project.save()
        messages.success(request, f"✅ Project '{project.name}' assigned to {team.name}.")
        return redirect("projects")

    projects_received = Project.objects.filter(assigned_team__isnull=True)
    teams = Team.objects.all()
    return render(request, "main/assignp.html", {"projects_received": projects_received, "teams": teams})

@require_POST

@login_required(login_url='manager_login')
def reject_project(request, pk):
    project = get_object_or_404(
        Project,
        pk=pk,
        assigned_team__isnull=True
    )

    reason = request.POST.get("reject_reason")

    project.status = "Rejected"
    project.reject_reason = reason
    project.save()
 
    messages.success(request, "Project rejected successfully")
 
    return redirect('main/projects')




# ======================
# 👨‍💼 Employees
# ======================

@login_required(login_url='manager_login')
def employees(request):
    user = request.user
    employees = User.objects.filter(role='Employee')
    teamleads = User.objects.filter(role='TeamLead')

    return render(request, "main/employees.html", {
        "manager":user,
        "employees": employees,
        "teamleads": teamleads
    })

@require_POST

@login_required(login_url='manager_login')
def upgrade_employee(request, emp_id):
    user = get_object_or_404(User, id=emp_id, role='Employee')

    user.role = 'TeamLead'
    user.save()

    messages.success(request, f"{user.username} promoted to Team Lead!")
    return redirect("employees")



from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
User = get_user_model()


@login_required(login_url='manager_login')
def newemp(request):
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        address = request.POST.get("address")
        salary = request.POST.get("salary")
        gender = request.POST.get("gender")
        date_of_birth = request.POST.get("date_of_birth")
        date_of_joining = request.POST.get("date_of_joining")
        status = request.POST.get("status")
        experience_years = request.POST.get("experience_years") or 0
        previous_company = request.POST.get("previous_company") or ""
        department = request.POST.get("department")
        profile_pic = request.FILES.get("profile_pic")
        document = request.FILES.get("document")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect("newemp")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("newemp")

        # Create employee
        user = User(
            username=username,
            email=email,
            phone=phone,
            address=address,
            salary=salary,
            gender=gender,
            date_of_birth=date_of_birth,
            date_of_joining=date_of_joining,
            status=status,
            experience_years=experience_years,
            previous_company=previous_company,
            department=department,
            profile_pic=profile_pic,
            document=document,
            role='Employee'
        )

        user.set_password(password)
        user.save()

        messages.success(request, f"Employee '{username}' added successfully.")
        return redirect("employees")

    return render(request, "main/newemp.html")




from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Project, ProjectWorkUpdate, DailyReport  # TL daily reports

@login_required(login_url='manager_login')
def projects(request):
    user = request.user

    # ✅ Projects assigned by HR to THIS manager but not yet assigned to a team
    projects_received = Project.objects.filter(
        assigned_manager=user,
        assigned_team__isnull=True
    )

    # ✅ Projects already assigned by THIS manager to teams
    projects_assigned = Project.objects.filter(
        assigned_manager=user,
        assigned_team__isnull=False
    ).select_related(
        "assigned_team__lead"
    ).prefetch_related(
        "assigned_team__members"
    )

    # ✅ Work updates only for THIS manager’s projects
    work_updates = ProjectWorkUpdate.objects.filter(
        project__assigned_manager=user
    ).select_related(
        "project",
        "team_lead"
    ).order_by("-created_at")

    # ✅ TL daily reports for projects managed by this manager
    manager_projects = Project.objects.filter(assigned_manager=user)  # queryset of Project objects
    tl_reports = DailyReport.objects.filter(
        project__in=manager_projects  # filter by ForeignKey
    ).select_related("teamlead", "project").order_by("-report_date")

    print("reports",)

    return render(request, "main/projects.html", {
        "manager": user,
        "projects_received": projects_received,
        "projects_assigned": projects_assigned,
        "work_updates": work_updates,
        
    })





@require_POST

@login_required(login_url='manager_login')
def approve_project(request, pk):
    project = get_object_or_404(Project, pk=pk)
    project.status = 'Approved'
    project.save()
    messages.success(request, f"Project {project.project_id} approved.")
    return redirect("main:projects")


@require_POST

@login_required(login_url='manager_login')
def decline_project(request, pk):
    project = get_object_or_404(Project, pk=pk)
    project.status = 'Declined'
    project.save()
    messages.info(request, f"Project {project.project_id} declined.")
    return redirect("main:projects")



from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required


@login_required(login_url='manager_login')   
def employee_profile1(request, emp_id):
    user = get_object_or_404(User, emp_id=emp_id)   


    projects = Project.objects.filter(assigned_team__members=user)
    attendance = calculate_attendance(user)

    context = {
        "employee": user,
       
        "projects": projects,
        "salary": user.salary,
         "attendance": attendance,
    }
    return render(request, "main/employee_profile1.html", context)



@require_POST

@login_required(login_url='manager_login')
def submit_work_update(request):
    project_id = request.POST.get("project_id")
    work_details = request.POST.get("work_details")
    file = request.FILES.get("file")

    project = get_object_or_404(Project, id=project_id)
    user = request.user

    if user.role != "TeamLead":
        messages.error(request, "Only Team Leads can submit work updates.")
        return redirect("main:projects")

    ProjectWorkUpdate.objects.create(
        project=project,
        team_lead=user,
        work_details=work_details,
        file=file
    )

    messages.success(request, "Work update submitted successfully.")
    return redirect("main:projects")



@login_required
@csrf_exempt
def update_manager_profile(request):
    if request.method == "POST":
        user = request.user

        user.username = request.POST.get("username")
        user.email = request.POST.get("email")
        user.phone = request.POST.get("phone")
        user.address = request.POST.get("address")

        if request.FILES.get("profile_pic"):
            user.profile_pic = request.FILES["profile_pic"]

        new_password = request.POST.get("new_password")
        if new_password:
            user.set_password(new_password)
            update_session_auth_hash(request, user)

        user.save()
        return JsonResponse({"status": "success"})




from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

@require_POST
@login_required(login_url='manager_login')
def manager_submit_to_hr(request):

    # 🔐 Only manager allowed
    if request.user.role != "Manager":
        messages.error(request, "Unauthorized access.")
        return redirect("main:projects")

    project_id = request.POST.get("project")
    details = request.POST.get("details")

    if not project_id:
        messages.error(request, "Please select a project.")
        return redirect("main:projects")

    # -----------------------
    # Get project of manager
    # -----------------------
    project = get_object_or_404(
        Project,
        id=project_id,
        assigned_manager=request.user
    )

    # -----------------------
    # Get latest TeamLead update
    # -----------------------
    work_update = ProjectWorkUpdate.objects.filter(
        project=project
    ).order_by("-created_at").first()

    if not work_update:
        messages.error(request, "Team Lead has not submitted work yet.")
        return redirect("main:projects")

    # -----------------------
    # Save submission to HR
    # -----------------------
    ManagerProjectSubmission.objects.create(
        manager=request.user,
        project=project,
        details=details,
        file=work_update.file  # 🔥 file comes from TeamLead
    )

    # -----------------------
    # Update project status
    # -----------------------
    project.status = "Completed"
    project.save()

    messages.success(request, "Project successfully submitted to HR.")
    return redirect("projects")

# **************************************************
           # *_______________tl views_______________*

# ***********************************************

from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model, login
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Project, Team

User = get_user_model()

def tl_login(request):
    return redirect('viewlogins')

from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import login_required
from .models import Project, Team

@login_required
def tl_project_list(request):
    team = Team.objects.filter(lead=request.user).first()

    if team:
        projects = list(Project.objects.filter(assigned_team=team))
    else:
        projects = []

    # Calculate statistics & attach computed fields
    active_count = 0
    completed_count = 0
    pending_tasks_count = 0
    
    for p in projects:
        # Status checks
        if p.status == 'Completed':
            completed_count += 1
        elif p.status != 'Rejected':
            active_count += 1
            
        # Task calculations
        p.total_tasks = p.task_set.count()
        p.completed_tasks = p.task_set.filter(status='Completed').count()
        p.progress_pct = int(p.completed_tasks * 100 / p.total_tasks) if p.total_tasks > 0 else 0
        pending_tasks_count += p.task_set.filter(status='Pending').count()

    # Fetch upcoming tasks and recent activities
    from datetime import date, timedelta
    upcoming_deadlines_count = 0
    upcoming_tasks = []
    recent_activities = []
    
    if team:
        from .models import Task, DailyReport
        upcoming_tasks = Task.objects.filter(project__assigned_team=team, status__in=['Pending', 'Submitted']).order_by('end_date')[:5]
        upcoming_deadlines_count = Task.objects.filter(project__assigned_team=team, status__in=['Pending', 'Submitted'], end_date__lte=date.today() + timedelta(days=14)).count()
        
        # Pull recent daily reports to show as activities
        recent_reports = DailyReport.objects.filter(project__assigned_team=team).order_by('-id')[:5]
        for r in recent_reports:
            recent_activities.append({
                'type': 'task_completed' if 'complete' in r.tasks_completed.lower() else 'report_submitted',
                'user_name': r.user.get_full_name() or r.user.username,
                'project_name': r.project.name if r.project else 'General',
                'detail': r.tasks_completed[:40] + '...' if len(r.tasks_completed) > 40 else r.tasks_completed,
            })
            
    # Default activities in case there are none
    if not recent_activities:
        recent_activities = [
            {'type': 'general', 'detail': 'Dashboard initialized.'}
        ]

    return render(request, 'tl/project_list.html', {
        'projects': projects,
        'active_projects_count': active_count,
        'completed_projects_count': completed_count,
        'pending_tasks_count': pending_tasks_count,
        'upcoming_deadlines_count': upcoming_deadlines_count or len(upcoming_tasks),
        'upcoming_tasks': upcoming_tasks,
        'recent_activities': recent_activities
    })



@login_required
def tl_dashboard(request):
    user = request.user

    if user.role != "TeamLead":
        messages.error(request, "Access denied")
        return redirect('tl_login')

    from .models import Task, DailyReport, Leave, Attendance
    
    team = Team.objects.filter(lead=user).first()

    members = User.objects.none()
    projects = Project.objects.none()
    reports = DailyReport.objects.none()
    upcoming_tasks = Task.objects.none()
    all_tasks = Task.objects.none()
    member_status_list = []
    recent_activities = []

    today = timezone.localdate()

    if team:
        members = team.members.all()
        projects = Project.objects.filter(assigned_team=team)

        # 1. Recent Reports
        reports = DailyReport.objects.filter(
            user__in=members
        ).select_related("user", "project").order_by("-report_date", "-id")[:5]

        # 2. Upcoming Deadlines
        upcoming_tasks = Task.objects.filter(
            project__assigned_team=team,
            status__in=['Pending', 'Submitted']
        ).select_related("project").order_by("end_date")[:5]

        # 3. All Tasks
        all_tasks = Task.objects.filter(project__assigned_team=team)

        # 4. Member Status
        today_attendances = {
            att.user_id: att for att in Attendance.objects.filter(user__in=members, date=today)
        }
        
        for m in members:
            att = today_attendances.get(m.id)
            check_in = att.check_in_time.strftime("%I:%M %p") if att and att.check_in_time else "—"
            att_status = att.status if att else "Absent"
            
            # Current task (Active task)
            curr_task = Task.objects.filter(members=m, status__in=["Pending", "Submitted"]).order_by("end_date").first()
            task_name = curr_task.task_name if curr_task else "No Active Task"
            task_status = curr_task.status if curr_task else "—"
            
            member_status_list.append({
                "member": m,
                "check_in": check_in,
                "attendance_status": att_status,
                "current_task": task_name,
                "task_status": task_status
            })

        # 5. Recent Activity Feed
        # Combine logs
        for r in reports[:3]:
            recent_activities.append({
                "type": "Report Submitted",
                "user": r.user.get_full_name() or r.user.username,
                "detail": f"submitted daily report for project: {r.project.name if r.project else 'General'}",
                "time_label": "Today" if r.report_date == today else r.report_date.strftime("%b %d"),
                "icon": "fa-file-invoice",
                "badge_class": "badge-success"
            })
            
        completed_tasks = Task.objects.filter(project__assigned_team=team, status="Completed").order_by("-id")[:3]
        for t in completed_tasks:
            recent_activities.append({
                "type": "Task Completed",
                "user": "Team Member",
                "detail": f"completed the task: '{t.task_name}'",
                "time_label": "Recently",
                "icon": "fa-circle-check",
                "badge_class": "badge-info"
            })

        recent_leaves = Leave.objects.filter(user__in=members).order_by("-id")[:2]
        for l in recent_leaves:
            recent_activities.append({
                "type": "Leave Request",
                "user": l.user.get_full_name() or l.user.username,
                "detail": f"applied for leave: '{l.reason[:30]}...'",
                "time_label": l.from_date.strftime("%b %d"),
                "icon": "fa-calendar-minus",
                "badge_class": "badge-purple"
            })

        # Sorting: leave recent logs
        recent_activities = recent_activities[:6]

    # Calculate cards metrics
    employee_count = members.count()
    project_count = projects.count()
    open_tasks_count = all_tasks.exclude(status="Completed").count()
    pending_reports_count = reports.count()
    
    # Efficiency calculation
    total_t = all_tasks.count()
    comp_t = all_tasks.filter(status="Completed").count()
    efficiency = int((comp_t / total_t) * 100) if total_t > 0 else 94

    # TL's own today attendance (for check-in/out button state in header)
    tl_attendance = Attendance.objects.filter(user=user, date=today).first()

    context = {
        "team": team,
        "members": members,
        "employee_count": employee_count,
        "projects": projects,
        "project_count": project_count,
        "open_tasks_count": open_tasks_count,
        "pending_reports_count": pending_reports_count,
        "efficiency": efficiency,
        "reports": reports,
        "upcoming_tasks": upcoming_tasks,
        "member_status_list": member_status_list,
        "recent_activities": recent_activities,
        "attendance": tl_attendance,  # TL's own today attendance
    }

    return render(request, "tl/tl_dashboard.html", context)


@login_required
def tl_employee_list(request):
    user = request.user

    if user.role != "TeamLead":
        return redirect('tl_login')

    team = Team.objects.filter(lead=user).first()

    if team:
        members = team.members.all()
    else:
        members = []

    # Attach projects and detailed KPI metrics to each employee
    for emp in members:
        emp.assigned_projects = Project.objects.filter(assigned_team__members=emp)
        
        # Attach project details (status, progress rate)
        for project in emp.assigned_projects:
            project_tasks = Task.objects.filter(project=project, members=emp)
            p_total = project_tasks.count()
            p_comp = project_tasks.filter(status="Completed").count()
            project.employee_completion_rate = int((p_comp / p_total) * 100) if p_total > 0 else 0

        # Compute stats metrics
        emp.projects_count = emp.assigned_projects.count()
        
        tasks = Task.objects.filter(members=emp)
        emp.total_tasks = tasks.count()
        emp.completed_tasks_count = tasks.filter(status="Completed").count()
        emp.pending_tasks_count = tasks.exclude(status="Completed").count()
        
        attendances = Attendance.objects.filter(user=emp)
        total_working_days = attendances.count()
        present_days = attendances.filter(status__icontains="Present").count()
        emp.attendance_percentage = int((present_days / total_working_days) * 100) if total_working_days > 0 else 94
        
        approved_leaves = Leave.objects.filter(user=emp, status__icontains="Approved").count()
        emp.leave_balance = 24 - approved_leaves

    context = {
        "team": team,
        "members": members,
    }
    return render(request, "tl/emps_list.html", context)








from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib.auth import get_user_model
from .models import Project, Task, Team

User = get_user_model()


@login_required
def assign_task(request):
    user = request.user

    # ✅ Get projects where this user is team lead
    projects = Project.objects.filter(
        assigned_team__lead=user
    ).distinct()

    # ✅ If no projects → show empty
    if not projects.exists():
        return render(request, "tl/assign_task.html", {
            "projects": [],
            "tasks": []
        })

    # ✅ Only tasks of this TL projects
    tasks = Task.objects.filter(
        project__assigned_team__lead=user
    ).distinct().order_by("-id")

    if request.method == "POST":
        project_id = request.POST.get("project")
        task_name = request.POST.get("task")
        start_date = request.POST.get("tfrom")
        end_date = request.POST.get("tto")
        file = request.FILES.get("taskFile")
        member_ids = request.POST.getlist("members")

        # 🔐 Ensure project belongs to THIS TL
        project = get_object_or_404(
            Project,
            id=project_id,
            assigned_team__lead=user
        )

        if not member_ids:
            messages.error(request, "Please select at least one employee.")
            return redirect("assign_task")

        # ✅ Only members of this TL team
        members = User.objects.filter(
            teams__lead=user,
            emp_id__in=member_ids
        ).distinct()

        try:
            with transaction.atomic():
                task = Task.objects.create(
                    project=project,
                    task_name=task_name,
                    start_date=start_date,
                    end_date=end_date,
                    file=file
                )
                task.members.set(members)

            messages.success(request, "Task assigned successfully!")
            return redirect("assign_task")

        except Exception as e:
            messages.error(request, f"Error: {e}")

    return render(request, "tl/assign_task.html", {
        "projects": projects,
        "tasks": tasks
    })


@login_required
def get_project_employees(request, project_id):
    user = request.user

    # 🔐 Ensure project belongs to this TL
    project = get_object_or_404(
        Project,
        id=project_id,
        assigned_team__lead=user
    )

    employees = User.objects.filter(
        teams__lead=user
    ).distinct()

    data = []
    for emp in employees:
        image_url = emp.profile_pic.url if hasattr(emp, 'profile_pic') and emp.profile_pic else "/static/img/user.png"

        data.append({
            "id": emp.emp_id,
            "name": emp.get_full_name() or emp.username,
            "role": emp.role,
            "image": image_url,
        })

    return JsonResponse({"employees": data})
    return JsonResponse({"employees": data})

def delete_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    task.delete()
    return redirect("assign_task")


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from .models import Task, Project
from django.contrib.auth import get_user_model
User = get_user_model()


def edit_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)

    # employees from project team
    employees = task.project.assigned_team.members.all()

    # get all projects
    projects = Project.objects.all()  # ✅ added this line

    if request.method == "POST":
        project_id = request.POST.get("project")
        task.task_name = request.POST.get("task")
        task.start_date = request.POST.get("tfrom")
        task.end_date = request.POST.get("tto")
        task.status = request.POST.get("status")

        member_ids = request.POST.getlist("members")
        project = get_object_or_404(Project, id=project_id)

        if not member_ids:
            messages.error(request, "Please select at least one employee.")
            return redirect("edit_task", task_id=task.id)

        members = User.objects.filter(emp_id__in=member_ids)

        try:
            with transaction.atomic():
                task.save()
                # ✅ update members
                task.members.set(members)

            messages.success(request, "Task updated successfully!")
            return redirect("assign_task")

        except Exception as e:
            messages.error(request, f"Error: {e}")

    return render(request, "tl/edit_task.html", {
        "task": task,
        "employees": employees,
        "projects": projects
    })

from django.contrib.auth import get_user_model

User = get_user_model()

def get_users(role=None, exclude_ids=None, search=None):
    """
    Returns a queryset of users filtered by role, excluding specific IDs, and optional search.
    
    :param role: str, the role to filter by, e.g., "Employee", "Manager", "TeamLead"
    :param exclude_ids: list of user IDs to exclude
    :param search: str, optional text to search in username or emp_id
    """
    qs = User.objects.all()

    if role:
        qs = qs.filter(role=role)
    
    if exclude_ids:
        qs = qs.exclude(id__in=exclude_ids)
    
    if search:
        qs = qs.filter(Q(username__icontains=search) | Q(emp_id__icontains=search))
    
    return qs





# hr/views.py
from django.shortcuts import render, redirect
from .forms import EmpUpdateForm
from .models import EmpUpdate

def submit_report(request):
    if request.method == 'POST':
        form = EmpUpdateForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.employee = request.user
            # Assuming each employee has a team lead set in their profile
            report.team_lead = request.user.profile.team_lead
            report.save()
            return redirect('employee_report_success')
    else:
        form = EmpUpdateForm()
    return render(request, 'submit_report.html', {'form': form})

# hr/views.py
from django.contrib.auth.decorators import login_required
# hr/views.py
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import EmpUpdate
from django.utils import timezone
from datetime import timedelta

@login_required
def tl_reports_view(request):
    reports = EmpUpdate.objects.filter(team_lead=request.user).order_by('-date', '-time')
    
    # Get filter params
    employee_id = request.GET.get('employee')
    date_filter = request.GET.get('date_filter')  # today / yesterday / custom
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if employee_id:
        reports = reports.filter(employee_id=employee_id)

    if date_filter == 'today':
        today = timezone.now().date()
        reports = reports.filter(date=today)
    elif date_filter == 'yesterday':
        yesterday = timezone.now().date() - timedelta(days=1)
        reports = reports.filter(date=yesterday)
    elif date_filter == 'custom' and start_date and end_date:
        reports = reports.filter(date__range=[start_date, end_date])

    # Pass list of employees for dropdown
    employees = reports.values_list('employee__id', 'employee__username').distinct()

    context = {
        'reports': reports,
        'employees': employees,
        'selected_employee': employee_id,
        'selected_date_filter': date_filter,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'tl/tl_reports.html', context)


from django.contrib.auth import get_user_model
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Team, Project

User = get_user_model()  # ✅ Get your user model dynamically

@login_required
def employee_detail(request, emp_id):
    # Get the member (user) from team
    emp = get_object_or_404(User, id=emp_id)

    # Get projects assigned to this employee
    assigned_projects = Project.objects.filter(assigned_team__members=emp)

    # Attach computed metrics to projects for the employee
    for project in assigned_projects:
        project_tasks = Task.objects.filter(project=project, members=emp)
        p_total = project_tasks.count()
        p_comp = project_tasks.filter(status="Completed").count()
        project.employee_completion_rate = int((p_comp / p_total) * 100) if p_total > 0 else 0
        project.employee_tasks_count = p_total
        project.employee_completed_tasks_count = p_comp
        project.team_size = project.assigned_team.members.count() if project.assigned_team else 1

    # Fetch tasks for the employee
    tasks = Task.objects.filter(members=emp)
    total_tasks = tasks.count()
    completed_tasks_count = tasks.filter(status="Completed").count()
    pending_tasks_count = tasks.exclude(status="Completed").count()

    # Calculate attendance metrics
    attendances = Attendance.objects.filter(user=emp)
    total_working_days = attendances.count()
    present_days = attendances.filter(status__icontains="Present").count()
    attendance_percentage = int((present_days / total_working_days) * 100) if total_working_days > 0 else 94
    if total_working_days == 0:
        total_working_days = 22  # default mock standard

    # Calculate leave balance
    approved_leaves = Leave.objects.filter(user=emp, status__icontains="Approved").count()
    leave_balance = 24 - approved_leaves

    # Get manager (Reporting Manager)
    manager = emp.reporting_manager

    # Fetch recent daily reports
    recent_reports = DailyReport.objects.filter(user=emp).order_by("-id")[:5]

    context = {
        "emp": emp,
        "projects": assigned_projects,
        "total_tasks": total_tasks,
        "completed_tasks_count": completed_tasks_count,
        "pending_tasks_count": pending_tasks_count,
        "total_working_days": total_working_days,
        "attendance_percentage": attendance_percentage,
        "leave_balance": leave_balance,
        "manager": manager,
        "recent_reports": recent_reports,
    }
    return render(request, "tl/employee_detail.html", context)





#___________________________________________*****

#***************md views********************

#___________________________________________*****

from django.shortcuts import render, redirect
from django.db import IntegrityError
from .models import User
from django.contrib.admin.views.decorators import staff_member_required


# login-page----

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect


def login_view(request):
    return redirect('viewlogins')



from django.shortcuts import render, redirect
from django.contrib.auth import get_user_model
from django.db import IntegrityError

User = get_user_model()

ROLE_CREATE_PERMISSION = {
    "MD": ["MD", "HR", "Manager", "TeamLead", "Employee"],
    "HR": ["Manager", "TeamLead", "Employee"],
    "Manager": ["TeamLead", "Employee"],
}

def user_data(request):
    if request.method == "POST":
        creator_role = request.user.role
        selected_role = request.POST.get("role")

        # Permission check
        if creator_role not in ROLE_CREATE_PERMISSION:
            return render(request, "md/register.html", {"error": "You are not allowed to create users."})

        if selected_role not in ROLE_CREATE_PERMISSION[creator_role]:
            return render(request, "md/register.html", {"error": f"You cannot create {selected_role}."})

        # Password check
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        if password != confirm_password:
            return render(request, "register.html", {"error": "Passwords do not match."})

        try:
            # Create user without using create_user() to control role and emp_id
            user = User(
                username=request.POST.get("username"),
                email=request.POST.get("email"),
                first_name=request.POST.get("fullname"),
                role=selected_role  # assign role BEFORE save
            )
            user.set_password(password)

            # Assign other fields
            user.phone = request.POST.get("phone")
            user.address = request.POST.get("address")
            user.department = request.POST.get("department")
            user.gender = request.POST.get("gender")
            user.date_of_birth = request.POST.get("date_of_birth")
            user.date_of_joining = request.POST.get("date_of_joining")
            user.salary = request.POST.get("salary")
            user.status = request.POST.get("status")
            user.experience_years = request.POST.get("experience_years") or 0
            user.previous_company = request.POST.get("previous_company")
            user.profile_pic = request.FILES.get("profile_pic")
            user.document = request.FILES.get("document")

            user.save()  # emp_id generated correctly based on role
            return redirect("md_dashboard")

        except IntegrityError:
            return render(request, "md/register.html", {"error": "Username already exists."})

    return render(request, "md/register.html")





# All counts----
from django.shortcuts import render
from django.contrib.auth import get_user_model
from .models import Project

User = get_user_model()
@staff_member_required 
def md_dashboard(request):

    total_hr = User.objects.filter(role='HR').count()
    total_mr = User.objects.filter(role='Manager').count()
    total_tl = User.objects.filter(role='TeamLead').count()
    total_emp = User.objects.filter(role='Employee').count()

    total_project = Project.objects.count()

    total_cmp = total_hr + total_mr + total_tl + total_emp
    
    # Retrieve all active users to render in the left-hand panel directory
    all_users = User.objects.filter(is_active=True).order_by('role', 'username')

    holiday_stats = {
        "all": Holiday.objects.count(),
        "pending": Holiday.objects.filter(status='Pending').count(),
        "approved": Holiday.objects.filter(status='Approved').count(),
        "draft": Holiday.objects.filter(status='Draft').count(),
    }

    from .models import AttendanceCorrection
    pending_corrections_count = AttendanceCorrection.objects.filter(status='Pending').count()

    context = {
        "total_hr": total_hr,
        "total_mr": total_mr,
        "total_tl": total_tl,
        "total_emp": total_emp,
        "total_project": total_project,
        "total_cmp": total_cmp,
        "all_users": all_users,
        "holiday_stats": holiday_stats,
        "pending_corrections_count": pending_corrections_count,
    }

    return render(request, "md/md_dashboard.html", context)
from django.contrib.auth import logout
from django.shortcuts import redirect

def md_logout(request):
    logout(request)
    return redirect("viewlogins")



def all(req):
    return render(req,"md/all.html")

# hr_lists--------
from django.shortcuts import render
from datetime import date
from .models import User
from django.http import  HttpResponse

from datetime import date
from django.contrib.auth import get_user_model
from django.shortcuts import render

User = get_user_model()

def hr_lists(request):
    data = User.objects.filter(role='HR')

    today = date.today()
    for user in data:
        if user.date_of_joining:
            user.total_days = (today - user.date_of_joining).days
        else:
            user.total_days = 0

    return render(request, "md/hr_list.html", {"data": data})



def mr_lists(request):
    datas = User.objects.filter(role='Manager')  # use lowercase 'role'
    return render(request, "md/mr_list.html", {"datas": datas})



from datetime import date

def tl_list(request):
    datetl = User.objects.filter(role__iexact='TeamLead')
    for user in datetl:
        user.total_days = (date.today() - user.date_of_joining).days if user.date_of_joining else 0
    return render(request, "md/tl_list.html", {"datetl": datetl})

from datetime import date
from django.shortcuts import render
from .models import User

def emp_list(request):
    dataemp = User.objects.filter(role__iexact='Employee')
    for user in dataemp:
        user.total_days = (date.today() - user.date_of_joining).days if user.date_of_joining else 0
    return render(request, "md/emp.html", {"dataemp": dataemp})

from django.shortcuts import get_object_or_404, redirect

def delete_hr(request, id):
    user = get_object_or_404(User, id=id)
    user.delete()
    return redirect('hr_list')   # make sure this URL name exists

def delete_mr(request, id):
    user = get_object_or_404(User, id=id)
    user.delete()
    return redirect('mr_list') 

from django.shortcuts import render, redirect
from .models import User

def hr_list(request):
    data = User.objects.all()
    return render(request, 'hr_list.html', {'data': data})


def update_hr(request):
    if request.method == "POST":
        id = request.POST.get('id')
        hr = User.objects.get(id=id)
        
        hr.username= request.POST.get('username')
        hr.email= request.POST.get('email')
        hr.role = request.POST.get('role')
        hr.address = request.POST.get('address')
        hr.phone = request.POST.get('phone')
        hr.salary = request.POST.get('salary')
        hr.gender = request.POST.get('gender')

        # ✅ FIX HERE
        dob = request.POST.get('date_of_birth')
        join = request.POST.get('date_of_joining')

        hr.date_of_birth = dob if dob else None
        hr.date_of_joining = join if join else None

        hr.status = request.POST.get('status')
        hr.experience_years = request.POST.get('experience_years')
        hr.previous_company = request.POST.get('previous_company')
        hr.department = request.POST.get('department')
        

        # ✅ FIXED IMAGE UPLOAD
        if request.FILES.get('profile_pic'):
            hr.profile_pic = request.FILES['profile_pic']

        hr.save()
        return redirect('hr_list')
    
def update_mr(request):
    if request.method == "POST":
        id = request.POST.get('id')
        mr = User.objects.get(id=id)

        # ✅ BASIC FIELDS
        mr.username = request.POST.get('username') or mr.username
        mr.email = request.POST.get('email') or mr.email

        # ⚠️ ROLE FIX (avoid null error)
        mr.role = request.POST.get('role') or "Manager"

        mr.address = request.POST.get('address')
        mr.phone = request.POST.get('phone')
        mr.salary = request.POST.get('salary')
        mr.gender = request.POST.get('gender')

        # ✅ DATE FIX (avoid empty string error)
        dob = request.POST.get('date_of_birth')
        join = request.POST.get('date_of_joining')

        mr.date_of_birth = dob if dob else None
        mr.date_of_joining = join if join else None

        mr.status = request.POST.get('status')
        mr.experience_years = request.POST.get('experience_years')
        mr.previous_company = request.POST.get('previous_company')
        mr.department = request.POST.get('department')

        # ✅ IMAGE UPLOAD FIX
        if request.FILES.get('profile_pic'):
            mr.profile_pic = request.FILES['profile_pic']

        mr.save()

        return redirect('mr_list')


def update_tl(request):
    if request.method == "POST":
        id = request.POST.get('id')
        tl = User.objects.get(id=id)

        # ✅ BASIC FIELDS
        tl.username = request.POST.get('username') or tl.username
        tl.email = request.POST.get('email') or tl.email

        # ⚠️ ROLE FIX (avoid null error)
        tl.role = request.POST.get('role') or "TeamLead"

        tl.address = request.POST.get('address')
        tl.phone = request.POST.get('phone')
        tl.salary = request.POST.get('salary')
        tl.gender = request.POST.get('gender')

        # ✅ DATE FIX (avoid empty string error)
        dob = request.POST.get('date_of_birth')
        join = request.POST.get('date_of_joining')

        tl.date_of_birth = dob if dob else None
        tl.date_of_joining = join if join else None

        tl.status = request.POST.get('status')
        tl.experience_years = request.POST.get('experience_years')
        tl.previous_company = request.POST.get('previous_company')
        tl.department = request.POST.get('department')
        tl.team_name = request.POST.get('team_name')

        # ✅ IMAGE UPLOAD FIX
        if request.FILES.get('profile_pic'):
            tl.profile_pic = request.FILES['profile_pic']

        tl.save()

        return redirect('tl_list')


def update_employee(request):
    if request.method == "POST":
        id = request.POST.get('id')
        emp = User.objects.get(id=id)

        # ✅ BASIC FIELDS
        emp.username = request.POST.get('username') or emp.username
        emp.email = request.POST.get('email') or emp.email

        # ⚠️ ROLE FIX (avoid null error)
        emp.role = request.POST.get('role') or "Employee"

        emp.address = request.POST.get('address')
        emp.phone = request.POST.get('phone')
        emp.salary = request.POST.get('salary')
        emp.gender = request.POST.get('gender')

        # ✅ DATE FIX (avoid empty string error)
        dob = request.POST.get('date_of_birth')
        join = request.POST.get('date_of_joining')

        emp.date_of_birth = dob if dob else None
        emp.date_of_joining = join if join else None

        emp.status = request.POST.get('status')
        emp.experience_years = request.POST.get('experience_years')
        emp.previous_company = request.POST.get('previous_company')
        emp.department = request.POST.get('department')
        emp.team_name = request.POST.get('team_name')

        # ✅ IMAGE UPLOAD FIX
        if request.FILES.get('profile_pic'):
            emp.profile_pic = request.FILES['profile_pic']

        emp.save()

        return redirect('emp_list')
    
from django.shortcuts import get_object_or_404, redirect
from .models import User

def delete_mr(request, id):
    if request.method == "POST":
        mr = get_object_or_404(User, id=id)
        mr.delete()
    return redirect('mr_list')

def delete_tl(request, id):
    if request.method == "POST":
        tl = get_object_or_404(User, id=id)
        tl.delete()
    return redirect('tl_list')

def delete_employee(request, id):
    if request.method == "POST":
        emp = get_object_or_404(User, id=id)
        emp.delete()
    return redirect('emp_list')



# Project Status Overview---------md---------------

from django.shortcuts import render
from .models import Project

def project_dashboard(request):
    completed_projects = list(
        Project.objects.filter(status='Completed').values_list('name', flat=True)
    )

    ongoing_projects = list(
        Project.objects.filter(status='Assigned to Team').values_list('name', flat=True)
    )

    pending_projects = list(
        Project.objects.filter(status='Pending').values_list('name', flat=True)
    )

    context = {
        "completed_count": len(completed_projects),
        "ongoing_count": len(ongoing_projects),
        "pending_count": len(pending_projects),

        "completed_projects": json.dumps(completed_projects),
        "ongoing_projects": json.dumps(ongoing_projects),
        "pending_projects": json.dumps(pending_projects),
    }

    return render(request, "md/project_dashboard.html", context)


def project_all(request):
    project_name = request.GET.get('name')  # get project name

    if project_name:
        projects = Project.objects.filter(name=project_name)
    else:
        projects = Project.objects.all()

    return render(request, "md/project_list.html", {"projects": projects})


# --------------------send_messages------------



def a_all(request):
    total_hr=User.objects.filter(Role='HR').count()
    total_mr=User.objects.filter(Role='MR').count()
    total_tl=User.objects.filter(Role='TL').count()
    total_emp=User.objects.filter(Role='EMP').count()
   
    total_project=Project.objects.all().count()


    context={
          "total_hr":total_hr,
          "total_mr":total_mr,
          "total_tl":total_tl,
          "total_emp":total_emp,
        
          'total_project': total_project,
          
    }

    return render(request,"md/a.html",context)


from django.contrib.auth import logout
from django.shortcuts import redirect

def user_logout(request):
    logout(request)
    return redirect('viewlogins')



from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import DailyReport
from hr.models import Project
from .forms import DailyReportForm

# --------------------------
# GET: View all reports
# --------------------------
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import DailyReport
from hr.models import Project
from .forms import DailyReportForm

# ----------------------------
# LIST: Show reports based on hierarchy
# ----------------------------
@login_required
def daily_report_list(request):
    user = request.user

    # ---------------- REPORTS BY ROLE ----------------
    if user.role == "TeamLead":
        reports = DailyReport.objects.filter(user__role="Employee")

    elif user.role == "Manager":
        reports = DailyReport.objects.filter(user__role="TeamLead")

    elif user.role in ["HR", "MD"]:
        reports = DailyReport.objects.all()

    else:
        reports = DailyReport.objects.none()

    # ---------------- DATE / MONTH FILTER ----------------
    selected_date = request.GET.get("date")
    selected_month = request.GET.get("month")

    if selected_date:
        reports = reports.filter(report_date=selected_date)

    if selected_month:
        year, month = selected_month.split("-")
        reports = reports.filter(
            report_date__year=year,
            report_date__month=month
        )

    reports = reports.select_related("user", "project").order_by("-report_date")

    # ---------------- BACK BUTTON & BASE TEMPLATE LOGIC ----------------
    role_to_base = {
        "MD": ("md/md_dashboard.html", "md_dashboard"),
        "HR": ("base.html", "home"),
        "Manager": ("main/mddash.html", "index"),
        "TeamLead": ("tl/side_bar.html", "tl_dashboard"),
        "Employee": ("employee/employee_base.html", "employee_dashboard"),
    }

    base_template, back_url = role_to_base.get(
        user.role,
        ("employee/employee_base.html", "employee_dashboard")
    )

    return render(request, "tl/daily_report_list.html", {
        "reports": reports,
        "selected_date": selected_date,
        "selected_month": selected_month,
        "back_url": back_url,
        "base_template": base_template,
    })



# ----------------------------
# SUBMIT: Submit a new report
# ----------------------------
@login_required
def daily_report_submit(request):
    user = request.user
    role = user.role
    roles = list(user.groups.values_list("name", flat=True))

    # Assign projects the user can submit reports for
    if role == "Employee" or "Employee" in roles:
        projects = Project.objects.filter(assigned_team__members=user).distinct()
        recipient_role = "TeamLead"
    elif role == "TeamLead" or "TeamLead" in roles:
        projects = Project.objects.filter(assigned_team__lead=user).distinct()
        recipient_role = "Manager"
    elif role == "Manager" or "Manager" in roles:
        projects = Project.objects.filter(assigned_manager=user).distinct()
        recipient_role = "HR"
    else:
        projects = Project.objects.none()
        recipient_role = "Unknown"

    if request.method == "POST":
        form = DailyReportForm(request.POST, request.FILES)
        form.fields['project'].queryset = projects
        if form.is_valid():
            report = form.save(commit=False)
            report.user = user
            report.recipient_role = recipient_role
            report.save()
            return redirect("daily_report_list")
    else:
        form = DailyReportForm()
        form.fields['project'].queryset = projects

    role_to_base = {
        "MD": ("md/md_dashboard.html", "md_dashboard"),
        "HR": ("base.html", "home"),
        "Manager": ("main/mddash.html", "index"),
        "TeamLead": ("tl/side_bar.html", "tl_dashboard"),
        "Employee": ("employee/employee_base.html", "employee_dashboard"),
    }

    base_template, back_url = role_to_base.get(
        user.role,
        ("employee/employee_base.html", "employee_dashboard")
    )

    return render(request, "tl/daily_report_submit.html", {
        "form": form,
        "projects": projects,
        "back_url": back_url,
        "base_template": base_template,
    })


  
 
# -----------all---apply_leaves--------------------

@login_required
def dashboard(request):

    role = request.user.role

    if role == "Employee":
        return render(request, "employee/employee_dase.html")

    elif role == "TeamLead":
        return render(request, "tl/side_bar.html")

    elif role == "Manager":
        return render(request, "md/mddash.html")

    elif role == "HR":
        return render(request, "base.html")

    elif role == "MD":
        return render(request, "md/md_header.html")

        

def leave_dashboard(request):
    role = request.user.role

    if role == "Employee":
        leaves = Leave.objects.filter(user=request.user).order_by('-id')
        template = "employee/employee_leave_dashboard.html"

    elif role == "TeamLead":
        leaves = Leave.objects.filter(approved_tl=False).order_by('-id')
        template = "tl/tl_leave_dashboard.html"

    elif role == "Manager":
        leaves = Leave.objects.filter(approved_tl=True, approved_manager=False).order_by('-id')
        template = "manager/manager_leave_dashboard.html"

    elif role == "HR":
        leaves = Leave.objects.filter(approved_manager=True, approved_hr=False).order_by('-id')
        template = "hr/hr_leave_dashboard.html"

    elif role == "MD":
        leaves = Leave.objects.filter(approved_hr=True, approved_md=False).order_by('-id')
        template = "md/md_leave_dashboard.html"

    return render(request, template, {"leaves": leaves})
# views.py
# hr/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Leave

@login_required
def apply_leave(request):
    if request.method == "POST":
        role = request.user.role

        leave = Leave.objects.create(
            user=request.user,
            from_date=request.POST['from_date'],
            to_date=request.POST['to_date'],
            reason=request.POST['reason'],
            leave_type=request.POST.get('leave_type', 'Paid'),
        )


        if role == "Employee":
            leave.status = "Pending TeamLead Approval"

        elif role == "TeamLead":
            leave.approved_tl = True
            leave.status = "Pending Manager Approval"

        elif role == "Manager":
            leave.approved_tl = True
            leave.approved_manager = True
            leave.status = "Pending HR Approval"

        elif role == "HR":
            leave.approved_tl = True
            leave.approved_manager = True
            leave.approved_hr = True
            leave.status = "Pending MD Approval"

        elif role == "MD":
            leave.approved_tl = True
            leave.approved_manager = True
            leave.approved_hr = True
            leave.approved_md = True
            leave.status = "Approved"

        leave.save()
        return redirect('leave_status')

    return render(request, 'apply_leave.html')



@login_required
def leave_dashboard(request):
    role = request.user.role

    if role == "TeamLead":
        leaves = Leave.objects.filter(
            approved_tl=False,
            status="Pending TeamLead Approval"
        ).order_by('-id')
        pending_count = leaves.count()

    elif role == "Manager":
        leaves = Leave.objects.filter(
            approved_tl=True,
            approved_manager=False,
            status="Pending Manager Approval"
        ).order_by('-id')
        pending_count = leaves.count()

    elif role == "HR":
        leaves = Leave.objects.filter(
            approved_manager=True,
            approved_hr=False,
            status="Pending HR Approval"
        ).order_by('-id')
        pending_count = leaves.count()

    elif role == "MD":
        leaves = Leave.objects.filter(
            approved_hr=True,
            approved_md=False,
            status="Pending MD Approval"
        ).order_by('-id')
        pending_count = leaves.count()

    else:
        leaves = Leave.objects.none()

        # Map role → header template
  
   

    return render(request, 'md/leave_dashboard.html', {
        'leaves': leaves,
        'countl': leaves.count()
    })
 


from django.shortcuts import get_object_or_404

@login_required
def approve_leave(request, leave_id):
    leave = get_object_or_404(Leave, id=leave_id)
    role = request.user.role

    if leave.status in ["Rejected", "Approved"]:
        return redirect('leave_dashboard')

    if role == "TeamLead":
        leave.approved_tl = True
        leave.status = "Pending Manager Approval"

    elif role == "Manager":
        leave.approved_manager = True
        leave.status = "Pending HR Approval"

    elif role == "HR":
        leave.approved_hr = True
        leave.status = "Pending MD Approval"

    elif role == "MD":
        leave.approved_md = True
        leave.status = "Approved"

    leave.save()
    return redirect('leave_dashboard')


@login_required
def reject_leave(request, leave_id):
    leave = get_object_or_404(Leave, id=leave_id)
    leave.status = "Rejected"
    leave.save()
    return redirect('leave_dashboard')
 
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta
from .models import Leave


@login_required
def leave_status(request):

    # Calculate 6 months ago date
    six_months_ago = timezone.now().date() - timedelta(days=180)

    # Delete leaves older than 6 months
    Leave.objects.filter(from_date__lt=six_months_ago).delete()

    # Show only current user's leaves
    leaves = Leave.objects.filter(user=request.user)

    context = {
        "hr_leaves": leaves
    }

    return render(request, "leave_status.html", context)
 

def all_leaves(request):
    all_leaves =Leave.objects.all()
    
    return render(request,"md/all_leaves.html",{"all_leaves":all_leaves})


from .models import User

def tl_leaves(request):
    user = request.user   # Logged in TL
    
    tl = Leave.objects.filter( approved_tl=True).order_by('-id')

    return render(request, "tl/tl_leaves_approved.html", {"tl": tl})


from django.shortcuts import render
from .models import Leave

def manager_approved_leaves(request):
    # Only leaves approved by manager
    leaves = Leave.objects.filter(approved_manager=True)

    return render(request, "main/manager_approved_leaves.html", {
        "leaves": leaves
    })

def hr_approved_leaves(request):
    leaves = Leave.objects.filter(
        approved_tl=True,
        approved_manager=True,
        approved_hr=False,
        
    )

    return render(request, "hr_approved_leaves.html", {
        "leaves": leaves
    })
    
from django.shortcuts import render
from .models import User   # make sure this is correct model import

# all_total_list-------
def all_member_list(request):
    total_list = User.objects.all()
    return render(request, "md/all_member.html", {"total_list": total_list})
    #________________________________________#

    #************employe views*****************#

    #_______________________________________#



    # employee/views.py
from django.contrib.auth import get_user_model, login
from django.contrib import messages
from django.shortcuts import render, redirect

User = get_user_model()

def employee_login(request):
    return redirect('viewlogins')



    # employee/views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Team, Project, Task, Payslip

@login_required
def employee_dashboard(request):
    user = request.user

    if user.role != "Employee":
        return redirect("employee_login")

    team = Team.objects.filter(members=user).select_related("lead").prefetch_related("members").first()

    team_lead = team.lead if team else None

    team_members = team.members.exclude(id=user.id) if team else []

    tasks = Task.objects.filter(members=user)

    # Fetch latest published payslip
    latest_payslip = Payslip.objects.filter(employee=user, is_published=True).order_by('-year', '-month').first()

    context = {
        "team_lead": team_lead,
        "team_members": team_members,
        "tasks": tasks,
        "latest_payslip": latest_payslip,
    }

    return render(request, "employee/emp_dashboard.html", context)


from django.contrib.auth import logout
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required

@login_required
def employee_logout(request):
    logout(request)
    return redirect("viewlogins")

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth import get_user_model
from .models import Team, Project
from datetime import datetime
from django.utils import timezone


User = get_user_model()


@login_required
def employee_profile_self(request):

    if request.user.role != "Employee":
        return redirect("home")

    employee = request.user

    teams = Team.objects.filter(members=employee)
    employee_projects = Project.objects.filter(assigned_team__in=teams)

    month = request.GET.get("month")

    if month:
        year, month = month.split("-")
        attendance = calculate_attendance(employee, int(year), int(month))
    else:
        now = timezone.now()
        attendance = calculate_attendance(employee, now.year, now.month)

    # Build leave summary safely
    leave_summary = None
    try:
        from django.db.models import Q
        now_year = timezone.now().year
        # Try common leave model names
        try:
            from .models import LeaveRequest as LeaveModel
        except ImportError:
            try:
                from .models import Leave as LeaveModel
            except ImportError:
                LeaveModel = None

        if LeaveModel is not None:
            emp_leaves = LeaveModel.objects.filter(employee=employee, applied_on__year=now_year)
            leave_summary = {
                "approved": emp_leaves.filter(status="Approved").count(),
                "pending": emp_leaves.filter(status="Pending").count(),
                "rejected": emp_leaves.filter(status="Rejected").count(),
            }
    except Exception:
        pass

    return render(request, "employee/profile.html", {
        "employee": employee,
        "employee_projects": employee_projects,
        "attendance": attendance,
        "reporting_manager": employee.reporting_manager,
        "leave_summary": leave_summary,
    })

@login_required
def edit_employee_self(request):
    if request.user.role != "Employee":
        return redirect("home")

    employee = request.user

    if request.method == "POST":
        employee.first_name = request.POST.get("first_name")
        employee.last_name = request.POST.get("last_name")
        employee.phone = request.POST.get("phone")

        # Profile Image Removal
        if request.POST.get("remove_profile_pic") == "true":
            if employee.profile_pic:
                employee.profile_pic.delete(save=False)
                employee.profile_pic = None

        # Profile Image Update
        elif request.FILES.get("profile_pic"):
            employee.profile_pic = request.FILES.get("profile_pic")

        # Password Update
        new_password = request.POST.get("new_password")
        if new_password:
            employee.set_password(new_password)
            employee.save()
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, employee)
        else:
            employee.save()

        messages.success(request, "Profile updated successfully")
        return redirect("employee_profile")

    return render(request, "employee/emp_edit_profile.html", {
        "employee": employee
    })

@login_required
def employee_all_tasks(request):
    user = request.user

    tasks = Task.objects.filter(members=user)

    context = {
        "pending_tasks": tasks.filter(status="Pending"),
        "completed_tasks": tasks.filter(status="Completed"),
        "submitted_tasks": tasks.filter(status="Submitted"),
    }
    return render(request, "employee/all_tasks.html", context)


@login_required
def submit_task(request, task_id):
    task = Task.objects.get(id=task_id, members=request.user)

    if request.method == "POST":
        task.status = "Submitted"

        if request.FILES.get("file"):
            task.file = request.FILES["file"]

        task.save()
        return redirect("employee_all_tasks")

    return render(request, "employee/submit_task.html", {"task": task})

@login_required
def complete_task(request, task_id):
    task = Task.objects.get(id=task_id)

    if request.user.role in ["HR", "Manager", "TeamLead"]:
        task.status = "Completed"
        task.save()

    return redirect("view_tasks")  # change to your HR task list page


#********** aatedence views***************

import base64
from django.core.files.base import ContentFile
import uuid

def handle_photo_upload(photo_data):
    if not photo_data:
        return None

    try:
        header, data = photo_data.split(',', 1)  # ✅ SAFE
    except ValueError:
        return None  # invalid image format

    file_ext = header.split('/')[-1].split(';')[0]
    file_name = f"{uuid.uuid4()}.{file_ext}"

    return ContentFile(base64.b64decode(data), name=file_name)



from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Attendance

@login_required
def emp_check_in(request):
    user = request.user
    today = timezone.localdate()

    # Get or create attendance for today
    attendance = Attendance.objects.get_or_create(
        user=request.user,
        date=today
    )[0]   

    if request.method == 'POST':
        action = request.POST.get("action")

    # ✅ Only check in if user clicked CHECK IN
    if action == "checkin" and not attendance.check_in_time:
        attendance.check_in_time = timezone.now()

        photo_data = request.POST.get("photo")
        if photo_data:
            attendance.check_in_photo = handle_photo_upload(photo_data)

        attendance.save()
   
    role_to_base = {
        "MD": "md_dashboard",
        "HR": "home",
        "Manager": "index",
        "TeamLead": "tl_dashboard",
        "Employee": "employee_dashboard",
    }

    back_url = role_to_base.get(user.role, "home")

    # Pass attendance to template
    context = {
        'attendance': attendance,
        "back_url": back_url,
    }
    return render(request, 'attedance/attendance_dashboard.html', context)


from django.utils.timezone import now
from datetime import date

def emp_check_out(request):
    user = request.user
    today = date.today()
    
    try:
        attendance = Attendance.objects.get(user=user, date=today)
        if not attendance.check_out_time:
            photo_data = request.POST.get('photo')
            attendance.check_out_time = now()
            attendance.check_out_photo = handle_photo_upload(photo_data)
            attendance.calculate_status()
            attendance.save()
    except Attendance.DoesNotExist:
        attendance = None  # or handle differently
    role_to_base = {
        "MD": "md_dashboard",
        "HR": "home",
        "Manager": "index",
        "TeamLead": "tl_dashboard",
        "Employee": "employee_dashboard",
    }

    back_url = role_to_base.get(user.role, "home")
    
    context = {
        'attendance': attendance,
          "back_url": back_url,
    }
    return render(request, 'attedance/attendance_dashboard.html', context)


from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import get_object_or_404, render, redirect
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from datetime import date, datetime
import calendar
import io
from django.http import HttpResponse
from openpyxl import Workbook
from collections import Counter
from django.contrib.auth import get_user_model

User = get_user_model()

def can_view_attendance(user, target_user):
    if user == target_user:
        return True
    if user.role in ['MD', 'HR']:
        return True
    if user.role == 'Manager':
        if target_user.reporting_manager == user:
            return True
        if target_user.reporting_manager and target_user.reporting_manager.reporting_manager == user:
            return True
        from .models import Team
        if Team.objects.filter(lead__reporting_manager=user, members=target_user).exists():
            return True
    if user.role == 'TeamLead':
        if target_user.reporting_manager == user:
            return True
        from .models import Team
        if Team.objects.filter(lead=user, members=target_user).exists():
            return True
    return False

def get_monthly_calendar_data(user, year, month):
    from datetime import date
    from .models import Attendance, Leave, Holiday
    
    first_weekday, num_days = calendar.monthrange(year, month)
    today = date.today()
    
    # Pre-fetch records
    attendance_records = Attendance.objects.filter(
        user=user,
        date__year=year,
        date__month=month
    )
    attendance_dict = {r.date: r for r in attendance_records}
    
    from django.db.models import Q
    holidays = Holiday.objects.filter(
        status='Approved',
        date__year=year,
        date__month=month
    ).filter(
        Q(department__isnull=True) | Q(department='') | Q(department=user.department)
    )
    holiday_dates = set(h.date for h in holidays)
    holiday_dict = {h.date: h.name for h in holidays}
    
    first_date = date(year, month, 1)
    last_date = date(year, month, num_days)
    approved_leaves = Leave.objects.filter(
        user=user,
        status='Approved',
        from_date__lte=last_date,
        to_date__gte=first_date
    )
    
    def get_leave_status(d):
        for lv in approved_leaves:
            if lv.from_date <= d <= lv.to_date:
                return lv.leave_type  # 'Paid' or 'Unpaid'
        return None

    days_data = []
    
    # Sunday is start of week in calendar grid.
    padding = (first_weekday + 1) % 7
    
    stats = {
        'total_working_days': 0,
        'present': 0,
        'absent': 0,
        'half_day': 0,
        'wfh': 0,
        'paid_leave': 0,
        'unpaid_leave': 0,
        'weekly_off': 0,
        'holiday': 0,
        'late': 0,
        'early': 0,
        'percentage': 0.0
    }
    
    for d_num in range(1, num_days + 1):
        curr_date = date(year, month, d_num)
        is_future = curr_date > today
        is_joined = True
        if user.date_of_joining and curr_date < user.date_of_joining:
            is_joined = False
            
        status = None
        color_class = 'calendar-empty'
        check_in = '--'
        check_out = '--'
        working_hours = 0
        remarks = ''
        overtime = 0
        
        # 1. Week Off (Sunday)
        if curr_date.weekday() == 6:
            status = 'Week Off'
            color_class = 'calendar-weekly-off'
            stats['weekly_off'] += 1
        # 2. Holiday
        elif curr_date in holiday_dates:
            status = 'Holiday'
            color_class = 'calendar-holiday'
            remarks = holiday_dict.get(curr_date, 'Public Holiday')
            stats['holiday'] += 1
        # 3. Before Joining
        elif not is_joined:
            status = 'Unpaid Leave'
            color_class = 'calendar-unpaid-leave'
            remarks = 'Before Date of Joining'
            stats['unpaid_leave'] += 1
        # 4. Future
        elif is_future:
            status = '--'
            color_class = 'calendar-future'
        # 5. Check attendance record or leave
        else:
            att = attendance_dict.get(curr_date)
            leave_type = get_leave_status(curr_date)
            
            if att:
                status_str = att.status or ''
                status_lower = status_str.lower()
                
                check_in = att.check_in_time.strftime('%I:%M %p') if att.check_in_time else '--'
                check_out = att.check_out_time.strftime('%I:%M %p') if att.check_out_time else '--'
                working_hours = att.total_hours or 0
                if working_hours > 9:
                    overtime = round(working_hours - 9, 2)
                    
                if att.is_late:
                    stats['late'] += 1
                if att.left_early:
                    stats['early'] += 1
                    
                if 'present' in status_lower:
                    status = 'Present'
                    color_class = 'calendar-present'
                    stats['present'] += 1
                    stats['total_working_days'] += 1
                elif 'half' in status_lower:
                    status = 'Half Day'
                    color_class = 'calendar-half-day'
                    stats['half_day'] += 1
                    stats['total_working_days'] += 1
                elif 'work from home' in status_lower or 'wfh' in status_lower:
                    status = 'Work From Home'
                    color_class = 'calendar-wfh'
                    stats['wfh'] += 1
                    stats['total_working_days'] += 1
                elif 'absent' in status_lower:
                    if leave_type:
                        if leave_type == 'Paid':
                            status = 'Paid Leave'
                            color_class = 'calendar-paid-leave'
                            stats['paid_leave'] += 1
                        else:
                            status = 'Unpaid Leave'
                            color_class = 'calendar-unpaid-leave'
                            stats['unpaid_leave'] += 1
                    else:
                        status = 'Absent'
                        color_class = 'calendar-absent'
                        stats['absent'] += 1
                        stats['total_working_days'] += 1
                else:
                    status = 'Present'
                    color_class = 'calendar-present'
                    stats['present'] += 1
                    stats['total_working_days'] += 1
            else:
                # No attendance record
                if leave_type:
                    if leave_type == 'Paid':
                        status = 'Paid Leave'
                        color_class = 'calendar-paid-leave'
                        stats['paid_leave'] += 1
                    else:
                        status = 'Unpaid Leave'
                        color_class = 'calendar-unpaid-leave'
                        stats['unpaid_leave'] += 1
                else:
                    status = 'Absent'
                    color_class = 'calendar-absent'
                    stats['absent'] += 1
                    stats['total_working_days'] += 1
                    
        days_data.append({
            'day_num': d_num,
            'date': curr_date.strftime('%Y-%m-%d'),
            'date_formatted': curr_date.strftime('%A, %b %d, %Y'),
            'status': status,
            'color_class': color_class,
            'check_in': check_in,
            'check_out': check_out,
            'working_hours': working_hours,
            'overtime': overtime,
            'remarks': remarks,
        })
        
    present_weight = stats['present'] + stats['wfh'] + (stats['half_day'] * 0.5) + stats['paid_leave']
    total_expected = stats['total_working_days']
    if total_expected > 0:
        stats['percentage'] = round((present_weight / total_expected) * 100, 2)
    else:
        stats['percentage'] = 0.0
        
    return days_data, padding, stats

@login_required
def admin_attendance_view(request):
    user = request.user
    role = user.role
    
    # Check if we are viewing a specific employee's detailed calendar/dashboard
    target_user_id = request.GET.get('user_id')
    
    if target_user_id or role == 'Employee':
        if role == 'Employee':
            target_user = user
        else:
            if target_user_id == 'me':
                target_user = user
            else:
                target_user = get_object_or_404(User, id=target_user_id)
                
            # Authorization Check
            if not can_view_attendance(user, target_user):
                raise PermissionDenied("You are not authorized to view this employee's attendance.")
        
        # Get selected month and year
        today = date.today()
        selected_month = request.GET.get('month')
        selected_year = request.GET.get('year')
        
        if not selected_month or not selected_year:
            selected_month = today.month
            selected_year = today.year
        else:
            selected_month = int(selected_month)
            selected_year = int(selected_year)
            
        # Calculate monthly attendance data for the calendar and stats
        days_data, padding, stats = get_monthly_calendar_data(target_user, selected_year, selected_month)
        
        # Today's attendance status
        today_record = Attendance.objects.filter(user=target_user, date=today).first()
        
        # Get history of attendance logs for the target user (with search and date filters)
        history_records = Attendance.objects.filter(user=target_user)
        
        # Apply search or date filter if requested
        q_month = request.GET.get('q_month')
        q_year = request.GET.get('q_year')
        if q_month:
            history_records = history_records.filter(date__month=q_month)
        if q_year:
            history_records = history_records.filter(date__year=q_year)
            
        history_records = history_records.order_by('-date')
        
        # Base template determination based on user's role
        role_to_base = {
            "MD": "md/md_base.html",
            "HR": "base.html",
            "Manager": "main/mddash.html",
            "TeamLead": "tl/side_bar.html",
            "Employee": "employee/employee_base.html",
        }
        base_template = role_to_base.get(role, "employee/employee_base.html")
        
        context = {
            'target_user': target_user,
            'days_data': days_data,
            'padding_range': range(padding),
            'stats': stats,
            'today_record': today_record,
            'history_records': history_records,
            'selected_month': selected_month,
            'selected_year': selected_year,
            'base_template': base_template,
            'q_month': q_month,
            'q_year': q_year,
        }
        return render(request, 'attedance/attendance_employee.html', context)
        
    else:
        # RENDER THE REGISTRY LIST VIEW (for TeamLead, Manager, HR, MD)
        query = request.GET.get('q', '')
        selected_date = request.GET.get('date', '')
        
        # Query set based on roles
        if role == 'TeamLead':
            records = Attendance.objects.filter(
                Q(user=user) | Q(user__reporting_manager=user)
            ).select_related('user')
            subordinates = User.objects.filter(reporting_manager=user)
        elif role == 'Manager':
            subordinates = User.objects.filter(
                Q(reporting_manager=user) |
                Q(reporting_manager__reporting_manager=user) |
                Q(teams__lead__reporting_manager=user)
            ).distinct()
            records = Attendance.objects.filter(
                Q(user=user) | Q(user__in=subordinates)
            ).select_related('user')
        elif role in ['HR', 'MD']:
            records = Attendance.objects.select_related('user')
            subordinates = User.objects.all()
        else:
            records = Attendance.objects.none()
            subordinates = User.objects.none()
            
        # Apply search query
        if query:
            records = records.filter(
                Q(user__username__icontains=query) |
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__emp_id__icontains=query)
            )
            
        # Apply date filter
        if selected_date:
            records = records.filter(date=selected_date)
            
        records = records.order_by('-date')
        
        # Determine base template
        role_to_base = {
            "MD": "md/md_base.html",
            "HR": "base.html",
            "Manager": "main/mddash.html",
            "TeamLead": "tl/side_bar.html",
        }
        base_template = role_to_base.get(role, "base.html")
        
        context = {
            'records': records,
            'subordinates': subordinates.order_by('first_name'),
            'base_template': base_template,
            'query': query,
            'selected_date': selected_date,
        }
        # HR and MD get a dedicated template that extends base.html (gives them the full sidebar)
        if role == 'HR':
            return render(request, 'attedance/hr_attendance_list.html', context)
        return render(request, 'attedance/attendance_list.html', context)

@login_required
def export_attendance_excel(request):
    user = request.user
    role = user.role
    
    # Get filters
    query = request.GET.get('q', '')
    selected_date = request.GET.get('date', '')
    target_user_id = request.GET.get('user_id')
    
    if target_user_id or role == 'Employee':
        if role == 'Employee':
            target_user = user
        else:
            if target_user_id == 'me':
                target_user = user
            else:
                target_user = get_object_or_404(User, id=target_user_id)
            if not can_view_attendance(user, target_user):
                raise PermissionDenied("You are not authorized to view this employee's attendance.")
        attendance_qs = Attendance.objects.filter(user=target_user).select_related('user').order_by('date')
    else:
        # Export for multiple employees
        if role == 'TeamLead':
            attendance_qs = Attendance.objects.filter(
                Q(user=user) | Q(user__reporting_manager=user)
            ).select_related('user')
        elif role == 'Manager':
            subordinates = User.objects.filter(
                Q(reporting_manager=user) |
                Q(reporting_manager__reporting_manager=user) |
                Q(teams__lead__reporting_manager=user)
            ).distinct()
            attendance_qs = Attendance.objects.filter(
                Q(user=user) | Q(user__in=subordinates)
            ).select_related('user')
        elif role in ['HR', 'MD']:
            attendance_qs = Attendance.objects.select_related('user').all()
        else:
            attendance_qs = Attendance.objects.none()
            
        if query:
            attendance_qs = attendance_qs.filter(
                Q(user__username__icontains=query) |
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__emp_id__icontains=query)
            )
        if selected_date:
            attendance_qs = attendance_qs.filter(date=selected_date)
            
        attendance_qs = attendance_qs.order_by('date')

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Write the header row
    ws.append(['Employee ID', 'Employee Name', 'Date', 'Check-in', 'Check-out', 'Status', 'Total Hours'])

    # Write each attendance record as a row in the Excel sheet
    for att in attendance_qs:
        ws.append([
            att.user.emp_id or '',
            att.user.get_full_name() or att.user.username,
            att.date.strftime('%Y-%m-%d') if att.date else '',
            att.check_in_time.strftime('%H:%M:%S') if att.check_in_time else '',
            att.check_out_time.strftime('%H:%M:%S') if att.check_out_time else '',
            att.status or '',
            att.total_hours or 0,
        ])

    # Save workbook to an in-memory bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Prepare the HTTP response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'

    return response

@login_required
def hr_attendance_edit(request, pk):
    """HR-only: Edit a single attendance record via AJAX POST."""
    if request.user.role != 'HR':
        return JsonResponse({'error': 'Permission denied. HR access only.'}, status=403)
    
    attendance = get_object_or_404(Attendance, pk=pk)
    
    if request.method == 'GET':
        # Return attendance data as JSON for the edit modal
        ci = attendance.check_in_time
        co = attendance.check_out_time
        return JsonResponse({
            'id': attendance.pk,
            'employee_name': attendance.user.get_full_name() or attendance.user.username,
            'employee_id': attendance.user.emp_id or '',
            'date': attendance.date.strftime('%Y-%m-%d'),
            'check_in_time': ci.strftime('%H:%M') if ci else '',
            'check_out_time': co.strftime('%H:%M') if co else '',
            'total_hours': attendance.total_hours or 0,
            'status': attendance.status or '',
            'remarks': attendance.remarks or '',
        })
    
    if request.method == 'POST':
        import json as _json
        try:
            data = _json.loads(request.body)
        except Exception:
            data = request.POST
        
        # Update fields
        new_date = data.get('date', '')
        new_checkin = data.get('check_in_time', '')
        new_checkout = data.get('check_out_time', '')
        new_status = data.get('status', attendance.status)
        new_remarks = data.get('remarks', '')
        
        from datetime import datetime as _datetime, date as _date
        from django.utils import timezone as _tz
        
        if new_date:
            try:
                attendance.date = _datetime.strptime(new_date, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if new_checkin:
            try:
                naive_dt = _datetime.combine(attendance.date, _datetime.strptime(new_checkin, '%H:%M').time())
                attendance.check_in_time = _tz.make_aware(naive_dt) if _tz.is_naive(naive_dt) else naive_dt
            except ValueError:
                pass
        elif not new_checkin:
            attendance.check_in_time = None
        
        if new_checkout:
            try:
                naive_dt = _datetime.combine(attendance.date, _datetime.strptime(new_checkout, '%H:%M').time())
                attendance.check_out_time = _tz.make_aware(naive_dt) if _tz.is_naive(naive_dt) else naive_dt
            except ValueError:
                pass
        elif not new_checkout:
            attendance.check_out_time = None
        
        # Compute total hours if both times are present
        if attendance.check_in_time and attendance.check_out_time:
            delta = attendance.check_out_time - attendance.check_in_time
            attendance.total_hours = round(delta.total_seconds() / 3600, 2)
        
        # Set status directly (HR can override auto-calculation)
        if new_status:
            attendance.status = new_status
        
        attendance.remarks = new_remarks
        attendance.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Attendance record updated successfully.',
            'status': attendance.status,
            'total_hours': attendance.total_hours,
        })
    
    return JsonResponse({'error': 'Invalid request method.'}, status=405)


@login_required
def hr_attendance_bulk_update(request):
    """HR-only: Bulk update attendance status for multiple records."""
    if request.user.role != 'HR':
        return JsonResponse({'error': 'Permission denied. HR access only.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required.'}, status=405)
    
    import json as _json
    try:
        data = _json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON.'}, status=400)
    
    record_ids = data.get('ids', [])
    new_status = data.get('status', '')
    remarks = data.get('remarks', '')
    
    if not record_ids or not new_status:
        return JsonResponse({'error': 'ids and status are required.'}, status=400)
    
    allowed_statuses = ['Present', 'Absent', 'Half Day', 'Work From Home', 'Paid Leave', 'Unpaid Leave', 'Holiday', 'Week Off']
    if new_status not in allowed_statuses:
        return JsonResponse({'error': f'Invalid status. Allowed: {allowed_statuses}'}, status=400)
    
    updated = Attendance.objects.filter(pk__in=record_ids).update(status=new_status, remarks=remarks)
    
    return JsonResponse({'success': True, 'updated': updated, 'message': f'{updated} records updated to "{new_status}".'})


@login_required
def hr_attendance_import(request):
    """HR-only: Import attendance records from an uploaded Excel file."""
    if request.user.role != 'HR':
        return JsonResponse({'error': 'Permission denied. HR access only.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required.'}, status=405)
    
    excel_file = request.FILES.get('file')
    if not excel_file:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)
    
    try:
        from openpyxl import load_workbook as _load_wb
        from datetime import datetime as _datetime
        from django.utils import timezone as _tz
        
        wb = _load_wb(excel_file)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return JsonResponse({'error': 'Empty Excel file.'}, status=400)
        
        # Expected columns: Employee ID, Date (YYYY-MM-DD), Check-In (HH:MM), Check-Out (HH:MM), Status, Remarks
        header = [str(c).strip().lower() if c else '' for c in rows[0]]
        
        imported = 0
        errors = []
        
        for row_num, row in enumerate(rows[1:], start=2):
            try:
                emp_id = str(row[0]).strip() if row[0] else ''
                date_val = row[1]
                checkin_val = row[2]
                checkout_val = row[3]
                status_val = str(row[4]).strip() if row[4] else ''
                remarks_val = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                
                if not emp_id:
                    errors.append(f'Row {row_num}: Missing Employee ID')
                    continue
                
                try:
                    emp_user = User.objects.get(emp_id__iexact=emp_id)
                except User.DoesNotExist:
                    errors.append(f'Row {row_num}: Employee ID "{emp_id}" not found')
                    continue
                
                # Parse date
                if isinstance(date_val, _datetime):
                    att_date = date_val.date()
                elif isinstance(date_val, str):
                    att_date = _datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
                else:
                    from datetime import date as _date
                    att_date = date_val
                
                # Parse check-in time
                check_in_dt = None
                if checkin_val:
                    if isinstance(checkin_val, str) and checkin_val.strip():
                        ci_time = _datetime.strptime(checkin_val.strip(), '%H:%M').time()
                        naive = _datetime.combine(att_date, ci_time)
                        check_in_dt = _tz.make_aware(naive)
                
                # Parse check-out time
                check_out_dt = None
                if checkout_val:
                    if isinstance(checkout_val, str) and checkout_val.strip():
                        co_time = _datetime.strptime(checkout_val.strip(), '%H:%M').time()
                        naive = _datetime.combine(att_date, co_time)
                        check_out_dt = _tz.make_aware(naive)
                
                # Calculate hours
                total_hours = None
                if check_in_dt and check_out_dt:
                    total_hours = round((check_out_dt - check_in_dt).total_seconds() / 3600, 2)
                
                # Upsert attendance record
                att, created = Attendance.objects.update_or_create(
                    user=emp_user,
                    date=att_date,
                    defaults={
                        'check_in_time': check_in_dt,
                        'check_out_time': check_out_dt,
                        'total_hours': total_hours,
                        'status': status_val or ('Present' if total_hours and total_hours >= 9 else 'Absent'),
                        'remarks': remarks_val,
                    }
                )
                imported += 1
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')
        
        return JsonResponse({
            'success': True,
            'imported': imported,
            'errors': errors,
            'message': f'{imported} records imported successfully. {len(errors)} errors.'
        })
    
    except Exception as e:
        return JsonResponse({'error': f'Failed to process file: {str(e)}'}, status=500)


@login_required
def monthly_attendance_summary(request):
    user = request.user
    role = user.role
    
    if role == 'Employee':
        return redirect('attendance_list')

    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')

    if not selected_month or not selected_year:
        today = date.today()
        selected_month = today.month
        selected_year = today.year
    else:
        selected_month = int(selected_month)
        selected_year = int(selected_year)

    # Filter employees list based on user's role
    if role == 'TeamLead':
        employees = User.objects.filter(reporting_manager=user)
    elif role == 'Manager':
        employees = User.objects.filter(
            Q(reporting_manager=user) |
            Q(reporting_manager__reporting_manager=user) |
            Q(teams__lead__reporting_manager=user)
        ).distinct()
    elif role in ['HR', 'MD']:
        employees = User.objects.all()
    else:
        employees = User.objects.none()

    summary = []

    for emp in employees:
        records = Attendance.objects.filter(
            user=emp,
            date__month=selected_month,
            date__year=selected_year
        )

        status_counter = Counter(
            r.status.split(" ")[0]
            for r in records
            if r.status
        )

        late_count = sum('Late' in r.status for r in records if r.status)
        early_count = sum('Left Early' in r.status for r in records if r.status)

        summary.append({
            'employee': emp,
            'present': status_counter.get("Present", 0) + status_counter.get("Work", 0),  # Present + WFH
            'half_day': status_counter.get("Half", 0),
            'absent': status_counter.get("Absent", 0),
            'off': status_counter.get("Off", 0),
            'late': late_count,
            'early': early_count,
            'total_days': records.count()
        })

    # Base template determination based on user's role
    role_to_base = {
        "MD": "md/md_base.html",
        "HR": "base.html",
        "Manager": "main/mddash.html",
        "TeamLead": "tl/side_bar.html",
    }
    base_template = role_to_base.get(role, "base.html")

    if role == 'HR':
        return render(request, 'attedance/hr_monthly_summary.html', {
            'summary': summary,
            'month': selected_month,
            'year': selected_year,
        })

    return render(request, 'attedance/monthly_summary.html', {
        'summary': summary,
        'month': selected_month,
        'year': selected_year,
        'base_template': base_template,
    })

from datetime import date, timedelta
from django.db.models import Sum, Q
from .models import Attendance


def calculate_attendance(user, year=None, month=None):

    if not user.date_of_joining:
        return {
            "total_days": 0,
            "present_days": 0,
            "absent_days": 0,
            "month_present": 0,
            "month_absent": 0,
            "total_hours": 0,
            "percentage": 0,
        }

    joining_date = user.date_of_joining
    today = date.today()

    # -------- TOTAL WORKING DAYS (Joining → Today) --------
    total_working_days = 0
    current = joining_date

    while current <= today:
        if current.weekday() != 6:
            total_working_days += 1
        current += timedelta(days=1)

    # -------- TOTAL PRESENT --------
    present_qs = Attendance.objects.filter(
        user=user
    ).filter(
        Q(status__startswith="Present") |
        Q(status__startswith="Half")
    )

    present_days = present_qs.count()

    # -------- TOTAL ABSENT --------
    absent_days = max(total_working_days - present_days, 0)

    # -------- TOTAL HOURS --------
    total_hours = present_qs.aggregate(
        total=Sum("total_hours")
    )["total"] or 0

    # -------- PERCENTAGE --------
    percentage = (
        (present_days / total_working_days) * 100
        if total_working_days else 0
    )

    # -------- MONTH FILTER --------
    month_present = 0
    month_absent = 0

    if year and month:

        month_records = Attendance.objects.filter(
            user=user,
            date__year=year,
            date__month=month
        ).count()

    return {
        "total_days": total_working_days,
        "present_days": present_days,
        "absent_days": absent_days,
        "month_present": month_present,
        "month_absent": month_absent,
        "total_hours": round(total_hours, 2),
        "percentage": round(percentage, 2),
    }


def calculate_payroll_data(employee, year, month):

    import calendar
    import decimal
    from decimal import Decimal
    from datetime import date, timedelta
    from django.db.models import Q
    from .models import Attendance, Leave, Holiday, SalaryStructure, HRSettings

    year = int(year)
    month = int(month)
    days_in_month = calendar.monthrange(year, month)[1]

    present_days = 0
    half_days = 0
    absent_days = 0
    paid_leave_days = 0
    unpaid_leave_days = 0
    holidays_count = 0
    week_offs_count = 0
    sandwich_leave_days = 0

    # ─── Data Lookups ───────────────────────────────────────────────────────────
    from django.db.models import Q
    holiday_dates = set(
        Holiday.objects.filter(
            status='Approved',
            date__year=year,
            date__month=month
        ).filter(
            Q(department__isnull=True) | Q(department='') | Q(department=employee.department)
        ).values_list('date', flat=True)
    )
    # Also grab ALL holidays for edge-of-month sandwich checks
    all_holiday_dates = set(
        Holiday.objects.filter(status='Approved').filter(
            Q(department__isnull=True) | Q(department='') | Q(department=employee.department)
        ).values_list('date', flat=True)
    )

    attendance_dict = {
        att.date: att
        for att in Attendance.objects.filter(user=employee, date__year=year, date__month=month)
    }
    approved_leaves = list(Leave.objects.filter(user=employee, status='Approved'))

    def get_leave_for_date(d):
        for lv in approved_leaves:
            if lv.from_date <= d <= lv.to_date:
                return lv
        return None

    # ─── Classify a single in-month day ─────────────────────────────────────────
    def classify_day(d):
        if employee.date_of_joining and d < employee.date_of_joining:
            return 'pre_joining'
        if d.weekday() == 6:
            return 'week_off'
        if d in all_holiday_dates:
            return 'holiday'
        att = attendance_dict.get(d)
        if att:
            status_lower = (att.status or '').lower()
            if 'present' in status_lower or 'work from home' in status_lower or 'wfh' in status_lower or 'on duty' in status_lower:
                return 'present'
            if 'half' in status_lower:
                return 'half'
        lv = get_leave_for_date(d)
        if lv:
            return 'paid_leave' if lv.leave_type == 'Paid' else 'unpaid_leave'
        return 'absent'

    # Classify a day that may be outside the current month (for sandwich edge check)
    def classify_any_day(d):
        if employee.date_of_joining and d < employee.date_of_joining:
            return 'pre_joining'
        if d.weekday() == 6:
            return 'week_off'
        if d in all_holiday_dates:
            return 'holiday'
        lv = get_leave_for_date(d)
        if lv:
            return 'paid_leave' if lv.leave_type == 'Paid' else 'unpaid_leave'
        att_any = Attendance.objects.filter(user=employee, date=d).first()
        if att_any:
            s = (att_any.status or '').lower()
            if 'present' in s or 'work from home' in s or 'wfh' in s or 'on duty' in s:
                return 'present'
            if 'half' in s:
                return 'half'
        return 'absent'

    # ─── Build day_status map ────────────────────────────────────────────────────
    day_status = {}
    for d_num in range(1, days_in_month + 1):
        curr_date = date(year, month, d_num)
        day_status[curr_date] = classify_day(curr_date)

    # ─── Accumulate base counts ──────────────────────────────────────────────────
    for d, status in day_status.items():
        if status == 'pre_joining':
            unpaid_leave_days += 1
        elif status == 'week_off':
            week_offs_count += 1
        elif status == 'holiday':
            holidays_count += 1
        elif status == 'present':
            present_days += 1
        elif status == 'half':
            half_days += 1
        elif status == 'paid_leave':
            paid_leave_days += 1
        elif status in ('unpaid_leave', 'absent'):
            if status == 'unpaid_leave':
                unpaid_leave_days += 1
            else:
                absent_days += 1

    # ─── Sandwich Leave Logic ────────────────────────────────────────────────────
    ABSENT_SIDE = {'absent', 'unpaid_leave'}

    hr_settings = HRSettings.objects.first() or HRSettings.objects.create()

    if hr_settings.sandwich_leave_enabled:
        processed = set()
        month_dates = sorted(day_status.keys())

        for start_d in month_dates:
            if start_d in processed:
                continue
            if day_status[start_d] not in ('week_off', 'holiday'):
                continue

            # Gather contiguous block of offs/holidays
            block = []
            cur = start_d
            while True:
                cur_status = day_status.get(cur) or classify_any_day(cur)
                if cur_status in ('week_off', 'holiday'):
                    block.append(cur)
                    cur = cur + timedelta(days=1)
                else:
                    break

            day_before = block[0] - timedelta(days=1)
            day_after  = block[-1] + timedelta(days=1)

            status_before = day_status.get(day_before) or classify_any_day(day_before)
            status_after  = day_status.get(day_after)  or classify_any_day(day_after)

            if status_before in ABSENT_SIDE and status_after in ABSENT_SIDE:
                for sw_d in block:
                    if sw_d in day_status:   # only convert days inside this month
                        orig = day_status[sw_d]
                        day_status[sw_d] = 'sandwich_absent'
                        sandwich_leave_days += 1
                        if orig == 'week_off':
                            week_offs_count -= 1
                        elif orig == 'holiday':
                            holidays_count -= 1
                        absent_days += 1

            for d in block:
                processed.add(d)

    # ─── Compute paid_days for proration ────────────────────────────────────────
    paid_days = (
        Decimal(present_days) +
        Decimal(paid_leave_days) +
        Decimal(holidays_count) +
        Decimal(week_offs_count) +
        (Decimal(half_days) * Decimal('0.5'))
    )

    prorate = Decimal(paid_days) / Decimal(days_in_month)

    # ─── Salary Structure ────────────────────────────────────────────────────────
    try:
        struct = employee.salary_structure
    except SalaryStructure.DoesNotExist:
        gross = employee.salary or Decimal('0.00')
        basic = gross * Decimal('0.50')
        hra = basic * Decimal('0.40')
        transport = Decimal('1600.00') if gross >= 10000 else Decimal('0.00')
        medical = Decimal('1250.00') if gross >= 10000 else Decimal('0.00')
        special = gross - (basic + hra + transport + medical)
        if special < 0:
            special = Decimal('0.00')

        struct = SalaryStructure(
            employee=employee,
            monthly_gross=gross,
            basic_salary=basic,
            hra=hra,
            transport_allowance=transport,
            medical_allowance=medical,
            special_allowance=special,
            bonus=Decimal('0.00'),
            pf_enabled=True,
            esi_enabled=True,
            pt_enabled=True,
            pt_amount=Decimal('200.00'),
            tds_amount=Decimal('0.00'),
            other_deductions=Decimal('0.00')
        )

    # Prorate earnings components
    actual_basic     = struct.basic_salary * prorate
    actual_hra       = struct.hra * prorate
    actual_transport = struct.transport_allowance * prorate
    actual_medical   = struct.medical_allowance * prorate
    actual_special   = struct.special_allowance * prorate
    actual_bonus     = struct.bonus  # bonus is fixed

    actual_gross = (
        actual_basic + actual_hra + actual_transport +
        actual_medical + actual_special + actual_bonus
    )

    # Calculate PF
    pf_ded = Decimal('0.00')
    if struct.pf_enabled:
        if struct.pf_amount > 0:
            pf_ded = struct.pf_amount * prorate
        else:
            pf_ded = actual_basic * (Decimal(str(struct.pf_rate)) / Decimal('100.00'))

    # Calculate ESI
    esi_ded = Decimal('0.00')
    if struct.esi_enabled:
        if struct.monthly_gross <= Decimal('21000.00'):
            if struct.esi_amount > 0:
                esi_ded = struct.esi_amount * prorate
            else:
                esi_ded = actual_gross * (Decimal(str(struct.esi_rate)) / Decimal('100.00'))

    # Calculate PT
    pt_ded = Decimal('0.00')
    if struct.pt_enabled:
        if paid_days > 0:
            pt_ded = struct.pt_amount

    tds_ded   = struct.tds_amount
    other_ded = struct.other_deductions

    total_ded = pf_ded + esi_ded + pt_ded + tds_ded + other_ded
    net_sal   = actual_gross - total_ded
    if net_sal < 0:
        net_sal = Decimal('0.00')

    loss_of_pay = round(max(struct.monthly_gross - actual_gross, Decimal('0.00')), 2)

    return {
        'days_in_month':       days_in_month,
        'working_days':        days_in_month - week_offs_count - holidays_count,
        'present_days':        present_days,
        'half_days':           half_days,
        'absent_days':         absent_days,
        'paid_leaves':         paid_leave_days,
        'unpaid_leaves':       unpaid_leave_days,
        'holidays':            holidays_count,
        'week_offs':           week_offs_count,
        'sandwich_leave_days': sandwich_leave_days,
        'sandwich_policy_on':  hr_settings.sandwich_leave_enabled,
        'paid_days':           float(paid_days),
        'prorate':             float(prorate),
        'basic_salary':        round(actual_basic, 2),
        'hra':                 round(actual_hra, 2),
        'transport_allowance': round(actual_transport, 2),
        'medical_allowance':   round(actual_medical, 2),
        'special_allowance':   round(actual_special, 2),
        'bonus':               round(actual_bonus, 2),
        'gross_salary':        round(actual_gross, 2),
        'pf_deduction':        round(pf_ded, 2),
        'esi_deduction':       round(esi_ded, 2),
        'professional_tax':    round(pt_ded, 2),
        'tds':                 round(tds_ded, 2),
        'other_deductions':    round(other_ded, 2),
        'total_deductions':    round(total_ded, 2),
        'net_salary':          round(net_sal, 2),
        'loss_of_pay':         loss_of_pay,
        'bank_name':           struct.bank_name or '',
        'account_number':      struct.account_number or '',
        'ifsc_code':           struct.ifsc_code or '',
        'pan':                 struct.pan or '',
        'uan':                 struct.uan or '',
        'aadhaar':             struct.aadhaar or '',
    }




from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Examuser, SELECT_TYPE

def exam_register(request):
    if request.method == 'POST':
        # Get data from form
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone_no = request.POST.get('phone_no')
        role = request.POST.get('role')
        password = request.POST.get('password')

        # Simple validation: check if email exists
        if Examuser.objects.filter(email=email).exists():
            messages.error(request, "Email already registered!")
            return redirect('exam_register')

        # Create user
        Examuser.objects.create(
            username=username,
            email=email,
            phone_no=phone_no,
            role=role,
            password=password # Note: In production, use make_password for security!
        )
        messages.success(request, "Registration successful! Please login.")
        return redirect('exam_login')

    return render(request, 'exam_register.html', {'roles': SELECT_TYPE})

def exam_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        user = Examuser.objects.filter(email=email, password=password).first()

        if user:  # ✅ check instance
            request.session['Examuser_id'] = user.id
            request.session['username'] = user.username
            return redirect('exam_user')
        else:
            messages.error(request, "Invalid credentials!")

    return render(request, 'exam_login.html')


from .models import Examuser, ExamSession, Result

def exam_user(request):

    Examuser_id = request.session.get('Examuser_id')  # ✅ same key

    if not Examuser_id:
        return redirect('exam_login')

    user = Examuser.objects.get(id=Examuser_id)

    exam = ExamSession.objects.filter(user=user, is_completed=True).last()

    result = None
    if exam:
        result = Result.objects.filter(exam=exam).first()

    return render(request, 'exam_user.html', {
        'user': user,
        'result': result
    })


def exam_logout(request):
    request.session.flush()
    return redirect('exam_login')


from django.shortcuts import redirect

@login_required(login_url='hr_login')
def create_question_page(request):
    if request.user.role != "HR":
        return redirect("hr_login")

    return render(request, "create_question_page.html")


@login_required(login_url='hr_login')
def view_questions(request):
    if request.user.role != "HR":
        return redirect("hr_login")

    return render(request, "view_questions.html")

from.models import Question
from django.http import JsonResponse
import json



def save_question(request):
    """
    Save a new question.
    """
    try:
        data = json.loads(request.body.decode("utf-8"))

        Question.objects.create(
            language=data.get("language"),
            question_text=data.get("question"),
            option_a=data.get("optionA"),
            option_b=data.get("optionB"),
            option_c=data.get("optionC"),
            option_d=data.get("optionD"),
            correct_option=data.get("correct"),
        )

        return JsonResponse({"status": "success", "message": "Question saved successfully!"})

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})




def get_questions(request):
    """
    Fetch questions filtered by language/role.
    """
    language = request.GET.get("language")
    if language:
        questions = Question.objects.filter(language=language)
    else:
        questions = Question.objects.all()

    data = [
        {
            "id": q.id,
            "language": q.language,
            "question": q.question_text,
            "optionA": q.option_a,
            "optionB": q.option_b,
            "optionC": q.option_c,
            "optionD": q.option_d,
            "correct": q.correct_option,
        }
        for q in questions
    ]
    return JsonResponse({"questions": data})

from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods



@csrf_exempt
@require_http_methods(["POST"])
def update_question(request, id):
    """Update an existing question"""
    try:
        data = json.loads(request.body.decode("utf-8"))
        q = Question.objects.get(id=id)

        q.language = data.get("language", q.language)
        q.question_text = data.get("question", q.question_text)
        q.option_a = data.get("optionA", q.option_a)
        q.option_b = data.get("optionB", q.option_b)
        q.option_c = data.get("optionC", q.option_c)
        q.option_d = data.get("optionD", q.option_d)
        q.correct_option = data.get("correct", q.correct_option)
        q.save()

        return JsonResponse({"status": "success", "message": "Question updated successfully!"})
    except Question.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Question not found."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})


@csrf_exempt
@require_http_methods(["DELETE"])
def delete_question(request, id):
    """Delete a question"""
    try:
        q = Question.objects.get(id=id)
        q.delete()
        return JsonResponse({"status": "success", "message": "Question deleted successfully!"})
    except Question.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Question not found."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

from .models import ExamSession,Result,UserAnswer


# ---------------- START EXAM ----------------
import random
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Examuser, Question, ExamSession, UserAnswer, Result


import random

def start_exam(request):

    Examuser_id = request.session.get("Examuser_id")

    if not Examuser_id:
        return redirect("exam_login")

    user = Examuser.objects.get(id=Examuser_id)

    existing_exam = ExamSession.objects.filter(
        user=user,
        is_completed=True
    ).first()

    if existing_exam:
        return redirect("result_page")

    all_questions = Question.objects.filter(language=user.role)

    if all_questions.count() < 30:
        messages.error(request, "Not enough questions available.")
        return redirect("exam_user")

    # 🔹 FIRST TIME LOAD
    if request.method == "GET":

        questions = random.sample(list(all_questions), 30)

        # Store question IDs in session
        request.session["exam_questions"] = [q.id for q in questions]

    else:
        # 🔹 On submit — fetch same questions
        question_ids = request.session.get("exam_questions")
        questions = Question.objects.filter(id__in=question_ids)

        exam = ExamSession.objects.create(
            user=user,
            language=user.role
        )

        correct_count = 0

        for question in questions:
            selected = request.POST.get(str(question.id))

            if selected:
                is_correct = selected == question.correct_option

                if is_correct:
                    correct_count += 1

                UserAnswer.objects.create(
                    exam=exam,
                    question=question,
                    selected_option=selected,
                    is_correct=is_correct
                )

        exam.is_completed = True
        exam.save()

        percentage = (correct_count / 30) * 100

        Result.objects.create(
            exam=exam,
            total_questions=30,
            correct_answers=correct_count,
            score_percentage=percentage
        )

        # Clear session
        del request.session["exam_questions"]

        return redirect("result_page")

    return render(request, "exam.html", {
        "questions": questions,
        "user": user,
        "exam_duration": 30
    })


# ---------------- RESULT PAGE ----------------
def result_page(request):

    user_id = request.session.get("Examuser_id")

    if not user_id:
        return redirect("login")

    user = Examuser.objects.get(id=user_id)

    exam = ExamSession.objects.filter(user=user, is_completed=True).last()

    if not exam:
        messages.error(request, "No completed exam found.")
        return redirect("dashboard")

    result = Result.objects.filter(exam=exam).first()

    if not result:
        messages.error(request, "Result not generated yet.")
        return redirect("dashboard")

    return render(request, "result.html", {
        "result": result,
        "user": user
    })

from django.shortcuts import redirect

@login_required(login_url='hr_login')
def all_exams_data(request):
    if request.user.role != "HR":
        return redirect("hr_login")

    all_users = Examuser.objects.all()
    return render(request, "all_exams_data.html", {"all": all_users})


@login_required(login_url='hr_login')
def exam_score(request):
    if request.user.role != "HR":
        return redirect("hr_login")

    score = Result.objects.all()
    return render(request, "score.html", {"score": score})


@login_required(login_url='hr_login')
def user_exam_alldata(request):
    if request.user.role != "HR":
        return redirect("hr_login")

    return render(request, "user_exam_alldata.html")


# report view
@login_required
def daily_reports(request):
    user = request.user

    # ---------------- REPORTS BY ROLE ----------------
    if user.role == "Employee":
        reports = DailyReport.objects.filter(user__role="Employee")
    if user.role == "TeamLead":
        reports = DailyReport.objects.filter(user__role="TeamLead")
        

    elif user.role == "Manager":
        reports = DailyReport.objects.filter(user__role="Manager")
    elif user.role == "HR":
        reports = DailyReport.objects.filter(user__role="HR")
    else:
        reports = DailyReport.objects.none()

    selected_date = request.GET.get("date")
    selected_month = request.GET.get("month")

    if selected_date:
        reports = reports.filter(report_date=selected_date)

    if selected_month:
        year, month = selected_month.split("-")
        reports = reports.filter(
            report_date__year=year,
            report_date__month=month
        )
    
    reports = reports.select_related("user", "project").order_by("-report_date")

    # ---------------- BACK BUTTON & BASE TEMPLATE LOGIC ----------------
    role_to_base = {
        "MD": ("md/md_dashboard.html", "md_dashboard"),
        "HR": ("base.html", "home"),
        "Manager": ("main/mddash.html", "index"),
        "TeamLead": ("tl/side_bar.html", "tl_dashboard"),
        "Employee": ("employee/employee_base.html", "employee_dashboard"),
    }

    base_template, back_url = role_to_base.get(
        user.role,
        ("employee/employee_base.html", "employee_dashboard")
    )

    return render(request, "tl/own_reports.html", {
        "reports": reports,
        "selected_date": selected_date,
        "selected_month": selected_month,
        "back_url": back_url,
        "base_template": base_template,
    })
    
    
    
def redirect_by_role(user):
    if user.role == "Employee":
        return redirect("employee_dashboard")
    elif user.role == "TeamLead":
        return redirect("tl_dashboard")
    elif user.role == "Manager":
        return redirect("index")
    elif user.role == "HR":
        return redirect("home")
    elif user.role == "MD":
        return redirect("md_dashboard")
    else:
        return redirect("home")

def viewall(request):
    from django.contrib.auth import login, get_user_model
    User = get_user_model()
    
    if request.user.is_authenticated:
        return redirect_by_role(request.user)

    if request.method == "POST":
        identifier = request.POST.get("username")
        password = request.POST.get("password")
        remember_me = request.POST.get("remember_me") == "on"

        from django.db.models import Q
        user_obj = User.objects.filter(
            Q(emp_id__iexact=identifier) |
            Q(email__iexact=identifier) |
            Q(username__iexact=identifier)
        ).first()

        if user_obj and user_obj.check_password(password):
            login(request, user_obj)
            if not remember_me:
                request.session.set_expiry(0)
            else:
                request.session.set_expiry(1209600)
            return redirect_by_role(user_obj)
        else:
            messages.error(request, "Invalid Employee ID/Email or Password")

    return render(request, 'viewall.html')
    
    
    
    # hr count view
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model

User = get_user_model()

@login_required
def members_list(request):
    # Only Manager, TeamLead, Employee
    members = User.objects.filter(
        role__in=["Manager", "TeamLead", "Employee"]
    ).order_by("role")

    total = members.count()

    return render(request, "members_list.html", {
        "members": members,
        "total": total
    })    
    
    # tl edit profile
    
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages

@login_required
def tledit_profile(request):
    user = request.user  # Directly using the custom User model

    if request.method == "POST":
        # Update User fields
        user.first_name = request.POST.get("first_name")
        user.last_name = request.POST.get("last_name")
        user.email = request.POST.get("email")
        user.phone = request.POST.get("phone")

        # Update profile image
        if request.FILES.get("profile_image"):
            user.profile_pic = request.FILES.get("profile_image")

        user.save()
        messages.success(request, "Profile updated successfully!")
        return redirect("tl_dashboard")

    context = {
        "user": user
    }
    return render(request, "tl/teamlead_edit_profile.html", context)


# ==================== PAYSLIP MODULE VIEWS ====================

import calendar
import decimal
from decimal import Decimal
from datetime import date, datetime
from django.db import models
from django.db.models import Q
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.core.files.base import ContentFile

from .models import Payslip, PayslipDownloadLog, Holiday, SalaryStructure, PayrollAuditLog, User

def number_to_words(num):
    num = int(num)
    if num <= 0:
        return 'Zero'
    a = ['','One','Two','Three','Four','Five','Six','Seven','Eight','Nine','Ten','Eleven','Twelve','Thirteen','Fourteen','Fifteen','Sixteen','Seventeen','Eighteen','Nineteen']
    b = ['','','Twenty','Thirty','Forty','Fifty','Sixty','Seventy','Eighty','Ninety']
    
    def to_words(n):
        if n < 20:
            return a[n]
        if n < 100:
            return b[n // 10] + ((' ' + a[n % 10]) if n % 10 else '')
        if n < 1000:
            return a[n // 100] + ' Hundred' + ((' ' + to_words(n % 100)) if n % 100 else '')
        if n < 100000:
            return to_words(n // 1000) + ' Thousand' + ((' ' + to_words(n % 1000)) if n % 1000 else '')
        if n < 10000000:
            return to_words(n // 100000) + ' Lakh' + ((' ' + to_words(n % 100000)) if n % 100000 else '')
        return to_words(n // 10000000) + ' Crore' + ((' ' + to_words(n % 10000000)) if n % 10000000 else '')
        
    return to_words(num) + ' Rupees Only'


_fonts_registered = False

def _register_pdf_fonts():
    global _fonts_registered
    if _fonts_registered:
        return
        
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.fonts import addMapping
    import xhtml2pdf.default
    import os

    font_options = [
        {
            'name': 'Arial',
            'paths': {
                'reg': r"C:\Windows\Fonts\arial.ttf",
                'bold': r"C:\Windows\Fonts\arialbd.ttf",
                'italic': r"C:\Windows\Fonts\ariali.ttf",
                'bi': r"C:\Windows\Fonts\arialbi.ttf"
            }
        },
        {
            'name': 'Calibri',
            'paths': {
                'reg': r"C:\Windows\Fonts\calibri.ttf",
                'bold': r"C:\Windows\Fonts\calibrib.ttf",
                'italic': r"C:\Windows\Fonts\calibrii.ttf",
                'bi': r"C:\Windows\Fonts\calibriz.ttf"
            }
        },
        {
            'name': 'SegoeUI',
            'paths': {
                'reg': r"C:\Windows\Fonts\segoeui.ttf",
                'bold': r"C:\Windows\Fonts\segoeuib.ttf",
                'italic': r"C:\Windows\Fonts\segoeuii.ttf",
                'bi': r"C:\Windows\Fonts\segoeuiz.ttf"
            }
        },
        {
            'name': 'DejaVuSans',
            'paths': {
                'reg': "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                'bold': "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                'italic': "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf",
                'bi': "/usr/share/fonts/truetype/dejavu/DejaVuSans-BoldOblique.ttf"
            }
        },
        {
            'name': 'LiberationSans',
            'paths': {
                'reg': "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                'bold': "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                'italic': "/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf",
                'bi': "/usr/share/fonts/truetype/liberation/LiberationSans-BoldItalic.ttf"
            }
        }
    ]

    selected_font = None
    for opt in font_options:
        reg_path = opt['paths']['reg']
        if os.path.exists(reg_path):
            try:
                pdfmetrics.registerFont(TTFont(opt['name'], reg_path))
                
                bold_path = opt['paths']['bold']
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont(f"{opt['name']}-Bold", bold_path))
                else:
                    pdfmetrics.registerFont(TTFont(f"{opt['name']}-Bold", reg_path))
                    
                italic_path = opt['paths']['italic']
                if os.path.exists(italic_path):
                    pdfmetrics.registerFont(TTFont(f"{opt['name']}-Italic", italic_path))
                else:
                    pdfmetrics.registerFont(TTFont(f"{opt['name']}-Italic", reg_path))
                    
                bi_path = opt['paths']['bi']
                if os.path.exists(bi_path):
                    pdfmetrics.registerFont(TTFont(f"{opt['name']}-BoldItalic", bi_path))
                else:
                    pdfmetrics.registerFont(TTFont(f"{opt['name']}-BoldItalic", reg_path))

                addMapping(opt['name'], 0, 0, opt['name'])
                addMapping(opt['name'], 1, 0, f"{opt['name']}-Bold")
                addMapping(opt['name'], 0, 1, f"{opt['name']}-Italic")
                addMapping(opt['name'], 1, 1, f"{opt['name']}-BoldItalic")

                selected_font = opt['name']
                break
            except Exception:
                continue

    if selected_font:
        xhtml2pdf.default.DEFAULT_FONT['customunicode'] = selected_font
        xhtml2pdf.default.DEFAULT_FONT[selected_font.lower()] = selected_font
        xhtml2pdf.default.DEFAULT_FONT['arial'] = selected_font
        xhtml2pdf.default.DEFAULT_FONT['calibri'] = selected_font
        xhtml2pdf.default.DEFAULT_FONT['sans-serif'] = selected_font
        _fonts_registered = True
        return selected_font
    return None


def generate_payslip_pdf_file(payslip):
    _register_pdf_fonts()
    from io import BytesIO
    template_path = 'employee/payslip_pdf.html'
    
    logo_file = finders.find('images/Logo.png')
    if logo_file:
        logo_path = "file:///" + logo_file.replace("\\", "/")
    else:
        logo_file2 = finders.find('images/logo1.png')
        logo_path = "file:///" + logo_file2.replace("\\", "/") if logo_file2 else ""

    context = {
        'payslip': payslip,
        'logo_path': logo_path,
        'amount_in_words': number_to_words(payslip.net_salary),
    }
    
    template = get_template(template_path)
    html = template.render(context)
    
    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=pdf_buffer, encoding='UTF-8')
    if not pisa_status.err:
        pdf_name = f"payslip_{payslip.employee.emp_id}_{payslip.month}_{payslip.year}.pdf"
        payslip.payslip_pdf.save(pdf_name, ContentFile(pdf_buffer.getvalue()), save=False)
        return True
    return False


# --------------- EMPLOYEE: My Payslips ---------------
@login_required
def my_payslips(request):
    user = request.user
    payslips = Payslip.objects.filter(employee=user, is_published=True).order_by('-year', '-month')

    # Filters
    filter_month = request.GET.get('month')
    filter_year = request.GET.get('year')
    filter_fy = request.GET.get('fy')

    if filter_month:
        payslips = payslips.filter(month=filter_month)
    if filter_year:
        payslips = payslips.filter(year=filter_year)
    if filter_fy:
        try:
            fy_start = int(filter_fy)
            payslips = payslips.filter(
                models.Q(year=fy_start, month__gte=4) |
                models.Q(year=fy_start + 1, month__lte=3)
            )
        except (ValueError, TypeError):
            pass

    # Dashboard widget data
    latest = payslips.first()
    total_count = payslips.count()

    # Available years for filter
    years = Payslip.objects.filter(employee=user, is_published=True).values_list('year', flat=True).distinct().order_by('-year')
    current_year = datetime.now().year
    fy_options = [(y, f"FY {y}-{str(y+1)[2:]}") for y in range(current_year, current_year - 5, -1)]

    context = {
        'payslips': payslips,
        'latest': latest,
        'total_count': total_count,
        'years': years,
        'fy_options': fy_options,
        'filter_month': filter_month,
        'filter_year': filter_year,
        'filter_fy': filter_fy,
        'month_choices': list(calendar.month_name)[1:],
    }
    return render(request, 'employee/payslips.html', context)


# --------------- EMPLOYEE: View single payslip ---------------
@login_required
def view_payslip(request, pk):
    payslip = get_object_or_404(Payslip, pk=pk)
    
    # Security check
    if request.user.role not in ('HR', 'MD') and payslip.employee != request.user:
        messages.error(request, "Access denied.")
        return redirect('home')
        
    if request.user.role not in ('HR', 'MD') and not payslip.is_published:
        messages.error(request, "This payslip is not published yet.")
        return redirect('my_payslips')
        
    # Fetch employee's published payslips for history (or all if admin view)
    if request.user.role in ('HR', 'MD'):
        history = Payslip.objects.filter(employee=payslip.employee).order_by('-year', '-month')
    else:
        history = Payslip.objects.filter(employee=payslip.employee, is_published=True).order_by('-year', '-month')
        
    return render(request, 'employee/payslip_detail.html', {
        'payslip': payslip,
        'is_admin_view': request.user.role in ('HR', 'MD'),
        'history': history
    })


# --------------- EMPLOYEE: Download payslip ---------------
@login_required
def download_payslip(request, pk):
    payslip = get_object_or_404(Payslip, pk=pk)
    
    if request.user.role not in ('HR', 'MD') and payslip.employee != request.user:
        messages.error(request, "Access denied.")
        return redirect('home')
        
    if request.user.role not in ('HR', 'MD') and not payslip.is_published:
        messages.error(request, "This payslip is not published yet.")
        return redirect('my_payslips')
        
    # Generate PDF if not exists
    if not payslip.payslip_pdf:
        generate_payslip_pdf_file(payslip)
        payslip.save()
        
    # Audit log
    ip = request.META.get('REMOTE_ADDR')
    PayslipDownloadLog.objects.create(payslip=payslip, downloaded_by=request.user, ip_address=ip)
    PayrollAuditLog.objects.create(
        payslip=payslip,
        performed_by=request.user,
        action='DOWNLOAD',
        ip_address=ip,
        details=f"Downloaded payslip of {payslip.employee.username} for {payslip.month_name} {payslip.year}"
    )
    
    if payslip.payslip_pdf:
        response = FileResponse(payslip.payslip_pdf.open('rb'), content_type='application/pdf')
        fname = f"{payslip.employee.get_full_name() or payslip.employee.username}_{payslip.month_name}_{payslip.year}_Payslip.pdf"
        fname = fname.replace(' ', '_')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        return response
    else:
        messages.error(request, "Error generating payslip PDF.")
        return redirect('my_payslips' if request.user.role not in ('HR', 'MD') else 'hr_payslip_list')


# --------------- HR ADMIN: Payslip Management ---------------
@login_required
def hr_payslip_list(request):
    user = request.user
    if user.role not in ('HR', 'MD', 'Manager'):
        messages.error(request, "Access denied.")
        return redirect('home')

    search = request.GET.get('search', '')
    filter_month = request.GET.get('month', '')
    filter_year = request.GET.get('year', '')
    filter_status = request.GET.get('status', '')

    payslips = Payslip.objects.select_related('employee').order_by('-year', '-month', 'employee__first_name')

    if user.role == 'Manager':
        # Filter to only show subordinate employees
        managed_user_ids = User.objects.filter(
            models.Q(reporting_manager=user) | 
            models.Q(teams__lead=user)
        ).values_list('id', flat=True).distinct()
        payslips = payslips.filter(employee_id__in=managed_user_ids)

    if search:
        payslips = payslips.filter(
            models.Q(employee_name__icontains=search) |
            models.Q(employee__first_name__icontains=search) |
            models.Q(employee__username__icontains=search) |
            models.Q(employee__emp_id__icontains=search)
        )
    if filter_month:
        payslips = payslips.filter(month=filter_month)
    if filter_year:
        payslips = payslips.filter(year=filter_year)
    if filter_status:
        payslips = payslips.filter(status=filter_status)

    employees = User.objects.filter(role__in=['Employee', 'TeamLead', 'Manager']).order_by('first_name')
    years = range(datetime.now().year, datetime.now().year - 5, -1)

    context = {
        'payslips': payslips,
        'employees': employees,
        'years': years,
        'search': search,
        'filter_month': filter_month,
        'filter_year': filter_year,
        'filter_status': filter_status,
        'month_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
    }
    return render(request, 'hr/payslip_admin.html', context)


@login_required
def hr_payslip_create(request):
    user = request.user
    if user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')

    if request.method == 'POST':
        emp_id = request.POST.get('employee')
        month = request.POST.get('month')
        year = request.POST.get('year')
        employee = get_object_or_404(User, id=emp_id)

        # Check duplicate
        if Payslip.objects.filter(employee=employee, month=month, year=year).exists():
            messages.error(request, f"Payslip for {employee.get_full_name() or employee.username} for {calendar.month_name[int(month)]} {year} already exists.")
            return redirect('hr_payslip_list')

        # Create
        payslip = Payslip(
            employee=employee,
            month=int(month),
            year=int(year),
            basic_salary=Decimal(request.POST.get('basic_salary') or 0),
            hra=Decimal(request.POST.get('hra') or 0),
            transport_allowance=Decimal(request.POST.get('transport_allowance') or 0),
            medical_allowance=Decimal(request.POST.get('medical_allowance') or 0),
            special_allowance=Decimal(request.POST.get('special_allowance') or 0),
            bonus=Decimal(request.POST.get('bonus') or 0),
            pf_deduction=Decimal(request.POST.get('pf_deduction') or 0),
            esi_deduction=Decimal(request.POST.get('esi_deduction') or 0),
            professional_tax=Decimal(request.POST.get('professional_tax') or 0),
            tds=Decimal(request.POST.get('tds') or 0),
            loan_deduction=Decimal(request.POST.get('loan_deduction') or 0),
            other_deductions=Decimal(request.POST.get('other_deductions') or 0),
            working_days=int(request.POST.get('working_days') or 26),
            days_present=int(request.POST.get('days_present') or 26),
            days_absent=int(request.POST.get('days_absent') or 0),
            leaves_taken=int(request.POST.get('leaves_taken') or 0),
            
            unpaid_leave_days=int(request.POST.get('unpaid_leave_days') or 0),
            half_days=int(request.POST.get('half_days') or 0),
            holidays=int(request.POST.get('holidays') or 0),
            week_offs=int(request.POST.get('week_offs') or 0),
            sandwich_leave_days=int(request.POST.get('sandwich_leave_days') or 0),
            
            status=request.POST.get('status', 'Pending'),
            payment_date=request.POST.get('payment_date') or None,
            is_published=request.POST.get('is_published') == 'on',
            notes=request.POST.get('notes', ''),
            created_by=user,
        )
        
        # Biographical details
        payslip.employee_name = employee.get_full_name() or employee.username
        payslip.designation = employee.role
        payslip.department = employee.get_department_display() or employee.department
        
        try:
            struct = employee.salary_structure
            payslip.bank_name = struct.bank_name
            payslip.account_number = struct.account_number
            payslip.ifsc_code = struct.ifsc_code
            payslip.pan = struct.pan
            payslip.uan = struct.uan
            payslip.aadhaar = struct.aadhaar
        except SalaryStructure.DoesNotExist:
            pass

        if request.FILES.get('payslip_pdf'):
            payslip.payslip_pdf = request.FILES['payslip_pdf']
            payslip.save()
        else:
            payslip.save()
            generate_payslip_pdf_file(payslip)
            payslip.save()
            
        # Log Audit
        ip = request.META.get('REMOTE_ADDR')
        PayrollAuditLog.objects.create(
            payslip=payslip,
            performed_by=request.user,
            action='GENERATE',
            ip_address=ip,
            details=f"Manually created payslip of {employee.username} for {payslip.month_name} {payslip.year}"
        )

        messages.success(request, f"Payslip created for {employee.get_full_name() or employee.username}.")
        return redirect('hr_payslip_list')

    employees = User.objects.filter(role__in=['Employee', 'TeamLead', 'Manager']).order_by('first_name')
    years = range(datetime.now().year, datetime.now().year - 5, -1)
    context = {
        'employees': employees,
        'years': years,
        'month_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
    }
    return render(request, 'hr/payslip_create.html', context)


@login_required
def hr_payslip_edit(request, pk):
    user = request.user
    if user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')

    payslip = get_object_or_404(Payslip, pk=pk)
    
    if payslip.is_locked:
        messages.error(request, "This payslip is locked and cannot be edited.")
        return redirect('hr_payslip_list')

    if request.method == 'POST':
        payslip.basic_salary = Decimal(request.POST.get('basic_salary') or 0)
        payslip.hra = Decimal(request.POST.get('hra') or 0)
        payslip.transport_allowance = Decimal(request.POST.get('transport_allowance') or 0)
        payslip.medical_allowance = Decimal(request.POST.get('medical_allowance') or 0)
        payslip.special_allowance = Decimal(request.POST.get('special_allowance') or 0)
        payslip.bonus = Decimal(request.POST.get('bonus') or 0)
        
        payslip.pf_deduction = Decimal(request.POST.get('pf_deduction') or 0)
        payslip.esi_deduction = Decimal(request.POST.get('esi_deduction') or 0)
        payslip.professional_tax = Decimal(request.POST.get('professional_tax') or 0)
        payslip.tds = Decimal(request.POST.get('tds') or 0)
        payslip.loan_deduction = Decimal(request.POST.get('loan_deduction') or 0)
        payslip.other_deductions = Decimal(request.POST.get('other_deductions') or 0)
        
        payslip.working_days = int(request.POST.get('working_days') or 26)
        payslip.days_present = int(request.POST.get('days_present') or 26)
        payslip.days_absent = int(request.POST.get('days_absent') or 0)
        payslip.leaves_taken = int(request.POST.get('leaves_taken') or 0)
        
        payslip.unpaid_leave_days = int(request.POST.get('unpaid_leave_days') or 0)
        payslip.half_days = int(request.POST.get('half_days') or 0)
        payslip.holidays = int(request.POST.get('holidays') or 0)
        payslip.week_offs = int(request.POST.get('week_offs') or 0)
        
        payslip.status = request.POST.get('status', 'Pending')
        payslip.payment_date = request.POST.get('payment_date') or None
        payslip.is_published = request.POST.get('is_published') == 'on'
        payslip.notes = request.POST.get('notes', '')
        
        if request.FILES.get('payslip_pdf'):
            payslip.payslip_pdf = request.FILES['payslip_pdf']
            payslip.save()
        else:
            payslip.save()
            generate_payslip_pdf_file(payslip)
            payslip.save()
            
        # Log Audit
        ip = request.META.get('REMOTE_ADDR')
        PayrollAuditLog.objects.create(
            payslip=payslip,
            performed_by=request.user,
            action='REGENERATE',
            ip_address=ip,
            details=f"Manually edited/regenerated payslip of {payslip.employee.username} for {payslip.month_name} {payslip.year}"
        )

        messages.success(request, "Payslip updated successfully.")
        return redirect('hr_payslip_list')

    employees = User.objects.filter(role__in=['Employee', 'TeamLead', 'Manager']).order_by('first_name')
    years = range(datetime.now().year, datetime.now().year - 5, -1)
    context = {
        'payslip': payslip,
        'employees': employees,
        'years': years,
        'month_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
    }
    return render(request, 'hr/payslip_edit.html', context)


@login_required
def hr_payslip_delete(request, pk):
    user = request.user
    if user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')

    payslip = get_object_or_404(Payslip, pk=pk)
    if payslip.is_locked:
        messages.error(request, "Locked payslips cannot be deleted.")
        return redirect('hr_payslip_list')
        
    name = str(payslip)
    payslip.delete()
    messages.success(request, f"Payslip deleted: {name}.")
    return redirect('hr_payslip_list')


@login_required
def hr_payslip_toggle_publish(request, pk):
    user = request.user
    if user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')

    payslip = get_object_or_404(Payslip, pk=pk)
    payslip.is_published = not payslip.is_published
    payslip.save()
    
    # Regenerate PDF to capture published status if visual indicator
    generate_payslip_pdf_file(payslip)
    payslip.save()
    
    status = "published" if payslip.is_published else "unpublished"
    messages.success(request, f"Payslip {status} successfully.")
    return redirect('hr_payslip_list')


@login_required
def hr_payslip_view(request, pk):
    user = request.user
    if user.role not in ('HR', 'MD', 'Manager'):
        messages.error(request, "Access denied.")
        return redirect('home')
    payslip = get_object_or_404(Payslip, pk=pk)
    if user.role == 'Manager':
        is_subordinate = User.objects.filter(
            models.Q(id=payslip.employee.id) & (
                models.Q(reporting_manager=user) |
                models.Q(teams__lead=user)
            )
        ).exists()
        if not is_subordinate:
            messages.error(request, "Access denied. Not your subordinate.")
            return redirect('hr_payslip_list')

    if request.user.role in ('HR', 'MD'):
        history = Payslip.objects.filter(employee=payslip.employee).order_by('-year', '-month')
    else:
        history = Payslip.objects.filter(employee=payslip.employee, is_published=True).order_by('-year', '-month')
        
    return render(request, 'employee/payslip_detail.html', {
        'payslip': payslip,
        'is_admin_view': request.user.role in ('HR', 'MD', 'Manager'),
        'history': history
    })


# --------------- HR ADMIN: AJAX Calculation Endpoint ---------------
@login_required
def hr_calculate_payroll_api(request):
    if request.user.role not in ('HR', 'MD'):
        return JsonResponse({'error': 'Access denied'}, status=403)
        
    emp_id = request.GET.get('employee_id')
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    if not all([emp_id, month, year]):
        return JsonResponse({'error': 'Missing parameters'}, status=400)
        
    employee = get_object_or_404(User, id=emp_id)
    try:
        data = calculate_payroll_data(employee, year, month)
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# --------------- HR ADMIN: Bulk Generation ---------------
@login_required
def hr_payslip_bulk_generate(request):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    from .models import AttendanceFinalization, HRSettings
    
    if request.method == 'POST':
        month = int(request.POST.get('month'))
        year = int(request.POST.get('year'))
        publish = request.POST.get('is_published') == 'on'
        force_recalculate = request.POST.get('force_recalculate') == 'on'
        
        # 1. Verify attendance is finalized
        if not AttendanceFinalization.objects.filter(month=month, year=year, is_finalized=True).exists():
            messages.error(request, f"Validation Error: Attendance for {calendar.month_name[month]} {year} has not been finalized yet. Please finalize attendance first.")
            return redirect('hr_payslip_list')
            
        # 2. Check duplicate / already generated
        existing_payslips = Payslip.objects.filter(month=month, year=year)
        if existing_payslips.exists():
            # If any payslip is Approved or Paid, prevent regeneration
            if existing_payslips.filter(status__in=['Approved', 'Paid']).exists():
                messages.error(request, f"Validation Error: Payroll for {calendar.month_name[month]} {year} has already been Approved or Paid. It cannot be generated again.")
                return redirect('hr_payslip_list')
            # If no needs_recalculation and force_recalculate is off, prevent duplication error
            if not existing_payslips.filter(needs_recalculation=True).exists() and not force_recalculate:
                messages.error(request, f"Validation Error: Payroll for {calendar.month_name[month]} {year} has already been generated. Check Force Recalculate to regenerate.")
                return redirect('hr_payslip_list')
                
        # 3. Filter active eligible employees who joined on or before the month
        from datetime import date
        last_day_of_month = date(year, month, calendar.monthrange(year, month)[1])
        employees = User.objects.filter(
            role__in=['Employee', 'TeamLead', 'Manager'],
            is_active=True,
            date_of_joining__lte=last_day_of_month
        ).order_by('first_name')
        
        if not employees.exists():
            messages.error(request, "Validation Error: No active employees who joined on or before the selected month were found.")
            return redirect('hr_payslip_list')
            
        # 4. Verify salary structures exist
        missing_salary_structs = []
        for emp in employees:
            if not hasattr(emp, 'salary_structure'):
                missing_salary_structs.append(emp.get_full_name() or emp.username)
        if missing_salary_structs:
            messages.error(request, f"Validation Error: The following employee(s) do not have a salary structure configured: {', '.join(missing_salary_structs)}")
            return redirect('hr_payslip_list')
            
        generated = 0
        updated = 0
        locked_count = 0
        
        for emp in employees:
            calc = calculate_payroll_data(emp, year, month)
            
            payslip, created = Payslip.objects.get_or_create(
                employee=emp,
                month=month,
                year=year,
                defaults={'created_by': request.user, 'status': 'Generated'}
            )
            
            if payslip.is_locked:
                locked_count += 1
                continue
                
            payslip.basic_salary = calc['basic_salary']
            payslip.hra = calc['hra']
            payslip.transport_allowance = calc['transport_allowance']
            payslip.medical_allowance = calc['medical_allowance']
            payslip.special_allowance = calc['special_allowance']
            payslip.bonus = calc['bonus']
            
            payslip.pf_deduction = calc['pf_deduction']
            payslip.esi_deduction = calc['esi_deduction']
            payslip.professional_tax = calc['professional_tax']
            payslip.tds = calc['tds']
            payslip.other_deductions = calc['other_deductions']
            
            payslip.working_days = calc['working_days']
            payslip.days_present = calc['present_days']
            payslip.days_absent = calc['absent_days']
            payslip.leaves_taken = calc['paid_leaves']
            
            # New metrics
            payslip.unpaid_leave_days = calc['unpaid_leaves']
            payslip.half_days = calc['half_days']
            payslip.holidays = calc['holidays']
            payslip.week_offs = calc['week_offs']
            payslip.sandwich_leave_days = calc.get('sandwich_leave_days', 0)
            
            # Extra fields
            payslip.loss_of_pay = calc.get('loss_of_pay', 0)
            
            # Biographical details copy
            payslip.employee_name = emp.get_full_name() or emp.username
            payslip.designation = emp.role
            payslip.department = emp.get_department_display() or emp.department
            payslip.bank_name = calc['bank_name']
            payslip.account_number = calc['account_number']
            payslip.ifsc_code = calc['ifsc_code']
            payslip.pan = calc['pan']
            payslip.uan = calc['uan']
            payslip.aadhaar = calc['aadhaar']
            
            payslip.is_published = publish
            
            if created:
                payslip.status = 'Generated'
                payslip.generated_by = request.user
                payslip.generated_at = timezone.now()
                payslip.needs_recalculation = False
            else:
                payslip.status = 'Generated' # reset back to Generated for review
                payslip.recalculated_by = request.user
                payslip.recalculated_at = timezone.now()
                payslip.needs_recalculation = False
                
            payslip.save()
            generate_payslip_pdf_file(payslip)
            payslip.save()
            
            ip = request.META.get('REMOTE_ADDR')
            PayrollAuditLog.objects.create(
                payslip=payslip,
                performed_by=request.user,
                action='GENERATE' if created else 'REGENERATE',
                ip_address=ip,
                details=f"Bulk generated payslip of {emp.username} for {payslip.month_name} {payslip.year}"
            )
            
            if created:
                generated += 1
            else:
                updated += 1
                
        messages.success(request, f"Payroll processed: {generated} generated, {updated} updated, {locked_count} locked skipped.")
        return redirect('hr_payslip_list')
        
    years = range(datetime.now().year, datetime.now().year - 5, -1)
    hr_settings = HRSettings.get_settings()
    
    # Check if current/selected month finalized
    filter_month = request.GET.get('month', str(datetime.now().month))
    filter_year = request.GET.get('year', str(datetime.now().year))
    is_finalized = AttendanceFinalization.objects.filter(month=filter_month, year=filter_year, is_finalized=True).exists()
    payslips_exist = Payslip.objects.filter(month=filter_month, year=filter_year).exists()
    
    context = {
        'years': years,
        'month_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'sandwich_policy_on': hr_settings.sandwich_leave_enabled,
        'is_finalized': is_finalized,
        'payslips_exist': payslips_exist,
        'filter_month': int(filter_month),
        'filter_year': int(filter_year),
    }
    return render(request, 'hr/payroll_generate.html', context)


# --------------- HR ADMIN: Email Payslip ---------------
@login_required
def hr_payslip_email(request, pk):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    payslip = get_object_or_404(Payslip, pk=pk)
    
    if not payslip.payslip_pdf:
        generate_payslip_pdf_file(payslip)
        payslip.save()
        
    from django.core.mail import EmailMessage
    from django.conf import settings
    
    subject = f"Payslip for {payslip.month_name} {payslip.year} - YGR Technologies"
    body = (
        f"Dear {payslip.employee_name or payslip.employee.get_full_name() or payslip.employee.username},\n\n"
        f"Please find attached your salary statement (payslip) for the month of {payslip.month_name} {payslip.year}.\n\n"
        f"Take-Home Net Salary: ₹{payslip.net_salary:,.2f}\n"
        f"Payment Status: {payslip.status}\n\n"
        f"Best Regards,\n"
        f"HR Team\n"
        f"YGR Technologies"
    )
    
    email = EmailMessage(
        subject,
        body,
        settings.DEFAULT_FROM_EMAIL or 'hr@ygrteam.com',
        [payslip.employee.email]
    )
    
    if payslip.payslip_pdf:
        email.attach_file(payslip.payslip_pdf.path)
        
    try:
        email.send()
        messages.success(request, f"Payslip emailed to {payslip.employee.email} successfully.")
        ip = request.META.get('REMOTE_ADDR')
        PayrollAuditLog.objects.create(
            payslip=payslip,
            performed_by=request.user,
            action='EMAIL',
            ip_address=ip,
            details=f"Emailed payslip to {payslip.employee.email}"
        )
    except Exception as e:
        messages.error(request, f"Failed to send email: {str(e)}")
        
    return redirect('hr_payslip_list')


# --------------- HR ADMIN: Lock Payslip ---------------
@login_required
def hr_payslip_lock(request, pk):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    payslip = get_object_or_404(Payslip, pk=pk)
    payslip.is_locked = not payslip.is_locked
    payslip.save()
    
    ip = request.META.get('REMOTE_ADDR')
    PayrollAuditLog.objects.create(
        payslip=payslip,
        performed_by=request.user,
        action='LOCK' if payslip.is_locked else 'UNLOCK',
        ip_address=ip,
        details=f"{'Locked' if payslip.is_locked else 'Unlocked'} payslip of {payslip.employee.username} for {payslip.month_name} {payslip.year}"
    )
    
    status = "locked" if payslip.is_locked else "unlocked"
    messages.success(request, f"Payslip {status} successfully.")
    return redirect('hr_payslip_list')


# --------------- HR ADMIN: Mark Paid ---------------
@login_required
def hr_payslip_mark_paid(request, pk):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    payslip = get_object_or_404(Payslip, pk=pk)
    payslip.status = 'Paid'
    payslip.payment_date = date.today()
    payslip.save()
    
    generate_payslip_pdf_file(payslip)
    payslip.save()
    
    ip = request.META.get('REMOTE_ADDR')
    PayrollAuditLog.objects.create(
        payslip=payslip,
        performed_by=request.user,
        action='PAID',
        ip_address=ip,
        details=f"Marked salary as Paid for {payslip.employee.username} for {payslip.month_name} {payslip.year}"
    )
    
    messages.success(request, f"Salary for {payslip.employee.username} marked as Paid.")
    return redirect('hr_payslip_list')


# --------------- HR ADMIN: Payroll Reports ---------------
@login_required
def hr_payroll_reports(request):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    filter_month = request.GET.get('month', str(datetime.now().month))
    filter_year = request.GET.get('year', str(datetime.now().year))
    
    payslips = Payslip.objects.filter(month=filter_month, year=filter_year).select_related('employee')
    
    total_gross = sum(p.gross_salary for p in payslips)
    total_ded = sum(p.total_deductions for p in payslips)
    total_net = sum(p.net_salary for p in payslips)
    total_sandwich = sum(p.sandwich_leave_days for p in payslips)
    
    years = range(datetime.now().year, datetime.now().year - 5, -1)
    
    context = {
        'payslips': payslips,
        'total_gross': total_gross,
        'total_deductions': total_ded,
        'total_net': total_net,
        'total_sandwich_days': total_sandwich,
        'filter_month': int(filter_month),
        'filter_year': int(filter_year),
        'years': years,
        'month_choices': [(i, calendar.month_name[i]) for i in range(1, 13)],
    }
    return render(request, 'hr/payroll_reports.html', context)


# --------------- HR ADMIN: Payroll Export (CSV / PDF) ---------------
@login_required
def hr_payroll_export(request):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    filter_month = int(request.GET.get('month', datetime.now().month))
    filter_year = int(request.GET.get('year', datetime.now().year))
    export_type = request.GET.get('type', 'csv')
    
    payslips = Payslip.objects.filter(month=filter_month, year=filter_year).select_related('employee')
    month_name = calendar.month_name[filter_month]
    
    if export_type == 'csv':
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="Payroll_Report_{month_name}_{filter_year}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Emp ID', 'Name', 'Designation', 'Department', 'Working Days', 'Present Days', 'Gross Salary', 'Total Deductions', 'Net Salary', 'Status'])
        
        for p in payslips:
            writer.writerow([
                p.employee.emp_id,
                p.employee_name or p.employee.get_full_name() or p.employee.username,
                p.designation or p.employee.role,
                p.department or p.employee.get_department_display(),
                p.working_days,
                p.days_present,
                p.gross_salary,
                p.total_deductions,
                p.net_salary,
                p.status
            ])
            
        writer.writerow([])
        writer.writerow(['Total', '', '', '', '', '', sum(p.gross_salary for p in payslips), sum(p.total_deductions for p in payslips), sum(p.net_salary for p in payslips), ''])
        return response
    else:
        context = {
            'payslips': payslips,
            'month_name': month_name,
            'year': filter_year,
            'total_gross': sum(p.gross_salary for p in payslips),
            'total_deductions': sum(p.total_deductions for p in payslips),
            'total_net': sum(p.net_salary for p in payslips),
        }
        
        template = get_template('hr/payroll_report_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Payroll_Report_{month_name}_{filter_year}.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response, encoding='UTF-8')
        if pisa_status.err:
            return HttpResponse('Error generating PDF')
        return response


# --------------- HR ADMIN: Salary Structures ---------------
@login_required
def hr_salary_structure_list(request):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    search = request.GET.get('search', '')
    employees = User.objects.filter(role__in=['Employee', 'TeamLead', 'Manager']).select_related('salary_structure')
    
    if search:
        employees = employees.filter(
            models.Q(first_name__icontains=search) |
            models.Q(username__icontains=search) |
            models.Q(emp_id__icontains=search)
        )
        
    return render(request, 'hr/salary_structure_list.html', {
        'employees': employees,
        'search': search,
    })


@login_required
def hr_salary_structure_detail(request, pk):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    employee = get_object_or_404(User, id=pk)
    try:
        struct = employee.salary_structure
    except SalaryStructure.DoesNotExist:
        struct = None
        
    if request.method == 'POST':
        monthly_gross = Decimal(request.POST.get('monthly_gross') or 0)
        basic_salary = Decimal(request.POST.get('basic_salary') or 0)
        hra = Decimal(request.POST.get('hra') or 0)
        transport_allowance = Decimal(request.POST.get('transport_allowance') or 0)
        medical_allowance = Decimal(request.POST.get('medical_allowance') or 0)
        special_allowance = Decimal(request.POST.get('special_allowance') or 0)
        bonus = Decimal(request.POST.get('bonus') or 0)
        
        pf_enabled = request.POST.get('pf_enabled') == 'on'
        pf_rate = Decimal(request.POST.get('pf_rate') or 12)
        pf_amount = Decimal(request.POST.get('pf_amount') or 0)
        
        esi_enabled = request.POST.get('esi_enabled') == 'on'
        esi_rate = Decimal(request.POST.get('esi_rate') or 0.75)
        esi_amount = Decimal(request.POST.get('esi_amount') or 0)
        
        pt_enabled = request.POST.get('pt_enabled') == 'on'
        pt_amount = Decimal(request.POST.get('pt_amount') or 200)
        
        tds_amount = Decimal(request.POST.get('tds_amount') or 0)
        other_deductions = Decimal(request.POST.get('other_deductions') or 0)
        
        bank_name = request.POST.get('bank_name')
        account_number = request.POST.get('account_number')
        ifsc_code = request.POST.get('ifsc_code')
        pan = request.POST.get('pan')
        uan = request.POST.get('uan')
        
        if not struct:
            struct = SalaryStructure(employee=employee)
            
        struct.monthly_gross = monthly_gross
        struct.basic_salary = basic_salary
        struct.hra = hra
        struct.transport_allowance = transport_allowance
        struct.medical_allowance = medical_allowance
        struct.special_allowance = special_allowance
        struct.bonus = bonus
        
        struct.pf_enabled = pf_enabled
        struct.pf_rate = pf_rate
        struct.pf_amount = pf_amount
        
        struct.esi_enabled = esi_enabled
        struct.esi_rate = esi_rate
        struct.esi_amount = esi_amount
        
        struct.pt_enabled = pt_enabled
        struct.pt_amount = pt_amount
        
        struct.tds_amount = tds_amount
        struct.other_deductions = other_deductions
        
        struct.bank_name = bank_name
        struct.account_number = account_number
        struct.ifsc_code = ifsc_code
        struct.pan = pan
        struct.uan = uan
        struct.save()
        
        # Update user salary field
        employee.salary = monthly_gross
        employee.save()
        
        messages.success(request, f"Salary structure updated for {employee.get_full_name() or employee.username}.")
        return redirect('hr_salary_structure_list')
        
    return render(request, 'hr/salary_structure_detail.html', {
        'employee': employee,
        'struct': struct,
    })


# --------------- HR ADMIN: Holiday Calendar ---------------

def notify_md_holiday_pending(holiday):
    md_users = User.objects.filter(role='MD')
    for md in md_users:
        HolidayNotification.objects.create(
            recipient=md,
            holiday=holiday,
            notif_type='pending',
            message=f"Holiday '{holiday.name}' ({holiday.date}) has been submitted by HR for your approval."
        )

def notify_holiday_approved(holiday):
    # Notify HR creator
    if holiday.created_by:
        HolidayNotification.objects.get_or_create(
            recipient=holiday.created_by,
            holiday=holiday,
            notif_type='approved',
            message=f"Your holiday request '{holiday.name}' ({holiday.date}) has been APPROVED by MD."
        )
    # Notify all active employees matching department
    employees = User.objects.all()
    if holiday.department and holiday.department != 'All':
        employees = employees.filter(department=holiday.department)
    
    for emp in employees:
        if emp != holiday.created_by:  # don't duplicate
            HolidayNotification.objects.create(
                recipient=emp,
                holiday=holiday,
                notif_type='approved',
                message=f"New Paid Holiday: '{holiday.name}' on {holiday.date} ({holiday.holiday_type})."
            )

def notify_holiday_rejected(holiday):
    if holiday.created_by:
        HolidayNotification.objects.create(
            recipient=holiday.created_by,
            holiday=holiday,
            notif_type='rejected',
            message=f"Your holiday request '{holiday.name}' ({holiday.date}) was REJECTED by MD. Remarks: {holiday.remarks or 'None'}"
        )

def notify_holiday_cancelled(holiday):
    # Notify HR creator
    if holiday.created_by:
        HolidayNotification.objects.create(
            recipient=holiday.created_by,
            holiday=holiday,
            notif_type='cancelled',
            message=f"Approved holiday '{holiday.name}' ({holiday.date}) has been CANCELLED."
        )
    # Notify all employees matching department
    employees = User.objects.all()
    if holiday.department and holiday.department != 'All':
        employees = employees.filter(department=holiday.department)
    
    for emp in employees:
        if emp != holiday.created_by:
            HolidayNotification.objects.create(
                recipient=emp,
                holiday=holiday,
                notif_type='cancelled',
                message=f"Holiday Announcement: The holiday '{holiday.name}' on {holiday.date} has been cancelled."
            )


@login_required
def hr_holiday_list(request):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
    
    status_filter = request.GET.get('status', 'All')
    year_filter = request.GET.get('year', '')
    
    holidays = Holiday.objects.all().order_by('date')
    if status_filter != 'All':
        holidays = holidays.filter(status=status_filter)
    if year_filter:
        holidays = holidays.filter(date__year=year_filter)
        
    # Get years for filter dropdown
    years = Holiday.objects.dates('date', 'year', order='DESC')
    years_list = [y.year for y in years]
    if timezone.now().year not in years_list:
        years_list.append(timezone.now().year)
    years_list = sorted(list(set(years_list)), reverse=True)
    
    # Calculate stats
    stats = {
        'all': Holiday.objects.count(),
        'draft': Holiday.objects.filter(status='Draft').count(),
        'pending': Holiday.objects.filter(status='Pending').count(),
        'approved': Holiday.objects.filter(status='Approved').count(),
        'rejected': Holiday.objects.filter(status='Rejected').count(),
        'cancelled': Holiday.objects.filter(status='Cancelled').count(),
    }
    
    return render(request, 'hr/holiday_list.html', {
        'holidays': holidays,
        'current_status': status_filter,
        'current_year': year_filter,
        'years_list': years_list,
        'stats': stats,
    })


@login_required
def hr_holiday_create(request):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    if request.method == 'POST':
        name = request.POST.get('name')
        date_str = request.POST.get('date')
        holiday_type = request.POST.get('holiday_type', 'Company')
        branch = request.POST.get('branch', 'All Branches')
        department = request.POST.get('department')
        description = request.POST.get('description', '')
        action = request.POST.get('action')  # 'Draft' or 'Submit'
        
        try:
            date_val = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Invalid date format.")
            return render(request, 'hr/holiday_create.html', {'departments': User.DEPARTMENT})
            
        status = 'Draft'
        if action == 'Submit':
            status = 'Pending'
            
        holiday = Holiday.objects.create(
            name=name,
            date=date_val,
            holiday_type=holiday_type,
            branch=branch,
            department=department,
            description=description,
            status=status,
            created_by=request.user,
            year=date_val.year
        )
        
        if status == 'Pending':
            holiday.submitted_by = request.user
            holiday.submitted_at = timezone.now()
            holiday.save()
            notify_md_holiday_pending(holiday)
            messages.success(request, f"Holiday '{name}' created and submitted for MD approval.")
        else:
            messages.success(request, f"Holiday '{name}' saved as Draft.")
            
        return redirect('hr_holiday_list')
        
    return render(request, 'hr/holiday_create.html', {
        'departments': User.DEPARTMENT
    })


@login_required
def hr_holiday_edit(request, pk):
    holiday = get_object_or_404(Holiday, id=pk)
    
    # Check permission
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    # Check if holiday is editable
    if not holiday.is_editable_by_hr and request.user.role == 'HR':
        messages.error(request, "Approved, Cancelled or Rejected holidays cannot be edited by HR.")
        return redirect('hr_holiday_list')
        
    if request.method == 'POST':
        name = request.POST.get('name')
        date_str = request.POST.get('date')
        holiday_type = request.POST.get('holiday_type')
        branch = request.POST.get('branch')
        department = request.POST.get('department')
        description = request.POST.get('description')
        action = request.POST.get('action') # 'Draft', 'Submit'
        
        try:
            date_val = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Invalid date format.")
            return redirect('hr_holiday_edit', pk=holiday.id)
            
        holiday.name = name
        holiday.date = date_val
        holiday.holiday_type = holiday_type
        holiday.branch = branch
        holiday.department = department
        holiday.description = description
        holiday.year = date_val.year
        holiday.last_modified_by = request.user
        
        if action == 'Submit':
            holiday.status = 'Pending'
            holiday.submitted_by = request.user
            holiday.submitted_at = timezone.now()
            notify_md_holiday_pending(holiday)
            messages.success(request, f"Holiday '{name}' updated and submitted for approval.")
        elif action == 'Draft':
            holiday.status = 'Draft'
            messages.success(request, f"Holiday '{name}' updated and saved as Draft.")
        else:
            messages.success(request, f"Holiday '{name}' updated successfully.")
            
        holiday.save()
        return redirect('hr_holiday_list')
        
    return render(request, 'hr/holiday_edit.html', {
        'holiday': holiday,
        'departments': User.DEPARTMENT
    })


@login_required
def hr_holiday_delete(request, pk):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    holiday = get_object_or_404(Holiday, id=pk)
    if holiday.status not in ('Draft', 'Pending'):
        messages.error(request, "Only Draft or Pending holidays can be deleted.")
        return redirect('hr_holiday_list')
        
    holiday.delete()
    messages.success(request, "Holiday deleted successfully.")
    return redirect('hr_holiday_list')


@login_required
def hr_holiday_submit(request, pk):
    if request.user.role != 'HR':
        messages.error(request, "Only HR can submit holidays.")
        return redirect('hr_holiday_list')
        
    holiday = get_object_or_404(Holiday, id=pk)
    if holiday.status != 'Draft':
        messages.error(request, "Only Draft holidays can be submitted.")
        return redirect('hr_holiday_list')
        
    holiday.status = 'Pending'
    holiday.submitted_by = request.user
    holiday.submitted_at = timezone.now()
    holiday.save()
    
    notify_md_holiday_pending(holiday)
    messages.success(request, f"Holiday '{holiday.name}' submitted for MD approval.")
    return redirect('hr_holiday_list')


@login_required
def hr_holiday_cancel(request, pk):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    holiday = get_object_or_404(Holiday, id=pk)
    if holiday.status != 'Approved':
        messages.error(request, "Only Approved holidays can be cancelled.")
        return redirect('hr_holiday_list')
        
    holiday.status = 'Cancelled'
    holiday.cancelled_by = request.user
    holiday.cancelled_at = timezone.now()
    if request.method == 'POST':
        holiday.remarks = request.POST.get('remarks')
    holiday.save()
    
    notify_holiday_cancelled(holiday)
    messages.success(request, f"Holiday '{holiday.name}' has been cancelled.")
    return redirect('hr_holiday_list')


@login_required
def md_holiday_approvals(request):
    if request.user.role != 'MD':
        messages.error(request, "Access denied.")
        return redirect('home')
        
    pending_holidays = Holiday.objects.filter(status='Pending').order_by('date')
    return render(request, 'md/holiday_approvals.html', {
        'pending_holidays': pending_holidays
    })


@login_required
def md_holiday_approve(request, pk):
    if request.user.role != 'MD':
        messages.error(request, "Access denied.")
        return redirect('home')
        
    holiday = get_object_or_404(Holiday, id=pk)
    if holiday.status != 'Pending':
        messages.error(request, "Holiday is not pending approval.")
        return redirect('md_holiday_approvals')
        
    holiday.status = 'Approved'
    holiday.approved_by = request.user
    holiday.approved_at = timezone.now()
    if request.method == 'POST':
        holiday.remarks = request.POST.get('remarks')
    holiday.save()
    
    notify_holiday_approved(holiday)
    messages.success(request, f"Holiday '{holiday.name}' approved successfully.")
    return redirect('md_holiday_approvals')


@login_required
def md_holiday_reject(request, pk):
    if request.user.role != 'MD':
        messages.error(request, "Access denied.")
        return redirect('home')
        
    holiday = get_object_or_404(Holiday, id=pk)
    if holiday.status != 'Pending':
        messages.error(request, "Holiday is not pending approval.")
        return redirect('md_holiday_approvals')
        
    holiday.status = 'Rejected'
    holiday.rejected_by = request.user
    holiday.rejected_at = timezone.now()
    if request.method == 'POST':
        holiday.remarks = request.POST.get('remarks')
    holiday.save()
    
    notify_holiday_rejected(holiday)
    messages.success(request, f"Holiday '{holiday.name}' was rejected.")
    return redirect('md_holiday_approvals')


@login_required
def hr_holiday_import_excel(request):
    if request.user.role != 'HR':
        messages.error(request, "Access denied.")
        return redirect('home')
        
    if request.method == 'POST' and request.FILES.get('file'):
        file_obj = request.FILES['file']
        filename = file_obj.name.lower()
        
        imported_count = 0
        errors = []
        import csv
        import datetime
        
        if filename.endswith('.xlsx'):
            try:
                import openpyxl
            except ImportError:
                messages.error(request, "openpyxl library not installed. Please contact administrator.")
                return redirect('hr_holiday_list')
                
            try:
                wb = openpyxl.load_workbook(file_obj, read_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(cell).strip().lower() for cell in rows[0] if cell is not None]
                    col_map = {
                        'name': -1, 'date': -1, 'type': -1,
                        'branch': -1, 'department': -1, 'description': -1
                    }
                    for idx, h in enumerate(headers):
                        for key in col_map.keys():
                            if key in h:
                                col_map[key] = idx
                                
                    if col_map['name'] == -1 or col_map['date'] == -1:
                        messages.error(request, "Invalid Excel structure. Name and Date columns are required.")
                        return redirect('hr_holiday_list')
                        
                    for row_idx, row in enumerate(rows[1:], start=2):
                        if len(row) <= max(col_map['name'], col_map['date']):
                            continue
                        name = row[col_map['name']]
                        date_val = row[col_map['date']]
                        
                        if not name or date_val is None:
                            continue
                        
                        parsed_date = None
                        if isinstance(date_val, datetime.date):
                            parsed_date = date_val
                        elif isinstance(date_val, datetime.datetime):
                            parsed_date = date_val.date()
                        else:
                            date_str = str(date_val).strip()
                            for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
                                try:
                                    parsed_date = datetime.datetime.strptime(date_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                                    
                        if not parsed_date:
                            errors.append(f"Row {row_idx}: Invalid date value '{date_val}'")
                            continue
                            
                        h_type = 'Company'
                        if col_map['type'] != -1 and row[col_map['type']]:
                            raw_type = str(row[col_map['type']]).strip()
                            for t_choice, _ in Holiday.HOLIDAY_TYPE_CHOICES:
                                if t_choice.lower() in raw_type.lower():
                                    h_type = t_choice
                                    break
                                    
                        branch = 'All Branches'
                        if col_map['branch'] != -1 and row[col_map['branch']]:
                            branch = str(row[col_map['branch']]).strip()
                            
                        dept = None
                        if col_map['department'] != -1 and row[col_map['department']]:
                            raw_dept = str(row[col_map['department']]).strip()
                            for dept_choice, dept_label in User.DEPARTMENT:
                                if raw_dept.lower() in (dept_choice.lower(), dept_label.lower()):
                                    dept = dept_choice
                                    break
                                    
                        desc = ''
                        if col_map['description'] != -1 and row[col_map['description']]:
                            desc = str(row[col_map['description']]).strip()
                            
                        Holiday.objects.create(
                            name=str(name).strip(),
                            date=parsed_date,
                            holiday_type=h_type,
                            branch=branch,
                            department=dept,
                            description=desc,
                            status='Draft',
                            created_by=request.user,
                            year=parsed_date.year
                        )
                        imported_count += 1
            except Exception as e:
                messages.error(request, f"Error processing Excel file: {str(e)}")
                return redirect('hr_holiday_list')
                
        elif filename.endswith('.csv'):
            try:
                decoded_file = file_obj.read().decode('utf-8-sig').splitlines()
                reader = csv.reader(decoded_file)
                rows = list(reader)
                if len(rows) > 1:
                    headers = [str(h).strip().lower() for h in rows[0]]
                    col_map = {
                        'name': -1, 'date': -1, 'type': -1,
                        'branch': -1, 'department': -1, 'description': -1
                    }
                    for idx, h in enumerate(headers):
                        for key in col_map.keys():
                            if key in h:
                                col_map[key] = idx
                                
                    if col_map['name'] == -1 or col_map['date'] == -1:
                        messages.error(request, "Invalid CSV structure. Name and Date columns are required.")
                        return redirect('hr_holiday_list')
                        
                    for row_idx, row in enumerate(rows[1:], start=2):
                        if len(row) <= max(col_map['name'], col_map['date']):
                            continue
                        name = row[col_map['name']].strip()
                        date_str = row[col_map['date']].strip()
                        
                        if not name or not date_str:
                            continue
                            
                        parsed_date = None
                        for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
                            try:
                                parsed_date = datetime.datetime.strptime(date_str, fmt).date()
                                break
                            except ValueError:
                                continue
                                
                        if not parsed_date:
                            errors.append(f"Row {row_idx}: Invalid date value '{date_str}'")
                            continue
                            
                        h_type = 'Company'
                        if col_map['type'] != -1 and len(row) > col_map['type'] and row[col_map['type']]:
                            raw_type = row[col_map['type']].strip()
                            for t_choice, _ in Holiday.HOLIDAY_TYPE_CHOICES:
                                if t_choice.lower() in raw_type.lower():
                                    h_type = t_choice
                                    break
                                    
                        branch = 'All Branches'
                        if col_map['branch'] != -1 and len(row) > col_map['branch'] and row[col_map['branch']]:
                            branch = row[col_map['branch']].strip()
                            
                        dept = None
                        if col_map['department'] != -1 and len(row) > col_map['department'] and row[col_map['department']]:
                            raw_dept = row[col_map['department']].strip()
                            for dept_choice, dept_label in User.DEPARTMENT:
                                if raw_dept.lower() in (dept_choice.lower(), dept_label.lower()):
                                    dept = dept_choice
                                    break
                                    
                        desc = ''
                        if col_map['description'] != -1 and len(row) > col_map['description'] and row[col_map['description']]:
                            desc = row[col_map['description']].strip()
                            
                        Holiday.objects.create(
                            name=name,
                            date=parsed_date,
                            holiday_type=h_type,
                            branch=branch,
                            department=dept,
                            description=desc,
                            status='Draft',
                            created_by=request.user,
                            year=parsed_date.year
                        )
                        imported_count += 1
            except Exception as e:
                messages.error(request, f"Error processing CSV file: {str(e)}")
                return redirect('hr_holiday_list')
        else:
            messages.error(request, "Unsupported file format. Please upload an Excel (.xlsx) or CSV (.csv) file.")
            return redirect('hr_holiday_list')
            
        if imported_count > 0:
            msg = f"Successfully imported {imported_count} holidays as Drafts."
            if errors:
                msg += f" However, there were some errors: {', '.join(errors[:5])}"
                if len(errors) > 5:
                    msg += "..."
                messages.warning(request, msg)
            else:
                messages.success(request, msg)
        elif errors:
            messages.error(request, f"Import failed with following errors: {'; '.join(errors)}")
        else:
            messages.error(request, "No valid holiday records found in the uploaded file.")
            
    return redirect('hr_holiday_list')


@login_required
def employee_holiday_calendar(request):
    year_filter = request.GET.get('year', '')
    if not year_filter:
        year_filter = str(timezone.now().year)
        
    from django.db.models import Q
    holidays = Holiday.objects.filter(
        status='Approved',
        date__year=year_filter
    ).filter(
        Q(department__isnull=True) | Q(department='') | Q(department=request.user.department)
    ).order_by('date')
    
    import calendar as cal_lib
    months_holidays = {i: [] for i in range(1, 13)}
    for h in holidays:
        months_holidays[h.date.month].append(h)
        
    months_data = []
    for m in range(1, 13):
        months_data.append({
            'num': m,
            'name': cal_lib.month_name[m],
            'holidays': months_holidays[m]
        })
        
    years = Holiday.objects.filter(status='Approved').dates('date', 'year', order='DESC')
    years_list = [y.year for y in years]
    if int(year_filter) not in years_list:
        years_list.append(int(year_filter))
    years_list = sorted(list(set(years_list)), reverse=True)
    
    return render(request, 'employee/holiday_calendar.html', {
        'months_data': months_data,
        'current_year': year_filter,
        'years_list': years_list,
    })


from django.http import JsonResponse
@login_required
def holiday_notifications_api(request):
    notifs = HolidayNotification.objects.filter(recipient=request.user, is_read=False)
    
    if request.method == 'POST':
        notifs.update(is_read=True)
        return JsonResponse({'status': 'success', 'unread_count': 0})
        
    notifs_list = []
    for n in notifs[:10]:
        notifs_list.append({
            'id': n.id,
            'message': n.message,
            'notif_type': n.notif_type,
            'created_at': n.created_at.strftime('%Y-%m-%d %H:%M'),
        })
        
    return JsonResponse({
        'unread_count': notifs.count(),
        'notifications': notifs_list
    })


# ==================== HR SETTINGS VIEW ====================

@login_required
def hr_settings_view(request):
    """Company-level HR settings: toggle Sandwich Leave Policy."""
    user = request.user
    if user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')

    from .models import HRSettings
    settings = HRSettings.get_settings()

    if request.method == 'POST':
        settings.sandwich_leave_enabled = request.POST.get('sandwich_leave_enabled') == 'on'
        settings.save()
        status = "enabled" if settings.sandwich_leave_enabled else "disabled"
        messages.success(request, f"Sandwich Leave Policy has been {status}.")
        return redirect('hr_settings')

    return render(request, 'hr/hr_settings.html', {
        'settings': settings,
    })


@login_required
def hr_finalize_attendance(request):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    if request.method == 'POST':
        month = int(request.POST.get('month'))
        year = int(request.POST.get('year'))
        
        from .models import AttendanceFinalization
        finalization, created = AttendanceFinalization.objects.get_or_create(
            month=month,
            year=year,
            defaults={'finalized_by': request.user, 'is_finalized': True}
        )
        if not created:
            finalization.is_finalized = True
            finalization.finalized_by = request.user
            finalization.save()
            
        messages.success(request, f"Attendance for {calendar.month_name[month]} {year} has been finalized successfully.")
    return redirect(request.META.get('HTTP_REFERER', 'hr_payslip_list'))


@login_required
def hr_unfinalize_attendance(request):
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, "Access denied.")
        return redirect('home')
        
    if request.method == 'POST':
        month = int(request.POST.get('month'))
        year = int(request.POST.get('year'))
        
        from .models import AttendanceFinalization
        AttendanceFinalization.objects.filter(month=month, year=year).delete()
        messages.success(request, f"Attendance for {calendar.month_name[month]} {year} has been unfinalized.")
    return redirect(request.META.get('HTTP_REFERER', 'hr_payslip_list'))


@login_required
def hr_payroll_submit_approval(request):
    if request.user.role != 'HR':
        messages.error(request, "Access denied. HR only.")
        return redirect('home')
    if request.method == 'POST':
        month = request.POST.get('month')
        year = request.POST.get('year')
        if not month or not year:
            messages.error(request, "Month and Year are required.")
            return redirect('hr_payslip_list')
        
        payslips = Payslip.objects.filter(month=month, year=year, status__in=['Draft', 'Generated'])
        if not payslips.exists():
            messages.error(request, "No generated/draft payslips found to submit for approval.")
            return redirect('hr_payslip_list')
            
        payslips.update(status='Pending Approval')
        
        # Log audit
        ip = request.META.get('REMOTE_ADDR')
        for ps in payslips:
            PayrollAuditLog.objects.create(
                payslip=ps,
                performed_by=request.user,
                action='LOCK',
                ip_address=ip,
                details=f"Submitted payslip to MD for approval for {ps.month_name} {ps.year}"
            )
        
        messages.success(request, f"Payroll for {calendar.month_name[int(month)]} {year} submitted to MD for approval.")
    return redirect('hr_payslip_list')


@login_required
def md_payroll_approve(request):
    if request.user.role != 'MD':
        messages.error(request, "Access denied. MD only.")
        return redirect('home')
    if request.method == 'POST':
        month = request.POST.get('month')
        year = request.POST.get('year')
        if not month or not year:
            messages.error(request, "Month and Year are required.")
            return redirect('hr_payslip_list')
            
        payslips = Payslip.objects.filter(month=month, year=year, status='Pending Approval')
        if not payslips.exists():
            messages.error(request, "No payroll records found pending approval for this month.")
            return redirect('hr_payslip_list')
            
        # Update each to Approved
        for ps in payslips:
            ps.status = 'Approved'
            ps.approved_by = request.user
            ps.approved_at = timezone.now()
            ps.save()
            
            # Log audit
            ip = request.META.get('REMOTE_ADDR')
            PayrollAuditLog.objects.create(
                payslip=ps,
                performed_by=request.user,
                action='LOCK',
                ip_address=ip,
                details=f"Approved payroll for {ps.month_name} {ps.year}"
            )
            
        messages.success(request, f"Payroll for {calendar.month_name[int(month)]} {year} approved successfully.")
    return redirect('hr_payslip_list')


@login_required
def md_payroll_release(request):
    if request.user.role != 'MD':
        messages.error(request, "Access denied. MD only.")
        return redirect('home')
    if request.method == 'POST':
        month = request.POST.get('month')
        year = request.POST.get('year')
        if not month or not year:
            messages.error(request, "Month and Year are required.")
            return redirect('hr_payslip_list')
            
        payslips = Payslip.objects.filter(month=month, year=year, status='Approved')
        if not payslips.exists():
            messages.error(request, "No approved payroll records found to release.")
            return redirect('hr_payslip_list')
            
        # Update each to Paid, publish, generate PDF
        for ps in payslips:
            ps.status = 'Paid'
            ps.payment_date = date.today()
            ps.is_published = True
            ps.save()
            
            generate_payslip_pdf_file(ps)
            ps.save()
            
            # Log audit
            ip = request.META.get('REMOTE_ADDR')
            PayrollAuditLog.objects.create(
                payslip=ps,
                performed_by=request.user,
                action='PAID',
                ip_address=ip,
                details=f"Released payment & published payslip for {ps.month_name} {ps.year}"
            )
            
        messages.success(request, f"Payroll for {calendar.month_name[int(month)]} {year} released and published successfully.")
    return redirect('hr_payslip_list')


# ─────────────────────────────────────────────────────────────────────────────
#  UNIVERSAL "MY PROFILE" – TL, Manager, MD (self-view, same template as employee)
# ─────────────────────────────────────────────────────────────────────────────

# Mapping role → base template for layouts that are sidebar-based
ROLE_BASE_TEMPLATE = {
    'Employee': 'employee/employee_base.html',
    'TeamLead': 'tl/side_bar.html',
    'Manager': 'main/mddash.html',
    'HR': 'base.html',
    'MD': 'md/md_base.html',
}


@login_required
def my_profile(request):
    """Universal My Profile page for TL, Manager, MD, HR."""
    user = request.user
    allowed = ('TeamLead', 'Manager', 'MD', 'HR')
    if user.role not in allowed:
        return redirect('employee_profile')  # Employees already have their own view

    from .models import Team, Project, Leave
    teams = Team.objects.filter(members=user)
    projects = Project.objects.filter(assigned_team__in=teams)

    month_param = request.GET.get('month')
    if month_param:
        yr, mo = month_param.split('-')
        attendance = calculate_attendance(user, int(yr), int(mo))
    else:
        now = timezone.now()
        attendance = calculate_attendance(user, now.year, now.month)

    leave_summary = None
    try:
        now_year = timezone.now().year
        emp_leaves = Leave.objects.filter(user=user, created_at__year=now_year)
        leave_summary = {
            'approved': emp_leaves.filter(status__icontains='Approved').count(),
            'pending':  emp_leaves.filter(status__icontains='Pending').count(),
            'rejected': emp_leaves.filter(status__icontains='Rejected').count(),
        }
    except Exception:
        pass

    base_template = ROLE_BASE_TEMPLATE.get(user.role, 'employee/employee_base.html')

    return render(request, 'shared/my_profile.html', {
        'employee': user,
        'employee_projects': projects,
        'attendance': attendance,
        'reporting_manager': user.reporting_manager,
        'leave_summary': leave_summary,
        'base_template': base_template,
    })


@login_required
def my_profile_edit(request):
    """Universal profile edit for TL, Manager, MD, HR."""
    user = request.user
    allowed = ('TeamLead', 'Manager', 'MD', 'HR')
    if user.role not in allowed:
        return redirect('edit_employee_self')

    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name  = request.POST.get('last_name', user.last_name)
        user.phone      = request.POST.get('phone', user.phone)
        user.email      = request.POST.get('email', user.email)

        if request.POST.get('remove_profile_pic') == 'true':
            if user.profile_pic:
                user.profile_pic.delete(save=False)
                user.profile_pic = None
        elif request.FILES.get('profile_pic'):
            user.profile_pic = request.FILES['profile_pic']

        new_password = request.POST.get('new_password')
        if new_password:
            user.set_password(new_password)
            user.save()
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, user)
        else:
            user.save()

        messages.success(request, 'Profile updated successfully.')
        return redirect('my_profile')

    base_template = ROLE_BASE_TEMPLATE.get(user.role, 'employee/employee_base.html')
    return render(request, 'shared/my_profile_edit.html', {
        'employee': user,
        'base_template': base_template,
    })


# ─────────────────────────────────────────────────────────────────────────────
#  UNIVERSAL "MY ATTENDANCE" – TL, MD (personal attendance calendar)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def my_attendance(request):
    """
    Personal attendance page for TL, Manager, MD, HR.
    Delegates to admin_attendance_view logic with user_id=me.
    """
    user = request.user
    allowed = ('TeamLead', 'Manager', 'MD', 'HR')
    if user.role not in allowed:
        return redirect('attendance_list')

    today = date.today()
    selected_month = int(request.GET.get('month', today.month))
    selected_year  = int(request.GET.get('year',  today.year))

    days_data, padding, stats = get_monthly_calendar_data(user, selected_year, selected_month)
    today_record = Attendance.objects.filter(user=user, date=today).first()

    # History
    history_records = Attendance.objects.filter(user=user)
    q_month = request.GET.get('q_month')
    q_year  = request.GET.get('q_year')
    if q_month:
        history_records = history_records.filter(date__month=q_month)
    if q_year:
        history_records = history_records.filter(date__year=q_year)
    history_records = history_records.order_by('-date')

    # Yearly summary
    yearly_stats = {}
    for yr in [today.year - 1, today.year]:
        yr_qs = Attendance.objects.filter(user=user, date__year=yr)
        yearly_stats[yr] = {
            'present':  yr_qs.filter(status__icontains='present').count(),
            'absent':   yr_qs.filter(status__icontains='absent').count(),
            'half':     yr_qs.filter(status__icontains='half').count(),
            'late':     yr_qs.filter(is_late=True).count(),
        }

    base_template = ROLE_BASE_TEMPLATE.get(user.role, 'employee/employee_base.html')

    context = {
        'target_user': user,
        'days_data': days_data,
        'padding_range': range(padding),
        'stats': stats,
        'today_record': today_record,
        'history_records': history_records,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'base_template': base_template,
        'q_month': q_month,
        'q_year': q_year,
        'yearly_stats': yearly_stats,
        'is_my_attendance': True,
    }
    return render(request, 'attedance/attendance_employee.html', context)


# ─────────────────────────────────────────────────────────────────────────────
#  UNIVERSAL "MY PAYSLIPS" – TL, Manager, MD, HR
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def my_payslips_universal(request):
    """Personal payslips for TL, Manager, MD, HR (mirrors my_payslips but for all roles)."""
    user = request.user
    allowed = ('TeamLead', 'Manager', 'MD', 'HR')
    if user.role not in allowed:
        return redirect('my_payslips')

    payslips = Payslip.objects.filter(employee=user, is_published=True).order_by('-year', '-month')

    filter_month = request.GET.get('month')
    filter_year  = request.GET.get('year')
    filter_fy    = request.GET.get('fy')

    if filter_month:
        payslips = payslips.filter(month=filter_month)
    if filter_year:
        payslips = payslips.filter(year=filter_year)
    if filter_fy:
        try:
            fy_start = int(filter_fy)
            payslips = payslips.filter(
                models.Q(year=fy_start, month__gte=4) |
                models.Q(year=fy_start + 1, month__lte=3)
            )
        except (ValueError, TypeError):
            pass

    latest      = payslips.first()
    total_count = payslips.count()
    years       = Payslip.objects.filter(employee=user, is_published=True).values_list('year', flat=True).distinct().order_by('-year')
    current_year = datetime.now().year
    fy_options  = [(y, f"FY {y}-{str(y+1)[2:]}") for y in range(current_year, current_year - 5, -1)]

    base_template = ROLE_BASE_TEMPLATE.get(user.role, 'employee/employee_base.html')

    context = {
        'payslips':      payslips,
        'latest':        latest,
        'total_count':   total_count,
        'years':         years,
        'fy_options':    fy_options,
        'filter_month':  filter_month,
        'filter_year':   filter_year,
        'filter_fy':     filter_fy,
        'month_choices': list(calendar.month_name)[1:],
        'base_template': base_template,
    }
    return render(request, 'shared/my_payslips.html', context)


# ─────────────────────────────────────────────────────────────────────────────
#  HR ATTENDANCE CORRECTION  (single + bulk corrections with MD approval flow)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def hr_attendance_correct(request):
    """Single-day attendance correction form – creates a pending AttendanceCorrection."""
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, 'Access denied.')
        return redirect('home')

    from .models import AttendanceCorrection

    STATUS_CHOICES = [
        'Present', 'Absent', 'Half Day', 'Paid Leave', 'Sick Leave',
        'Casual Leave', 'Earned Leave', 'Work From Home', 'Holiday', 'Weekly Off',
    ]

    if request.method == 'POST':
        att_id         = request.POST.get('attendance_id')
        new_status     = request.POST.get('new_status', '').strip()
        new_check_in   = request.POST.get('new_check_in') or None
        new_check_out  = request.POST.get('new_check_out') or None
        new_hours      = request.POST.get('new_total_hours') or None
        new_remarks    = request.POST.get('new_remarks', '')
        reason         = request.POST.get('reason', '').strip()

        if not att_id or not reason:
            messages.error(request, 'Attendance record and reason are required.')
            return redirect(request.path)

        try:
            att = Attendance.objects.get(pk=att_id)
        except Attendance.DoesNotExist:
            messages.error(request, 'Attendance record not found.')
            return redirect(request.path)

        # Parse datetime strings if provided
        from django.utils.dateparse import parse_datetime
        parsed_in  = parse_datetime(new_check_in)  if new_check_in  else None
        parsed_out = parse_datetime(new_check_out) if new_check_out else None

        # Snapshot original
        AttendanceCorrection.objects.create(
            attendance          = att,
            original_check_in   = att.check_in_time,
            original_check_out  = att.check_out_time,
            original_status     = att.status,
            original_total_hours= att.total_hours,
            original_remarks    = att.remarks,
            new_check_in        = parsed_in,
            new_check_out       = parsed_out,
            new_status          = new_status,
            new_total_hours     = float(new_hours) if new_hours else att.total_hours,
            new_remarks         = new_remarks,
            reason              = reason,
            edited_by           = request.user,
            status              = 'Pending',
        )

        # Mark attendance as pending (visible signal to MD)
        att.status = 'Pending Approval'
        att.save()

        messages.success(request, f'Correction submitted for {att.user.get_full_name()} on {att.date}. Awaiting MD approval.')
        return redirect('attendance_list')

    # GET  – show form with optional pre-filled employee/date
    emp_id    = request.GET.get('emp_id')
    date_str  = request.GET.get('date')
    att       = None
    employees = User.objects.filter(role__in=['Employee', 'TeamLead', 'Manager'], is_active=True).order_by('first_name')

    if emp_id and date_str:
        att = Attendance.objects.filter(user_id=emp_id, date=date_str).first()

    return render(request, 'hr/hr_attendance_correction.html', {
        'att':            att,
        'employees':      employees,
        'status_choices': STATUS_CHOICES,
        'prefill_emp':    emp_id,
        'prefill_date':   date_str,
    })


@login_required
def hr_attendance_correct_bulk(request):
    """Bulk attendance correction – one POST creates multiple AttendanceCorrection records."""
    if request.user.role not in ('HR', 'MD'):
        messages.error(request, 'Access denied.')
        return redirect('home')

    from .models import AttendanceCorrection
    from django.utils.dateparse import parse_date

    STATUS_CHOICES = [
        'Present', 'Absent', 'Half Day', 'Paid Leave', 'Sick Leave',
        'Casual Leave', 'Earned Leave', 'Work From Home', 'Holiday', 'Weekly Off',
    ]

    if request.method == 'POST':
        emp_id     = request.POST.get('emp_id')
        from_date  = parse_date(request.POST.get('from_date', ''))
        to_date    = parse_date(request.POST.get('to_date', ''))
        new_status = request.POST.get('new_status', '').strip()
        new_remarks= request.POST.get('new_remarks', '')
        reason     = request.POST.get('reason', '').strip()

        if not (emp_id and from_date and to_date and new_status and reason):
            messages.error(request, 'All fields are required.')
            return redirect(request.path)

        try:
            emp = User.objects.get(pk=emp_id)
        except User.DoesNotExist:
            messages.error(request, 'Employee not found.')
            return redirect(request.path)

        if from_date > to_date:
            messages.error(request, 'From date must be before To date.')
            return redirect(request.path)

        records = Attendance.objects.filter(user=emp, date__range=(from_date, to_date))
        count = 0
        for att in records:
            AttendanceCorrection.objects.create(
                attendance           = att,
                original_check_in    = att.check_in_time,
                original_check_out   = att.check_out_time,
                original_status      = att.status,
                original_total_hours = att.total_hours,
                original_remarks     = att.remarks,
                new_check_in         = att.check_in_time,
                new_check_out        = att.check_out_time,
                new_status           = new_status,
                new_total_hours      = att.total_hours,
                new_remarks          = new_remarks or att.remarks,
                reason               = reason,
                edited_by            = request.user,
                status               = 'Pending',
            )
            att.status = 'Pending Approval'
            att.save()
            count += 1

        messages.success(request, f'{count} attendance record(s) submitted for correction. Awaiting MD approval.')
        return redirect('attendance_list')

    employees = User.objects.filter(role__in=['Employee', 'TeamLead', 'Manager'], is_active=True).order_by('first_name')
    return render(request, 'hr/hr_attendance_correction_bulk.html', {
        'employees':      employees,
        'status_choices': STATUS_CHOICES,
    })


# ─────────────────────────────────────────────────────────────────────────────
#  MD ATTENDANCE APPROVAL
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def md_attendance_approvals_list(request):
    """MD view: list of all pending attendance correction requests."""
    if request.user.role != 'MD':
        messages.error(request, 'Access denied. MD only.')
        return redirect('home')

    from .models import AttendanceCorrection

    status_filter = request.GET.get('status', 'Pending')
    corrections = AttendanceCorrection.objects.select_related(
        'attendance__user', 'edited_by', 'approved_by'
    ).order_by('-created_at')

    if status_filter:
        corrections = corrections.filter(status=status_filter)

    return render(request, 'md/md_attendance_approvals.html', {
        'corrections':    corrections,
        'status_filter':  status_filter,
        'pending_count':  AttendanceCorrection.objects.filter(status='Pending').count(),
    })


@login_required
def md_attendance_approval_action(request, pk):
    """MD approves or rejects a single attendance correction."""
    if request.user.role != 'MD':
        messages.error(request, 'Access denied. MD only.')
        return redirect('home')

    from .models import AttendanceCorrection

    correction = get_object_or_404(AttendanceCorrection, pk=pk)

    if request.method == 'POST':
        action     = request.POST.get('action')  # 'approve' or 'reject'
        md_remarks = request.POST.get('md_remarks', '')

        att = correction.attendance

        if action == 'approve':
            # Apply proposed changes to the Attendance record
            if correction.new_check_in:
                att.check_in_time  = correction.new_check_in
            if correction.new_check_out:
                att.check_out_time = correction.new_check_out
            if correction.new_status:
                att.status = correction.new_status
            if correction.new_total_hours is not None:
                att.total_hours = correction.new_total_hours
            if correction.new_remarks:
                att.remarks = correction.new_remarks
            att.save()

            correction.status      = 'Approved'
            correction.approved_by = request.user
            correction.approved_at = timezone.now()
            correction.md_remarks  = md_remarks
            correction.save()

            messages.success(request, f'Correction approved for {att.user.get_full_name()} on {att.date}.')

        elif action == 'reject':
            # Restore original status on Attendance
            att.status = correction.original_status
            att.save()

            correction.status      = 'Rejected'
            correction.approved_by = request.user
            correction.approved_at = timezone.now()
            correction.md_remarks  = md_remarks
            correction.save()

            messages.success(request, f'Correction rejected for {att.user.get_full_name()} on {att.date}.')
        else:
            messages.error(request, 'Invalid action.')

    return redirect('md_attendance_approvals_list')

